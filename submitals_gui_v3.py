#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
================================================================================
 submitals_gui_v3.py  --  Interfaz principal v3.1.0 (Generador de Submittals ES)
================================================================================
Interfaz de la version 3, basada en una **Base de Datos centralizada en GitHub**
de fichas reutilizables. Coexiste con el sistema v2.6 (generacion desde carpetas)
sin modificarlo.

Menu principal (2x2):
   [ Generar desde BD ]        [ Abrir submittal existente ]
   [ Cargar ficha a BD ]       [ Gestionar BD ]

Este archivo contiene DOS partes:
  A) Funciones de ORQUESTACION de entregables (sin tkinter): reutilizan el motor
     de carATulas v2.6 (``generate_caratulas.py``) y replican la compilacion de
     CMPs/Excel con ``pypdf``/``openpyxl``. Son importables y testeables solas.
  B) La GUI (tkinter), que solo se carga si tkinter esta disponible.

Sincronizacion (v3.1.0, reemplaza a OneDrive):
  - Al abrir: ``sync_indice()`` (git pull) en segundo plano, con barra de estado.
  - Al cargar fichas o generar un submittal: ``git_push()``.
  - Los conflictos se resuelven solos (fusion por registro) y se avisa al usuario.
  - Sin conexion: se trabaja con la copia local y se sube al reconectar.
  - YA NO EXISTE el ``.lock`` ni el dialogo de "forzar acceso": git se encarga.
================================================================================
"""

import os
import re
import sys
import json
import shutil
import logging
import threading
import importlib.util
from pathlib import Path

import bd_manager
import fuzzy_search
import git_bd
import nomenclatura
import ocr_extractor
import updater_gh

VERSION = "3.3.8"
BASE_DIR = Path(__file__).resolve().parent
PIN_MODO_DEV = "9119"

# ==========================================================================
# SISTEMA DE DISENO v3.4.0
# --------------------------------------------------------------------------
# Refresh visual: neutros CALIDOS + un unico acento rojo ES, sidebar oscuro
# permanente y tipografia IBM Plex (Sans para texto, Mono para todo dato
# tabular: consecutivos, conteos, rutas, horas). Reemplaza a la paleta
# azul/naranja de v3.3.x.
#
# Los valores vienen del diseno en oklch, convertidos a sRGB (Tkinter solo
# entiende hex). Contraste texto/fondo >= 4.5:1 en todos los pares en uso
# (ej. blanco sobre ACENTO = 5.35:1).
# ==========================================================================
# --- superficies
FONDO = "#F4F1F0"            # fondo de dialogos
WINDOW_BG = "#FCF9F9"        # fondo del area de contenido de la ventana
SUPERFICIE = "#FFFFFF"       # tarjetas, tablas, barras
SUPERFICIE_2 = "#FBFAF9"     # encabezados de tarjeta, barras de herramientas
SUPERFICIE_3 = "#F7F4F3"     # cabeceras de tabla
# --- sidebar (modo oscuro fijo, independiente del tema claro del resto)
SIDEBAR = "#221D1C"
SIDEBAR_HOVER = "#302A28"
SIDEBAR_CARD = "#2F2827"
SIDEBAR_BTN = "#433B39"
SIDEBAR_BTN_H = "#554A48"
SIDEBAR_BORDE = "#524B4A"
SIDEBAR_TXT = "#F7F4F3"      # titulo / valores
SIDEBAR_TXT_2 = "#B6AFAD"    # items de navegacion en reposo
SIDEBAR_TXT_3 = "#918B88"    # metadatos (version, "hace un momento")
SIDEBAR_LABEL = "#7A7370"    # encabezados de seccion (MENU, CATALOGO)
SIDEBAR_NUM = "#E8E3E2"      # conteos por categoria
# --- acento (unico: rojo ES)
ACENTO = "#CC2827"
ACENTO_HOVER = "#B00C15"
ACENTO_TXT = "#8D0005"       # texto de acento sobre fondo claro
ACENTO_SUAVE = "#FFF3F0"     # fondo de tarjeta destacada
ACENTO_SUAVE_H = "#FFE9E5"
ACENTO_BORDE = "#EF958B"
ACENTO_BORDE_FUERTE = "#D9544B"   # borde de la tarjeta destacada en hover
CHIP_BG = "#F8E6E4"          # pills de filtro y fila seleccionada
FILA_HOVER = "#FFF4F2"
# --- texto
TEXTO = "#211B19"
TEXTO_2 = "#3F3937"
TEXTO_SUAVE = "#6D6765"
TEXTO_TENUE = "#8B8583"
# --- bordes
BORDE = "#DED9D8"
BORDE_FUERTE = "#D1CCCB"
BORDE_TENUE = "#EEEAE9"
# --- estado
VERDE = "#007136"
VERDE_PUNTO = "#46B86E"
AMBAR = "#8E3800"
ROJO = "#B32322"
ROJO_BORDE = "#F9AEA5"
ROJO_SUAVE = "#FFECE9"
ROJO_CLARO = "#F0918A"       # rojo legible SOBRE el sidebar oscuro
# --- consola / registro (unico bloque oscuro del area de contenido)
LOG_BG = "#16181E"
LOG_TXT = "#CACED4"
LOG_OK = "#98D5A8"
LOG_WARN = "#D5B36A"
LOG_LABEL = "#7A808D"

# Alias de la paleta v3.3.x: el resto del archivo (dialogos que no cambian de
# estructura) los sigue usando y asi hereda la paleta nueva sin tocarse.
AZUL_ES = ACENTO             # accion primaria
AZUL_CLARO = ACENTO_HOVER    # acentos/hover
NARANJA_CTA = ACENTO         # el diseno tiene UN solo acento
GRIS_BG = FONDO
GRIS_TEXTO = TEXTO
GRIS_TEXTO_SUAVE = TEXTO_SUAVE
BORDE_SUAVE = BORDE
ROJO_ES = ROJO               # Danger (acciones destructivas)
VERDE_OK = VERDE             # Success

IMG_EXT = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")

# Plantilla -> (archivo, logo relativo, clave de contexto del logo)
CARATULAS = {
    "clasica": ("template_caratula.html",
                "Tabla visual refresh/assets/logo_es_crop.png", "logo_path"),
    "ministerio_salud": ("template_ministerio_salud.html",
                         "Tabla visual refresh/assets/ministerio_salud_banner.png",
                         "logo_ministerio"),
}

# --------------------------------------------------------------------------
# LOGGING
# --------------------------------------------------------------------------
LOG_PATH = bd_manager.dir_appdata() / "app.log"
logger = logging.getLogger("submitals_v3")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    try:
        fh = logging.FileHandler(LOG_PATH, encoding="utf-8")
        fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
        logger.addHandler(fh)
    except Exception:
        pass


# ==========================================================================
# PARTE A) ORQUESTACION DE ENTREGABLES  (reutiliza el motor v2.6)
# ==========================================================================
def _cargar_motor(base_dir=BASE_DIR):
    """Carga ``generate_caratulas.py`` como modulo (import dinamico, igual que
    hace la GUI v2.6). Devuelve el modulo ``gc``."""
    ruta = Path(base_dir) / "generate_caratulas.py"
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontro el motor de carATulas: {ruta}")
    spec = importlib.util.spec_from_file_location("generate_caratulas", str(ruta))
    gc = importlib.util.module_from_spec(spec)
    sys.modules["generate_caratulas"] = gc
    spec.loader.exec_module(gc)
    return gc


def _ctx_proyecto(tipo, proyecto, item):
    """Construye el contexto de proyecto para la plantilla de carATula."""
    dp = proyecto.get("datos_procedimiento", {}) or {}
    if tipo == "ministerio_salud":
        proj = proyecto.get("datos_proyecto", {}) or {}
        return {
            "proyecto": proj.get("proyecto", dp.get("detalle", "")),
            "cliente": proj.get("cliente", dp.get("institucion", "")),
            "contrato": proj.get("contrato", dp.get("numero_procedimiento", "")),
            "monto": proj.get("monto", dp.get("monto", "")),
            "plazo": proj.get("plazo", dp.get("plazo", "")),
            "nombre_cargo": proj.get("nombre_cargo", ""),
            "fecha": proj.get("fecha", ""),
            "fecha_emision": proj.get("fecha_emision", ""),
            "fecha_revision": proj.get("fecha_revision", ""),
            "registro": item.get("consecutivo", ""),
            "revisa": proj.get("revisa", ""),
            "version": proj.get("version", "v1"),
            "documentacion_tecnica": "",
            "observaciones_material": item.get("aspectos_adicionales", ""),
            "observaciones_respuesta": "",
            "estado": "",
        }
    # clasica
    return {
        "numero_procedimiento": dp.get("numero_procedimiento", ""),
        "nombre_institucion": dp.get("institucion", ""),
        "detalle_procedimiento": dp.get("detalle", ""),
        "duracion_contrato": dp.get("plazo", ""),
        "monto": dp.get("monto", ""),
    }


def generar_caratulas(base_dir, datos, tipo, proyecto, log=print):
    """Genera la carATula PDF de cada material usando el motor v2.6."""
    from jinja2 import Template
    gc = _cargar_motor(base_dir)
    engines = gc.available_engines()
    if not engines:
        raise RuntimeError("No hay motor de PDF (instale playwright + chromium)")

    tpl_rel, logo_rel, logo_key = CARATULAS.get(tipo, CARATULAS["clasica"])
    tpl_text = (Path(base_dir) / tpl_rel).read_text(encoding="utf-8")
    template = Template(tpl_text)

    logo_file = Path(base_dir) / logo_rel
    if logo_file.exists():
        gc.LOGO_URI = gc.file_uri(logo_file)
    else:
        gc.LOGO_URI = ""
        log(f"AVISO: logo no encontrado ({logo_rel}); se genera sin logo")

    ok = 0
    for item in datos["materiales"]:
        extra = _ctx_proyecto(tipo, proyecto, item)
        if tipo == "ministerio_salud":
            extra["logo_ministerio"] = gc.LOGO_URI
        try:
            gc.process_material(item, template, engines, extra_ctx=extra)
            ok += 1
        except Exception as e:
            log(f"ERROR carATula {item.get('consecutivo')}: {e}")
    log(f"CarATulas generadas: {ok}/{len(datos['materiales'])}")
    return ok


# --- Compilacion de PDFs (pypdf), replicando la logica de v2.6 ---------------
def imagen_a_pdf_reader(path):
    """Convierte una imagen a un PdfReader de una pagina (via PIL)."""
    import io
    from PIL import Image
    from pypdf import PdfReader
    with Image.open(path) as im:
        im = im.convert("RGB")
        buf = io.BytesIO()
        im.save(buf, format="PDF", resolution=150.0)
    buf.seek(0)
    return PdfReader(buf)


def _caratula_reader_para_compilar(caratula_path, prefijo_campos="", log=print):
    """PdfReader de la 1a pagina de la caratula, lista para entrar a un
    compilado: apariencia visual refrescada y (si se indica
    ``prefijo_campos``) campos renombrados para que sean unicos.

    Historia del problema (v3.3.6 y antes): la caratula se genera con campos
    AcroForm con ``NeedAppearances`` -- le pide al lector que DIBUJE el valor
    el mismo. Acrobat lo hace; Chrome, Preview y varios visores livianos no,
    asi que el campo se veia vacio dentro del compilado.

    v3.3.7 lo arreglo aplanando (``fitz.Document.bake``): el texto quedaba
    incrustado en el contenido de la pagina, visible en cualquier lector, pero
    el campo dejaba de ser editable.

    Este cambio resuelve las 2 cosas SIN perder edicion:
      1. ``widget.update()`` regenera la apariencia visual correcta de cada
         campo (se ve bien en cualquier visor) sin destruir el widget -- sigue
         siendo un campo real, editable donde el lector soporte formularios
         (Acrobat, Foxit, PDF-XChange, navegadores de escritorio).
      2. Con ``prefijo_campos`` se renombra cada campo (ej. ``ARQ01_marca``).
         Todas las caratulas salen del mismo template, con los MISMOS nombres
         de campo -- al fusionar varias en un compilado (``CMP SUBMITTAL
         <disciplina>.pdf``, que puede juntar decenas), el PDF trata los
         campos homonimos como un solo campo logico: Acrobat termina
         mostrando el valor de uno solo y deja el resto en blanco (bug real,
         reportado en un compilado de +500 paginas donde solo se veian los
         datos de la 1a caratula). Con el prefijo cada campo queda unico y
         esto no puede volver a pasar, sin importar cuantas se fusionen.

    Si PyMuPDF no esta disponible o falla, se cae a la caratula tal cual (los
    campos quedan como estaban, sin renombrar ni refrescar).
    """
    import io
    from pypdf import PdfReader
    ruta = str(caratula_path)
    try:
        import fitz
        doc = fitz.open(ruta)
        try:
            for page in doc:
                for widget in page.widgets():
                    if prefijo_campos:
                        widget.field_name = f"{prefijo_campos}_{widget.field_name}"
                    widget.update()
            datos = doc.tobytes(garbage=3, deflate=True)
        finally:
            doc.close()
        return PdfReader(io.BytesIO(datos))
    except Exception as e:
        log(f"AVISO: no se pudo preparar la caratula para el compilado ({e}); "
            "se usa la version original")
        return PdfReader(ruta)


def generar_compilado(caratula_path, doc_paths, out_path, log=print):
    """Une la carATula (solo 1a pagina, con la apariencia de sus campos
    refrescada para que se vea bien en cualquier visor y siga siendo
    editable) con las fichas de la carpeta -> PDF ``-CMP.pdf``.
    Replica ``generar_compilado`` v2.6.
    """
    from pypdf import PdfWriter, PdfReader
    w = PdfWriter()
    if caratula_path and Path(caratula_path).exists():
        try:
            w.append(_caratula_reader_para_compilar(caratula_path, log=log), pages=(0, 1))
        except Exception:
            w.add_page(PdfReader(str(caratula_path)).pages[0])
    anexados = 0
    for d in sorted(doc_paths, key=lambda x: x.name.lower()):
        ext = d.suffix.lower()
        try:
            if ext == ".pdf":
                w.append(PdfReader(str(d)))
                anexados += 1
            elif ext in IMG_EXT:
                w.append(imagen_a_pdf_reader(d))
                anexados += 1
        except Exception as e:
            log(f"AVISO: no se anexo '{d.name}' ({e})")
    tmp = out_path.with_suffix(".cmp.tmp")
    with open(tmp, "wb") as f:
        w.write(f)
    os.replace(tmp, out_path)
    w.close()
    return anexados


def _docs_carpeta(carpeta):
    """Documentos anexables de una carpeta (excluye carATula y compilados)."""
    out = []
    for p in sorted(carpeta.iterdir()):
        if not p.is_file():
            continue
        up = p.name.upper()
        if up.startswith("CARATULA") or up.endswith("-CMP.PDF"):
            continue
        if p.suffix.lower() == ".pdf" or p.suffix.lower() in IMG_EXT:
            out.append(p)
    return out


def compilar_cmps(datos, log=print):
    """Genera el ``-CMP.pdf`` de cada material (carATula + fichas)."""
    total = 0
    for mat in datos["materiales"]:
        carpeta = Path(mat["ruta_carpeta"])
        if not carpeta.is_dir():
            continue
        caratulas = sorted(carpeta.glob("CARATULA*.pdf"))
        caratula = caratulas[0] if caratulas else None
        docs = _docs_carpeta(carpeta)
        nombre = bd_manager.sanitizar_nombre(f"{mat['consecutivo']}-{mat['nombre']}-CMP") + ".pdf"
        out = carpeta / nombre
        try:
            generar_compilado(caratula, docs, out, log=log)
            mat["compilado_generado"] = nombre
            total += 1
        except Exception as e:
            log(f"ERROR CMP {mat['consecutivo']}: {e}")
    log(f"CMPs por material generados: {total}")
    return total


def compilar_disciplinas(destino, log=print):
    """Genera ``CMP SUBMITTAL <DISCIPLINA>.pdf`` por cada carpeta madre presente.
    Replica ``compilar_por_disciplina`` v2.6."""
    from pypdf import PdfWriter, PdfReader
    destino = Path(destino)
    generados = []
    for cat, madre in bd_manager.CATEGORIAS.items():
        carpeta_disc = destino / madre
        if not carpeta_disc.is_dir():
            continue
        entradas = []
        for sub in sorted(carpeta_disc.iterdir()):
            if not sub.is_dir():
                continue
            m = re.match(rf"^({cat})(\d+)-(.*)$", sub.name)
            if m:
                entradas.append((int(m.group(2)), sub))
        entradas.sort(key=lambda e: e[0])
        if not entradas:
            continue
        w = PdfWriter()
        procesados = 0
        for n, sub in entradas:
            caratulas = sorted(sub.glob("CARATULA*.pdf"))
            if not caratulas:
                log(f"AVISO {madre}: '{sub.name}' sin carATula, se omite")
                continue
            try:
                # Prefijo unico por caratula (ej. "ARQ01"): sin esto, al
                # fusionar varias del mismo template (mismos nombres de
                # campo) el PDF resultante colapsa los campos homonimos y
                # solo se ve el valor de la 1a caratula en cualquier lector.
                prefijo = f"{cat}{n:02d}"
                w.append(_caratula_reader_para_compilar(
                    caratulas[0], prefijo_campos=prefijo, log=log), pages=(0, 1))
            except Exception as e:
                log(f"AVISO {madre}: carATula ilegible en '{sub.name}' ({e})")
                continue
            for d in _docs_carpeta(sub):
                try:
                    if d.suffix.lower() == ".pdf":
                        w.append(PdfReader(str(d)))
                    elif d.suffix.lower() in IMG_EXT:
                        w.append(imagen_a_pdf_reader(d))
                except Exception as e:
                    log(f"AVISO {madre}: no se anexo '{sub.name}/{d.name}' ({e})")
            procesados += 1
        if procesados == 0:
            w.close()
            continue
        singular = bd_manager.DISCIPLINA_SINGULAR.get(madre, madre)
        out = carpeta_disc / f"CMP SUBMITTAL {singular}.pdf"
        tmp = out.with_suffix(".tmp")
        with open(tmp, "wb") as f:
            w.write(f)
        os.replace(tmp, out)
        w.close()
        generados.append(out.name)
    log(f"Compilados por disciplina: {', '.join(generados) or '(ninguno)'}")
    return generados


# --- Excel (openpyxl), replicando columnas de v2.6 ---------------------------
def _agrupar_por_disciplina(materiales):
    nombres = {"ARQ": "Arquitectónicos", "ESTR": "Estructurales",
               "MEC": "Mecánicos", "ELEC": "Eléctricos"}
    grupos = {}
    for it in materiales:
        pref = re.sub(r"\d.*", "", str(it.get("consecutivo", ""))).upper()
        grupos.setdefault(nombres.get(pref, pref or "Otros"), []).append(it)
    return grupos


def generar_excel_submittal(datos, destino):
    """``Guía Submittal.xlsx`` (para administracion): Consecutivo | Descripción |
    Aprobación | Observaciones (las 2 ultimas en blanco)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    grupos = _agrupar_por_disciplina(datos["materiales"])
    if not any(grupos.values()):
        return {"exitoso": False, "error": "Sin materiales"}
    wb = Workbook(); wb.remove(wb.active)
    hfill = PatternFill("solid", fgColor="4472C4")
    hfont = Font(color="FFFFFF", bold=True)
    halign = Alignment(horizontal="center")
    for hoja, items in grupos.items():
        ws = wb.create_sheet(title=hoja[:31])
        for col, h in enumerate(["Consecutivo", "Descripción", "Aprobación", "Observaciones"], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill; c.font = hfont; c.alignment = halign
        for i, it in enumerate(items, 2):
            ws.cell(row=i, column=1, value=it.get("consecutivo", ""))
            ws.cell(row=i, column=2, value=it.get("nombre", ""))
        for col, w in zip("ABCD", (15, 60, 20, 40)):
            ws.column_dimensions[col].width = w
    out = Path(destino) / "Guía Submittal.xlsx"
    wb.save(out)
    return {"exitoso": True, "archivo": str(out), "total_materiales": len(datos["materiales"])}


def generar_excel_materiales(datos, destino):
    """``Guía interna materiales.xlsx``: Consecutivo | Familia | Descripción |
    Normativa | Estado | Proveedor."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    grupos = _agrupar_por_disciplina(datos["materiales"])
    if not any(grupos.values()):
        return {"exitoso": False, "error": "Sin materiales"}
    wb = Workbook(); wb.remove(wb.active)
    hfill = PatternFill("solid", fgColor="4472C4")
    hfont = Font(color="FFFFFF", bold=True)
    halign = Alignment(horizontal="center")
    for hoja, items in grupos.items():
        ws = wb.create_sheet(title=hoja[:31])
        for col, h in enumerate(["Consecutivo", "Familia", "Descripción", "Normativa",
                                 "Estado", "Proveedor"], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill; c.font = hfont; c.alignment = halign
        for i, it in enumerate(items, 2):
            ws.cell(row=i, column=1, value=it.get("consecutivo", ""))
            ws.cell(row=i, column=2, value=it.get("marca", ""))
            ws.cell(row=i, column=3, value=it.get("nombre", ""))
            ws.cell(row=i, column=4, value=it.get("normativa", "SIN ESPECIFICAR"))
            estado = "DISPONIBLE" if (it.get("estado") == "FICHA_DISPONIBLE"
                                      and not it.get("carpeta_vacia")) else "FALTANTE"
            ws.cell(row=i, column=5, value=estado)
        for col, w in zip("ABCDEF", (15, 25, 60, 30, 15, 25)):
            ws.column_dimensions[col].width = w
    out = Path(destino) / "Guía interna materiales.xlsx"
    wb.save(out)
    return {"exitoso": True, "archivo": str(out), "total_materiales": len(datos["materiales"])}


def limpiar_entregables(destino, log=print):
    """Borra carATulas, CMPs por material y compilados/Excel previos para poder
    regenerar TODO limpiamente (usado al editar un submittal)."""
    destino = Path(destino)
    for patron in ("Guía Submittal.xlsx", "Guía interna materiales.xlsx"):
        p = destino / patron
        if p.exists():
            try: p.unlink()
            except Exception: pass
    for madre in bd_manager.CATEGORIAS.values():
        cm = destino / madre
        if not cm.is_dir():
            continue
        for p in cm.glob("CMP SUBMITTAL *.pdf"):
            try: p.unlink()
            except Exception: pass
        for sub in cm.iterdir():
            if not sub.is_dir():
                continue
            for p in list(sub.glob("CARATULA*.pdf")) + list(sub.glob("*-CMP.pdf")):
                try: p.unlink()
                except Exception: pass
    log("Entregables previos limpiados")


def generar_entregables(bd, proyecto, destino, tipo="clasica", log=print,
                        con_caratulas=True, limpiar=True):
    """Pipeline completo de generacion de entregables desde la BD:

      1. Valida el submittal.
      2. Materializa la carpeta destino (arbol + copia de fichas + JSONs).
      3. (opcional) Genera carATulas con el motor v2.6.
      4. Compila los CMP por material.
      5. Compila los CMP por disciplina.
      6. Genera los dos Excel.

    Devuelve un dict con el resultado. Lanza ``bd_manager.BDError`` si la
    validacion falla.
    """
    ok, errores = bd.validar_proyecto(proyecto)
    if not ok:
        raise bd_manager.BDError("Validacion fallida:\n - " + "\n - ".join(errores))

    destino = Path(destino)
    if limpiar and destino.exists():
        limpiar_entregables(destino, log=log)

    json_path = bd.materializar_proyecto(proyecto, destino)
    datos = json.loads(json_path.read_text(encoding="utf-8"))
    log(f"Materializado en {destino} ({len(datos['materiales'])} materiales)")

    if con_caratulas:
        generar_caratulas(destino if (destino / "generate_caratulas.py").exists() else BASE_DIR,
                          datos, tipo, proyecto, log=log)

    compilar_cmps(datos, log=log)
    compilar_disciplinas(destino, log=log)
    r1 = generar_excel_submittal(datos, destino)
    r2 = generar_excel_materiales(datos, destino)

    # Persistir estado del submittal
    proyecto["entregables_generados"] = True
    proyecto["tipo_caratula"] = tipo
    bd.guardar_submittal(proyecto, destino=destino)
    # Reescribir datos_materiales con compilado_generado actualizado
    json_path.write_text(json.dumps(datos, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"destino": str(destino), "materiales": len(datos["materiales"]),
            "excel_submittal": r1.get("exitoso"), "excel_materiales": r2.get("exitoso")}


# ==========================================================================
# PARTE B) GUI (tkinter)
# ==========================================================================
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    try:
        import customtkinter as ctk
    except ImportError:
        # Dependencia nueva del refresh visual v3.3.0: si falta, se instala
        # sola (solo tiene sentido corriendo el .py con un interprete real;
        # el .exe empaquetado ya la trae incluida via el .spec de PyInstaller).
        if getattr(sys, "frozen", False):
            raise
        import subprocess
        print("Instalando dependencia 'customtkinter' (primera vez)...")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "customtkinter"],
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
        import customtkinter as ctk
    _TK_OK = True
except Exception:
    _TK_OK = False


if _TK_OK:

    # -------------------------------------------------------------------
    # Sistema de diseno v3.3.0: modo claro fijo (ver AVOID en el spec: nada
    # de modo oscuro por defecto), tipografia Inter con reserva a Segoe UI
    # si el sistema no la tiene instalada.
    # -------------------------------------------------------------------
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("blue")

    # Tipografia del diseno: IBM Plex Sans para texto e IBM Plex Mono para
    # datos (consecutivos, conteos, rutas, horas). Si la PC no las tiene
    # instaladas se cae a las del sistema, que son metricamente parecidas.
    FUENTE = "Segoe UI"
    FUENTE_MONO = "Consolas"
    try:
        import tkinter.font as _tkfont
        _r = tk.Tk()
        _r.withdraw()
        _fams = set(_tkfont.families())
        for _cand in ("IBM Plex Sans", "Inter", "Segoe UI Variable Text", "Segoe UI"):
            if _cand in _fams:
                FUENTE = _cand
                break
        for _cand in ("IBM Plex Mono", "Cascadia Mono", "Consolas"):
            if _cand in _fams:
                FUENTE_MONO = _cand
                break
        _r.destroy()
    except Exception:
        pass

    def _fuente(size=11, weight="normal"):
        """Fuente del sistema de diseno (``CTkFont``: escala con el DPI)."""
        return ctk.CTkFont(family=FUENTE, size=size, weight=weight)

    def _mono(size=11, weight="normal"):
        """Monoespaciada del sistema de diseno: TODO dato tabular (consecutivo,
        conteo, ruta, hora, porcentaje) va en esta fuente."""
        return ctk.CTkFont(family=FUENTE_MONO, size=size, weight=weight)

    def _configurar_estilo_ttk():
        """Reskin del ``ttk.Treeview`` (unico widget ttk que sigue en uso;
        CustomTkinter no trae reemplazo) a la paleta nueva.

        El diseno usa cabeceras claras con texto tenue en mayusculas y la fila
        seleccionada como TINTE del acento (no un bloque solido de color): asi
        la seleccion no tapa el dato ni pelea con el resto de la pantalla.
        """
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Treeview", background=SUPERFICIE,
                        fieldbackground=SUPERFICIE, foreground=TEXTO_2,
                        rowheight=28, borderwidth=0, font=(FUENTE, 10))
        style.configure("Treeview.Heading", background=SUPERFICIE_3,
                        foreground=TEXTO_SUAVE, font=(FUENTE_MONO, 9),
                        relief="flat", borderwidth=0, padding=(6, 6))
        style.map("Treeview.Heading", background=[("active", BORDE_TENUE)])
        style.map("Treeview", background=[("selected", CHIP_BG)],
                  foreground=[("selected", ACENTO_TXT)])
        style.layout("Treeview", [("Treeview.treearea", {"sticky": "nswe"})])
        # Scrollbars al tono de los bordes (antes quedaban gris azulado).
        style.configure("Vertical.TScrollbar", background=BORDE,
                        troughcolor=SUPERFICIE_3, bordercolor=SUPERFICIE_3,
                        arrowcolor=TEXTO_SUAVE, relief="flat")
        style.configure("Horizontal.TScrollbar", background=BORDE,
                        troughcolor=SUPERFICIE_3, bordercolor=SUPERFICIE_3,
                        arrowcolor=TEXTO_SUAVE, relief="flat")

    # Color de "hover" de cada color solido del tema (mismo tono, mas oscuro).
    _HOVER = {ACENTO: ACENTO_HOVER, ACENTO_HOVER: ACENTO_TXT, ROJO: "#8D0005",
             VERDE: "#005A2B", "white": SUPERFICIE_3, SUPERFICIE: SUPERFICIE_3,
             SIDEBAR_BTN: SIDEBAR_BTN_H}

    def _tarjeta(parent, **kw):
        """Frame estilo 'tarjeta': fondo blanco, esquinas redondeadas suaves y
        borde de 1px. El diseno es plano (sin sombras ni simil-vidrio): la
        jerarquia la dan el borde y el fondo, no la profundidad."""
        kw.setdefault("fg_color", SUPERFICIE)
        kw.setdefault("corner_radius", 9)
        kw.setdefault("border_width", 1)
        kw.setdefault("border_color", BORDE)
        return ctk.CTkFrame(parent, **kw)

    def _boton(parent, text, command, color=ACENTO, texto_color="white",
              ancho=140, alto=34, **kw):
        """Boton de accion primaria: solido, esquinas de 6px."""
        return ctk.CTkButton(parent, text=text, command=command, fg_color=color,
                             hover_color=_HOVER.get(color, color),
                             text_color=texto_color, width=ancho, height=alto,
                             corner_radius=6, font=_fuente(12, "bold"), **kw)

    def _boton_secundario(parent, text, command, ancho=140, alto=32, **kw):
        """Boton de accion secundaria: contorno neutro sobre fondo blanco (en
        el diseno el color se reserva para la accion primaria de cada
        pantalla, para que siempre haya UNA sola obvia)."""
        return ctk.CTkButton(parent, text=text, command=command,
                             fg_color=SUPERFICIE, hover_color=SUPERFICIE_3,
                             text_color=TEXTO_2, border_width=1,
                             border_color=BORDE, width=ancho, height=alto,
                             corner_radius=6, font=_fuente(12), **kw)

    def _boton_peligro(parent, text, command, ancho=140, alto=32, **kw):
        """Accion destructiva: contorno rojo, no bloque rojo. Se distingue de
        la accion primaria (solida) sin gritar en toda la pantalla."""
        return ctk.CTkButton(parent, text=text, command=command,
                             fg_color=SUPERFICIE, hover_color=ROJO_SUAVE,
                             text_color=ROJO, border_width=1,
                             border_color=ROJO_BORDE, width=ancho, height=alto,
                             corner_radius=6, font=_fuente(12), **kw)

    def _enlace(parent, text, command, **kw):
        """Accion terciaria de bajo peso (estilo enlace): mono, sin recuadro."""
        b = ctk.CTkButton(parent, text=text, command=command,
                          fg_color="transparent", hover_color=SUPERFICIE_3,
                          text_color=TEXTO_SUAVE, width=1, height=24,
                          corner_radius=5, font=_mono(11), **kw)
        return b

    def _etiqueta_seccion(parent, texto, color=TEXTO_SUAVE, **kw):
        """Rotulo de seccion del diseno: mono y en mayusculas. (El diseno le
        pone letter-spacing; Tkinter no lo soporta, asi que se compensa con la
        mono, que ya es mas abierta.)"""
        return ctk.CTkLabel(parent, text=texto.upper(), font=_mono(10),
                            text_color=color, **kw)

    def _consola(parent, **kw):
        """Caja de registro (el unico bloque oscuro del area de contenido)."""
        kw.setdefault("corner_radius", 8)
        kw.setdefault("fg_color", LOG_BG)
        kw.setdefault("text_color", LOG_TXT)
        kw.setdefault("border_width", 0)
        return ctk.CTkTextbox(parent, font=(FUENTE_MONO, 10), **kw)

    def _tabla_ttk(parent, columnas, titulos, anchos, alineados=(), alto=10):
        """``ttk.Treeview`` dentro de una tarjeta con borde, como en el diseno
        (la tabla es una superficie con marco, no un widget suelto).

        Devuelve ``(marco, tree, pie)``; ``pie`` es la franja inferior de la
        tarjeta, lista para poner el conteo de resultados (o vacia).
        """
        marco = _tarjeta(parent, corner_radius=8)
        marco.grid_rowconfigure(0, weight=1)
        marco.grid_columnconfigure(0, weight=1)
        interior = ctk.CTkFrame(marco, fg_color=SUPERFICIE, corner_radius=8)
        interior.grid(row=0, column=0, sticky="nsew", padx=1, pady=(1, 0))
        tree = ttk.Treeview(interior, columns=columnas, show="headings", height=alto)
        for c in columnas:
            # El encabezado se alinea igual que su columna (por defecto ttk lo
            # centra, y quedaba desfasado del dato que rotula).
            alineacion = "center" if c in alineados else "w"
            tree.heading(c, text=titulos[c], anchor=alineacion)
            tree.column(c, width=anchos[c], minwidth=40, anchor=alineacion)
        vsb = ttk.Scrollbar(interior, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        interior.grid_rowconfigure(0, weight=1)
        interior.grid_columnconfigure(0, weight=1)
        pie = ctk.CTkFrame(marco, fg_color=SUPERFICIE_2, corner_radius=0)
        pie.grid(row=1, column=0, sticky="ew", padx=1, pady=(0, 1))
        return marco, tree, pie

    def _clicable(marco, command, normal="transparent", hover=SUPERFICIE_3,
                  borde_normal=None, borde_hover=None):
        """Convierte un ``CTkFrame`` (con todo lo que tenga adentro) en una
        tarjeta clicable con hover, como las del diseno.

        En Tkinter el clic y el ``<Enter>``/``<Leave>`` NO se heredan: hay que
        atarlos a cada hijo, si no la tarjeta "no responde" justo donde esta el
        texto. Las ataduras se reemplazan (no se acumulan) para poder repintar
        el mismo marco varias veces.
        """
        def _pintar(estado):
            kw = {"fg_color": hover if estado else normal}
            if borde_hover is not None or borde_normal is not None:
                kw["border_color"] = (borde_hover if estado else borde_normal) or BORDE
            try:
                marco.configure(**kw)
            except Exception:
                pass

        def _recorrer(w):
            w.bind("<Button-1>", lambda _ev: command())
            w.bind("<Enter>", lambda _ev: _pintar(True))
            w.bind("<Leave>", lambda _ev: _pintar(False))
            try:
                w.configure(cursor="hand2")
            except Exception:
                pass
            for hijo in w.winfo_children():
                _recorrer(hijo)

        _recorrer(marco)
        return marco

    def _pista(entry, var, texto, size=11):
        """Texto de ayuda DENTRO de un campo vacio.

        ``CTkEntry`` ignora ``placeholder_text`` cuando el campo tiene
        ``textvariable`` (limitacion de CustomTkinter), y todos los campos de
        esta app usan variables. Sin esto, los filtros y la API key se ven como
        recuadros vacios sin decir que va adentro.
        """
        lbl = ctk.CTkLabel(entry, text=texto, font=_fuente(size),
                           text_color=TEXTO_TENUE, fg_color="transparent")

        def _refrescar(*_a):
            try:
                if var.get():
                    lbl.place_forget()
                else:
                    lbl.place(x=11, rely=0.5, anchor="w")
            except Exception:
                pass

        lbl.bind("<Button-1>", lambda _ev: entry.focus_set())
        var.trace_add("write", _refrescar)
        _refrescar()
        return entry

    def _tooltip(widget, texto):
        """Globo de ayuda al pasar el mouse. Se usa en los botones que quedan
        como icono solo (↑ / ↓): sin esto la accion no se puede saber sin
        probarla."""
        estado = {"win": None}

        def _mostrar(_ev=None):
            if estado["win"] is not None:
                return
            try:
                x = widget.winfo_rootx() + 6
                y = widget.winfo_rooty() + widget.winfo_height() + 6
                win = tk.Toplevel(widget)
                win.overrideredirect(True)
                win.attributes("-topmost", True)
                win.geometry(f"+{x}+{y}")
                tk.Label(win, text=texto, bg=TEXTO, fg="#FFFFFF",
                         font=(FUENTE, 9), padx=8, pady=4, bd=0).pack()
                estado["win"] = win
            except Exception:
                estado["win"] = None

        def _ocultar(_ev=None):
            win = estado["win"]
            estado["win"] = None
            if win is not None:
                try:
                    win.destroy()
                except Exception:
                    pass

        widget.bind("<Enter>", _mostrar, add="+")
        widget.bind("<Leave>", _ocultar, add="+")
        widget.bind("<Button-1>", _ocultar, add="+")
        return widget

    _RE_EMOJI_INICIAL = re.compile(r"^[^\w¿¡(]+\s*")

    def _limpiar_estado(texto):
        """Quita el emoji inicial de los mensajes de estado: en el diseno el
        estado se comunica con el punto de color, no con iconos."""
        return _RE_EMOJI_INICIAL.sub("", str(texto)).strip() or str(texto)

    class _EtiquetaSync(ctk.CTkLabel):
        """Etiqueta de estado de sincronizacion del menu lateral.

        Todo el codigo de sincronizacion/actualizacion de v3.3.x le escribe con
        los colores del tema CLARO (``AZUL_ES``, ``VERDE_OK``, ``ROJO_ES``,
        ``GRIS_TEXTO_SUAVE``); sobre el sidebar oscuro esos colores no se leen.
        Esta subclase traduce el color recibido a un par (texto legible, color
        del punto de estado), asi los flujos existentes no se tocan.
        """

        _punto = None
        _MAPA = {ACENTO: (SIDEBAR_TXT, SIDEBAR_TXT_3),
                 VERDE: (SIDEBAR_TXT, VERDE_PUNTO),
                 ROJO: (ROJO_CLARO, ROJO_CLARO),
                 TEXTO_SUAVE: (SIDEBAR_TXT_3, SIDEBAR_TXT_3)}

        def configure(self, require_redraw=False, **kw):
            if "text" in kw:
                kw["text"] = _limpiar_estado(kw["text"])
            color = kw.get("text_color")
            if color is not None:
                texto_color, punto_color = self._MAPA.get(
                    color, (SIDEBAR_TXT, SIDEBAR_TXT_3))
                kw["text_color"] = texto_color
                if self._punto is not None:
                    try:
                        self._punto.configure(text_color=punto_color)
                    except Exception:
                        pass
            super().configure(require_redraw=require_redraw, **kw)

    def _hover_filas(tree, color=FILA_HOVER):
        """Resalta la fila bajo el cursor (el diseno lo hace con ``:hover``;
        en ``ttk.Treeview`` hay que seguir el mouse a mano)."""
        tree.tag_configure("_hover", background=color)
        estado = {"iid": None}

        def _mover(ev):
            iid = tree.identify_row(ev.y)
            if iid == estado["iid"]:
                return
            anterior = estado["iid"]
            if anterior and tree.exists(anterior):
                tags = [t for t in tree.item(anterior, "tags") if t != "_hover"]
                tree.item(anterior, tags=tags)
            estado["iid"] = iid
            if iid and tree.exists(iid):
                tags = list(tree.item(iid, "tags"))
                # Las filas con tag propio (ej. la animacion de "recien
                # agregado") conservan el suyo: _hover solo se suma.
                if "_hover" not in tags:
                    tree.item(iid, tags=tags + ["_hover"])

        def _salir(_ev):
            iid = estado["iid"]
            if iid and tree.exists(iid):
                tree.item(iid, tags=[t for t in tree.item(iid, "tags")
                                     if t != "_hover"])
            estado["iid"] = None

        tree.bind("<Motion>", _mover, add="+")
        tree.bind("<Leave>", _salir, add="+")

    def _avisar_resultado_git(parent, r, mensaje_ok):
        """Muestra el resultado de un ``git_push()`` (subido/offline/sin token/
        error), igual que ``VentanaCatalogo._avisar_push`` pero reutilizable
        desde cualquier ventana."""
        if r.get("offline"):
            messagebox.showinfo("Sin conexión",
                                mensaje_ok + "\n\nEl cambio se guardó localmente "
                                "y se subirá al reconectar.", parent=parent)
        elif r.get("auth"):
            messagebox.showwarning("Falta el token",
                                   mensaje_ok + "\n\nNo se pudo subir a GitHub: "
                                   "configure el token.", parent=parent)
        elif r.get("subido") or r.get("nada_que_subir") or r.get("desactivado"):
            messagebox.showinfo("Listo", mensaje_ok, parent=parent)
        else:
            messagebox.showwarning("No se pudo subir",
                                   mensaje_ok + f"\n\n{r.get('error', '')}", parent=parent)

    def _traer_al_frente(win):
        """Fuerza que una ventana nueva quede al frente y con el foco.

        En Windows, un Toplevel recien creado a veces aparece detras de la
        ventana principal (el usuario "pierde" la ventana que tenia abierta).
        lift()/focus_force() no siempre alcanzan; el truco de -topmost
        momentaneo si funciona de forma consistente.
        """
        win.lift()
        win.attributes("-topmost", True)
        win.after_idle(lambda: win.attributes("-topmost", False))
        win.focus_force()

    def _dimensionar_ventana(win, ancho=0, alto=0, margen=90):
        """Abre la ventana con un tamano que garantice que TODOS los controles
        (incluidos los botones de abajo y los de la derecha) queden visibles, y
        la centra en pantalla.

        Se usa el tamano REQUERIDO por el contenido (``winfo_reqwidth/height``,
        que ya suma la barra de botones inferior y la fila de botones mas ancha)
        como piso: la ventana nunca abre — ni el usuario puede encogerla —
        por debajo de lo que hace falta para ver todo. ``ancho``/``alto`` son un
        tamano comodo preferido; si el contenido pide mas, se respeta el
        contenido. Todo acotado al tamano de pantalla menos un margen.
        """
        win.update_idletasks()
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        req_w, req_h = win.winfo_reqwidth(), win.winfo_reqheight()
        min_w = min(max(req_w, 320), sw - margen)
        min_h = min(max(req_h, 240), sh - margen)
        w = min(max(ancho, min_w), sw - margen)
        h = min(max(alto, min_h), sh - margen)
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h) // 2 - 20)
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.minsize(min_w, min_h)

    def _dimensionar_principal(win, ancho=1240, alto=800, min_w=1100, min_h=670):
        """Abre la ventana PRINCIPAL al tamano del diseno (1240x800), centrada y
        acotada a la pantalla.

        A diferencia de ``_dimensionar_ventana`` (para dialogos), aca NO se usa
        el tamano requerido por el contenido como piso: el shell tiene tablas y
        paneles que se estiran, y en un monitor de 1366x768 exigir el tamano
        "ideal" dejaria la ventana mas grande que la pantalla. Se fija un minimo
        razonable, siempre menor que la pantalla disponible.
        """
        win.update_idletasks()
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        min_w = min(min_w, max(720, sw - 60))
        min_h = min(min_h, max(520, sh - 60))
        w = max(min_w, min(ancho, sw - 80))
        h = max(min_h, min(alto, sh - 90))
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h) // 2 - 20)
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.minsize(min_w, min_h)

    def _vista_previa_ficha(master, bd, ficha):
        """Abre el PDF de una ficha tecnica con el visor predeterminado."""
        if not ficha:
            return
        try:
            ruta = bd.ruta_local_ficha(ficha)
        except Exception as e:
            messagebox.showerror("Vista previa", f"No se pudo obtener el PDF:\n{e}",
                                 parent=master.winfo_toplevel())
            return
        try:
            os.startfile(ruta)
        except Exception as e:
            messagebox.showerror("Vista previa", f"No se pudo abrir el archivo:\n{e}",
                                 parent=master.winfo_toplevel())

    class DatosProyectoDialog(ctk.CTkToplevel):
        """Dialogo para capturar los datos del procedimiento (ninguno obligatorio:
        a veces se arma la caratula antes de tener el numero de procedimiento o
        el monto definitivos)."""

        def __init__(self, master, datos=None):
            super().__init__(master)
            self.title("Datos del Proyecto")
            self.resultado = None
            self.configure(fg_color=GRIS_BG, padx=16, pady=16)
            self.grab_set()
            datos = datos or {}
            campos = [
                ("numero_procedimiento", "Número de procedimiento"),
                ("institucion", "Institución"),
                ("detalle", "Detalle"),
                ("plazo", "Plazo"),
                ("monto", "Monto"),
            ]
            tarjeta = _tarjeta(self)
            tarjeta.pack(fill="both", expand=True, padx=4, pady=4)
            ctk.CTkLabel(tarjeta, text="Datos del Proyecto", font=_fuente(14, "bold"),
                        text_color=GRIS_TEXTO).grid(
                row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 10))
            self.vars = {}
            for i, (clave, etiqueta) in enumerate(campos, start=1):
                ctk.CTkLabel(tarjeta, text=etiqueta + ":", text_color=GRIS_TEXTO).grid(
                    row=i, column=0, sticky="e", pady=4, padx=(16, 6))
                v = tk.StringVar(value=datos.get(clave, ""))
                ctk.CTkEntry(tarjeta, textvariable=v, width=280, height=32,
                            corner_radius=8, border_color=BORDE_SUAVE).grid(
                    row=i, column=1, pady=4, padx=(0, 16))
                self.vars[clave] = v
            barra = ctk.CTkFrame(tarjeta, fg_color="transparent")
            barra.grid(row=len(campos) + 1, column=0, columnspan=2, pady=(14, 16))
            _boton(barra, "Guardar", self._guardar, color=AZUL_ES).pack(side="left", padx=6)
            _boton_secundario(barra, "Cancelar", self.destroy).pack(side="left", padx=6)
            _dimensionar_ventana(self, 460, 360)
            _traer_al_frente(self)

        def _guardar(self):
            self.resultado = {k: v.get().strip() for k, v in self.vars.items()}
            self.destroy()


    class _PinDialog(ctk.CTkToplevel):
        """Pide un PIN enmascarado (reemplaza ``simpledialog.askstring`` con
        ``show='*'``, que CustomTkinter no ofrece de forma nativa)."""

        def __init__(self, master, titulo, mensaje):
            super().__init__(master)
            self.title(titulo)
            self.resultado = None
            self.configure(fg_color=GRIS_BG, padx=16, pady=16)
            self.resizable(False, False)
            self.grab_set()
            tarjeta = _tarjeta(self)
            tarjeta.pack(fill="both", expand=True)
            ctk.CTkLabel(tarjeta, text=mensaje, text_color=GRIS_TEXTO,
                        font=_fuente(11)).pack(padx=18, pady=(18, 8))
            self.v_pin = tk.StringVar()
            e = ctk.CTkEntry(tarjeta, textvariable=self.v_pin, show="*", width=200,
                             height=32, corner_radius=8, border_color=BORDE_SUAVE,
                             justify="center")
            e.pack(padx=18, pady=4)
            e.bind("<Return>", lambda _ev: self._ok())
            barra = ctk.CTkFrame(tarjeta, fg_color="transparent")
            barra.pack(pady=(12, 18))
            _boton(barra, "Aceptar", self._ok, color=AZUL_ES, ancho=110).pack(
                side="left", padx=6)
            _boton_secundario(barra, "Cancelar", self.destroy, ancho=110).pack(
                side="left", padx=6)
            _dimensionar_ventana(self)
            _traer_al_frente(self)
            e.focus_set()

        def _ok(self):
            self.resultado = self.v_pin.get()
            self.destroy()


    class _BuscadorFichas(ctk.CTkFrame):
        """Buscador de fichas reutilizable, con la mejor coincidencia de
        primera, filtros (categoria / marca / modelo / nombre), columnas para
        distinguir un material entre muchos, y lista con scroll.

        Se usa tanto al armar un submittal (``TablaMateriales``, doble clic =
        agregar) como al gestionar la BD (``PantallaBD``, doble clic =
        editar). El widget no sabe que se hace con la ficha elegida: avisa por
        ``on_activar(ficha)`` y expone ``ficha_seleccionada()``.
        """

        def __init__(self, master, bd, on_activar=None, permitir_inactivas=False,
                     alto_filas=12, top_n=80, compacto=False):
            super().__init__(master, fg_color="transparent")
            self.bd = bd
            self.on_activar = on_activar
            self.permitir_inactivas = permitir_inactivas
            self.top_n = top_n
            # ``compacto``: version de 4 columnas para el panel angosto del
            # armado de submittal (el diseno muestra ahi Nombre/Marca/Cat.);
            # la vista completa se usa al gestionar la BD.
            self.compacto = compacto
            self._resultados = []
            self._debounce = None
            self._build(alto_filas)
            self.refrescar()

        def _build(self, alto_filas):
            filtros = ctk.CTkFrame(self, fg_color="transparent")
            filtros.pack(fill="x")
            # Fila 1: busqueda libre (fuzzy) + categoria (+ ver desactivadas)
            f1 = ctk.CTkFrame(filtros, fg_color="transparent")
            f1.pack(fill="x", pady=(0, 6))
            self.var_q = tk.StringVar()
            ent = ctk.CTkEntry(f1, textvariable=self.var_q, height=32,
                               corner_radius=6, border_color=BORDE,
                               fg_color=SUPERFICIE, text_color=TEXTO,
                               font=_fuente(12))
            _pista(ent, self.var_q,
                   "Buscar por nombre, marca, modelo o sinónimo…", size=12)
            ent.pack(side="left", fill="x", expand=True)
            caja_cat = ctk.CTkFrame(f1, fg_color="transparent")
            caja_cat.pack(side="left", padx=(8, 0))
            _etiqueta_seccion(caja_cat, "cat", color=TEXTO_TENUE).pack(side="left",
                                                                      padx=(0, 4))
            self.var_cat = tk.StringVar(value="TODAS")
            ctk.CTkComboBox(caja_cat, variable=self.var_cat, width=104, height=32,
                            corner_radius=6, border_color=BORDE,
                            fg_color=SUPERFICIE, text_color=TEXTO,
                            font=_fuente(12), dropdown_font=_fuente(12),
                            button_color=SUPERFICIE, button_hover_color=SUPERFICIE_3,
                            text_color_disabled=TEXTO_TENUE,
                            state="readonly", dropdown_fg_color=SUPERFICIE,
                            dropdown_text_color=TEXTO,
                            dropdown_hover_color=CHIP_BG,
                            values=["TODAS"] + list(bd_manager.CATEGORIAS)).pack(side="left")
            if self.permitir_inactivas:
                self.var_inact = tk.BooleanVar(value=False)
                ctk.CTkCheckBox(f1, text="Ver desactivadas", variable=self.var_inact,
                               command=self.refrescar, text_color=TEXTO_2,
                               font=_fuente(12), checkbox_width=18, checkbox_height=18,
                               corner_radius=4, border_width=1, border_color=BORDE_FUERTE,
                               fg_color=ACENTO, hover_color=ACENTO_HOVER).pack(
                    side="left", padx=(10, 0))
            else:
                self.var_inact = None

            # Fila 2: filtros por campo. El diseno los dibuja como "pills"; se
            # mantienen como campos SIEMPRE visibles a proposito: esconder un
            # filtro detras de un clic extra le cuesta trabajo a quien arma
            # submittals todo el dia.
            f2 = ctk.CTkFrame(filtros, fg_color="transparent")
            f2.pack(fill="x", pady=(0, 8))
            self.var_marca = tk.StringVar()
            self.var_modelo = tk.StringVar()
            self.var_nombre = tk.StringVar()
            for pista, var in (("Marca", self.var_marca),
                              ("Modelo", self.var_modelo),
                              ("Nombre", self.var_nombre)):
                # La etiqueta va DENTRO del campo en vez de al lado: son tres
                # filtros y el panel es angosto; asi entran los tres en una
                # linea sin recortar nada.
                e = ctk.CTkEntry(f2, textvariable=var, width=104, height=28,
                                corner_radius=14, border_color=BORDE,
                                fg_color=SUPERFICIE, text_color=TEXTO,
                                font=_fuente(11))
                _pista(e, var, pista)
                e.pack(side="left", padx=(0, 6))
            _enlace(f2, "Limpiar", self._limpiar).pack(side="left")

            # Resultados: Treeview con scroll dentro de una tarjeta con borde.
            if self.compacto:
                cols = ("nombre", "marca", "cat", "match")
            else:
                cols = ("nombre", "marca", "modelo", "dim", "cat", "match")
            if self.permitir_inactivas:
                cols = cols + ("estado",)
            # Anchos de arranque chicos a proposito: las columnas se estiran con
            # la ventana (``stretch`` por defecto), y si arrancan anchas obligan
            # a la ventana a abrir mas grande que la pantalla.
            anchos = ({"nombre": 190, "marca": 96, "cat": 46, "match": 54,
                       "estado": 78} if self.compacto else
                      {"nombre": 260, "marca": 110, "modelo": 120, "dim": 100,
                       "cat": 50, "match": 62, "estado": 80})
            titulos = {"nombre": "Nombre", "marca": "Marca", "modelo": "Modelo / espec.",
                       "dim": "Dimensiones", "cat": "Cat.", "match": "Coincid.",
                       "estado": "Estado"}
            marco, self.tree, pie = _tabla_ttk(self, cols, titulos, anchos,
                                               alineados=("cat", "match", "estado"),
                                               alto=alto_filas)
            marco.pack(fill="both", expand=True)
            self.tree.bind("<Double-Button-1>", lambda _ev: self._activar())
            self.tree.bind("<Return>", lambda _ev: self._activar())
            _hover_filas(self.tree)

            # Pie de la tarjeta: conteo de resultados, en mono (es un dato).
            self.lbl_estado = ctk.CTkLabel(pie, text="", text_color=TEXTO_SUAVE,
                                           font=_mono(10))
            self.lbl_estado.pack(anchor="w", padx=12, pady=4)

            # Refresco con pequeno "debounce" para no releer en cada tecla en
            # catalogos grandes.
            for var in (self.var_q, self.var_marca, self.var_modelo, self.var_nombre,
                        self.var_cat):
                var.trace_add("write", lambda *_a: self._pedir_refresco())

        def _pedir_refresco(self):
            if self._debounce is not None:
                try:
                    self.after_cancel(self._debounce)
                except Exception:
                    pass
            self._debounce = self.after(140, self.refrescar)

        def _limpiar(self):
            for v in (self.var_q, self.var_marca, self.var_modelo, self.var_nombre):
                v.set("")
            self.var_cat.set("TODAS")

        def refrescar(self):
            self._debounce = None
            cat = self.var_cat.get()
            cat = None if cat == "TODAS" else cat
            inact = bool(self.var_inact.get()) if self.var_inact is not None else False
            self._resultados = fuzzy_search.buscar(
                self.var_q.get(),
                self.bd.listar_fichas(incluir_inactivas=inact),
                categoria=cat, marca=self.var_marca.get(),
                modelo=self.var_modelo.get(), nombre=self.var_nombre.get(),
                top_n=self.top_n, incluir_inactivas=inact)
            for i in self.tree.get_children():
                self.tree.delete(i)
            for idx, f in enumerate(self._resultados):
                sim = f.get("_similitud", 0)
                match = f"{int(sim * 100)}%" if sim else "—"
                modelo = (f.get("especificacion") or f.get("tipo_producto") or "").strip()
                if self.compacto:
                    vals = [bd_manager.BDManager.nombre_de(f), f.get("marca", ""),
                            f.get("categoria", ""), match]
                else:
                    vals = [bd_manager.BDManager.nombre_de(f), f.get("marca", ""),
                            modelo, f.get("dimensiones", ""),
                            f.get("categoria", ""), match]
                if self.permitir_inactivas:
                    estado = f.get("estado", "activo")
                    vals.append("desactivada" if estado != "activo" else "activa")
                self.tree.insert("", "end", iid=str(idx), values=vals)
            # La MEJOR coincidencia queda seleccionada de primera.
            if self._resultados:
                self.tree.selection_set("0")
                self.tree.focus("0")
                self.tree.see("0")
            n = len(self._resultados)
            tope = "" if n < self.top_n else f" · se muestran los primeros {self.top_n}"
            self.lbl_estado.configure(
                text=(f"{n} resultado(s){tope}" if n else
                      "Sin resultados — probá con menos texto o revisá los filtros"),
                text_color=(TEXTO_SUAVE if n else AMBAR))

        def ficha_seleccionada(self):
            sel = self.tree.selection()
            if not sel:
                return None
            try:
                return self._resultados[int(sel[0])]
            except (ValueError, IndexError):
                return None

        def _activar(self):
            f = self.ficha_seleccionada()
            if f and callable(self.on_activar):
                self.on_activar(f)


    class TablaMateriales(ctk.CTkFrame):
        """Tabla reutilizable de materiales seleccionados con agregar/editar/
        eliminar y renumeracion automatica de consecutivos."""

        def __init__(self, master, bd, materiales=None):
            super().__init__(master, fg_color="transparent")
            self.bd = bd
            self.materiales = list(materiales or [])
            self._build()
            self._refrescar()

        def _build(self):
            # Dos paneles lado a lado, como en el diseno: el catalogo a la
            # izquierda (mas ancho) y los materiales ya elegidos a la derecha.
            # Antes estaban apilados y obligaban a hacer scroll para ver lo que
            # se acababa de agregar.
            # Sin ``uniform``: con columnas uniformes, la mas angosta se estira
            # hasta igualar la proporcion de la mas ancha y la pantalla exigia
            # ~300 px mas de los que tiene un monitor de 1366.
            self.grid_rowconfigure(0, weight=1)
            self.grid_columnconfigure(0, weight=115)
            self.grid_columnconfigure(2, weight=100)

            izq = ctk.CTkFrame(self, fg_color="transparent")
            izq.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
            sep = ctk.CTkFrame(self, width=1, fg_color=BORDE)
            sep.grid(row=0, column=1, sticky="ns")
            der = ctk.CTkFrame(self, fg_color="transparent")
            der.grid(row=0, column=2, sticky="nsew", padx=(12, 0))
            # ``der`` queda expuesto: la pantalla de submittal cuelga ahi el
            # bloque de registro, debajo de la lista de materiales.
            self.der = der

            # ---------------------------------------------- panel: catalogo
            top = ctk.CTkFrame(izq, fg_color="transparent")
            top.pack(fill="x", pady=(0, 8))
            titulo = ctk.CTkFrame(top, fg_color="transparent")
            titulo.pack(side="left")
            ctk.CTkLabel(titulo, text="Catálogo de fichas", text_color=TEXTO,
                        font=_fuente(13, "bold")).pack(side="left")
            ctk.CTkLabel(titulo, text="  · doble clic para agregar",
                        text_color=TEXTO_TENUE, font=_fuente(11)).pack(side="left")
            # "Cargar ficha" queda como boton de icono (con globo de ayuda): el
            # panel es angosto y la accion completa esta en Inicio.
            b_nueva = _boton_secundario(top, "＋", self._cargar_ficha, ancho=30,
                                        alto=24)
            b_nueva.pack(side="right")
            _tooltip(b_nueva, "Cargar una ficha nueva a la Base de Datos")
            _enlace(top, "Actualizar", self._actualizar_catalogo).pack(
                side="right", padx=(0, 8))

            # Buscador reutilizable (mejor coincidencia primero + filtros + scroll)
            self.buscador = _BuscadorFichas(izq, self.bd, compacto=True,
                                            on_activar=self._agregar_ficha, alto_filas=8)
            self.buscador.pack(fill="both", expand=True)

            # ------------------------------------------- panel: materiales
            cab = ctk.CTkFrame(der, fg_color="transparent")
            cab.pack(fill="x", pady=(0, 8))
            ctk.CTkLabel(cab, text="Materiales del submittal", text_color=TEXTO,
                        font=_fuente(13, "bold")).pack(side="left")
            self.lbl_conteo = ctk.CTkLabel(cab, text="0 fichas", text_color=TEXTO_SUAVE,
                                           font=_mono(11))
            self.lbl_conteo.pack(side="right")

            marco, self.tree, pie = _tabla_ttk(
                der, ("cons", "nombre", "marca"),
                {"cons": "Consec.", "nombre": "Nombre", "marca": "Marca"},
                {"cons": 64, "nombre": 170, "marca": 96},
                alineados=("cons",), alto=7)
            marco.pack(fill="both", expand=True)
            _hover_filas(self.tree)
            self.lbl_vacio = ctk.CTkLabel(
                pie, text="Agregá fichas desde el catálogo de la izquierda",
                text_color=TEXTO_TENUE, font=_fuente(11))
            self.lbl_vacio.pack(anchor="w", padx=12, pady=4)

            bar = ctk.CTkFrame(der, fg_color="transparent")
            bar.pack(fill="x", pady=(8, 0))
            b_subir = _boton_secundario(bar, "↑", lambda: self._mover(-1), ancho=34,
                                        alto=30)
            b_subir.pack(side="left")
            b_bajar = _boton_secundario(bar, "↓", lambda: self._mover(1), ancho=34,
                                        alto=30)
            b_bajar.pack(side="left", padx=(6, 0))
            _tooltip(b_subir, "Subir el material dentro de su especialidad")
            _tooltip(b_bajar, "Bajar el material dentro de su especialidad")
            _boton_secundario(bar, "Marcas", self._editar, ancho=84,
                              alto=30).pack(side="left", padx=(6, 0))
            _boton_secundario(bar, "Ver PDF", self._vista_previa, ancho=86,
                              alto=30).pack(side="left", padx=(6, 0))
            _boton_peligro(bar, "Eliminar", self._eliminar, ancho=82, alto=30).pack(
                side="right")

        def _actualizar_catalogo(self):
            """Trae los cambios de otras PCs (git pull) y vuelve a correr la
            busqueda actual: cubre el caso de agregar una ficha nueva y que el
            buscador no la reconociera hasta cerrar y reabrir esta ventana."""
            self.bd.sincronizar()
            self.buscador.refrescar()

        def _agregar_ficha(self, ficha):
            cat = ficha["categoria"]
            cons = self._siguiente_consecutivo(cat)
            # El material hereda el nombre descriptivo SIN la marca (la marca ya
            # tiene su propia columna en carátulas y Excel). Así dos tubos del
            # mismo tipo no aparecen como dos filas idénticas.
            nombre = nomenclatura.nombre_sin_marca(
                bd_manager.BDManager.nombre_de(ficha), ficha.get("marca", ""))
            material = {
                "consecutivo": cons, "id_ficha_bd": ficha["id"],
                "nombre_material": nombre or ficha["nombre_material"],
                "marca": ficha["marca"],
                "categoria": cat, "marcas_alternativas": [], "justificacion_stock": False,
            }
            # ¿Este material tiene otras marcas por stock (misma especificación,
            # distinta marca) ya cargadas en la BD? Si las hay, se ofrece
            # adjuntarlas todas de una vez, sin tener que abrir "Editar material"
            # y buscarlas una por una. Al aceptar, quedan como marcas
            # alternativas con su ficha real y se activa la justificación por
            # stock (el texto legal de la carátula se genera solo al generar).
            try:
                familia = self.bd.fichas_misma_familia(ficha)
            except Exception:
                familia = []      # la detección nunca debe impedir agregar
            if familia:
                elegidas = self._preguntar_familia(ficha, familia)
                if elegidas:
                    material["marcas_alternativas"] = [
                        {"id_ficha_bd": fa["id"], "marca": fa.get("marca", "")}
                        for fa in elegidas]
                    material["justificacion_stock"] = True
            self.materiales.append(material)
            self._refrescar(resaltar=cons)

        def _preguntar_familia(self, ficha, familia):
            """Popup: ``ficha`` tiene N marcas por stock disponibles. Devuelve la
            lista de fichas alternativas elegidas (``[]`` si el usuario prefiere
            solo la marca principal o cierra la ventana)."""
            top = ctk.CTkToplevel(self)
            top.title("Varias marcas disponibles")
            top.configure(fg_color=GRIS_BG, padx=16, pady=14)
            top.grab_set()
            tarjeta = _tarjeta(top)
            tarjeta.pack(fill="both", expand=True, padx=4, pady=4)
            nombre = nomenclatura.nombre_sin_marca(
                bd_manager.BDManager.nombre_de(ficha), ficha.get("marca", ""))
            n_total = 1 + len(familia)
            ctk.CTkLabel(tarjeta, text=f"Este material tiene {n_total} marcas por stock",
                        font=_fuente(13, "bold"), text_color=GRIS_TEXTO).pack(
                anchor="w", padx=16, pady=(16, 2))
            ctk.CTkLabel(
                tarjeta,
                text=f"«{nombre}» está disponible con estas marcas. Se recomienda "
                     "incluirlas todas para no depender de una sola en la aprobación "
                     "(se adjunta la ficha de cada una y se justifica por stock).",
                text_color=GRIS_TEXTO_SUAVE, justify="left", wraplength=460).pack(
                anchor="w", padx=16, pady=(0, 10))

            # Marca principal: fija, siempre incluida (es la ficha que se agrega).
            fila_p = ctk.CTkFrame(tarjeta, fg_color="transparent")
            fila_p.pack(fill="x", padx=16, pady=2)
            ctk.CTkLabel(fila_p, text=f"✓  {ficha.get('marca', '(sin marca)')}",
                        font=_fuente(11, "bold"), text_color=AZUL_ES).pack(side="left")
            ctk.CTkLabel(fila_p, text="marca principal (siempre incluida)",
                        text_color=GRIS_TEXTO_SUAVE).pack(side="left", padx=8)

            variables = []
            for fa in familia:
                v = tk.BooleanVar(value=True)
                variables.append((v, fa))
                fila = ctk.CTkFrame(tarjeta, fg_color="transparent")
                fila.pack(fill="x", padx=16, pady=2)
                ctk.CTkCheckBox(fila, variable=v, text=fa.get("marca", "(sin marca)"),
                               font=_fuente(11, "bold"), text_color=GRIS_TEXTO,
                               fg_color=AZUL_ES, hover_color=_HOVER[AZUL_ES]).pack(
                    side="left")
                ctk.CTkLabel(fila, text="ficha en catálogo",
                            text_color=GRIS_TEXTO_SUAVE).pack(side="left", padx=8)

            resultado = {"v": []}

            def _agregar_todas():
                resultado["v"] = [fa for v, fa in variables if v.get()]
                top.destroy()

            def _solo_principal():
                resultado["v"] = []
                top.destroy()

            barra = ctk.CTkFrame(tarjeta, fg_color="transparent")
            barra.pack(pady=(14, 16))
            _boton(barra, "Agregar las marcas marcadas", _agregar_todas,
                  color=NARANJA_CTA, ancho=250).pack(side="left", padx=6)
            _boton_secundario(barra, "Solo esta marca", _solo_principal,
                             ancho=150).pack(side="left", padx=6)
            # Cerrar con la X = no agregar alternativas (equivale a "solo esta
            # marca"): el usuario ya pidió agregar el material, no se pierde.
            top.protocol("WM_DELETE_WINDOW", _solo_principal)
            _dimensionar_ventana(top, 520, 360)
            _traer_al_frente(top)
            self.wait_window(top)
            return resultado["v"]

        def _siguiente_consecutivo(self, cat):
            nums = [int(re.match(rf"{cat}(\d+)", m["consecutivo"]).group(1))
                    for m in self.materiales
                    if re.match(rf"{cat}(\d+)", m.get("consecutivo", ""))]
            return f"{cat}{(max(nums) + 1 if nums else 1):02d}"

        def _renumerar(self):
            # Se numera respetando el ORDEN ACTUAL de self.materiales (no el
            # consecutivo previo) para que "Subir"/"Bajar" puedan decidir
            # manualmente que material ocupa cada numero dentro de su categoria.
            contadores = {}
            for m in self.materiales:
                cat = m["categoria"]
                contadores[cat] = contadores.get(cat, 0) + 1
                m["consecutivo"] = f"{cat}{contadores[cat]:02d}"

        def _refrescar(self, resaltar=None):
            self._renumerar()
            self.materiales.sort(key=lambda x: bd_manager._clave_orden(x["consecutivo"]))
            for i in self.tree.get_children():
                self.tree.delete(i)
            for m in self.materiales:
                marca = bd_manager._marcas_material(m, {})
                self.tree.insert("", "end", iid=m["consecutivo"],
                                 values=(m["consecutivo"], m["nombre_material"], marca))
            n = len(self.materiales)
            self.lbl_conteo.configure(text=f"{n} ficha" if n == 1 else f"{n} fichas")
            # El pie de la tabla solo dice que hacer cuando esta vacia; con
            # materiales adentro estorba.
            if n:
                self.lbl_vacio.pack_forget()
            else:
                self.lbl_vacio.pack(anchor="w", padx=12, pady=4)
            if resaltar and self.tree.exists(resaltar):
                self.tree.selection_set(resaltar)
                self.tree.see(resaltar)
                self._animar_ingreso(resaltar)

        def _animar_ingreso(self, iid):
            """Resalta en verde el renglon recien agregado y lo desvanece a
            blanco en unos pocos pasos, para dar feedback visual inmediato de
            que el material entro a la lista del submittal."""
            # Se desvanece hacia el tinte de acento del diseno (no verde: el
            # verde en esta paleta significa "sincronizado/ok", no "nuevo").
            colores = [CHIP_BG, "#FBEEEC", FILA_HOVER, SUPERFICIE]
            tag = f"nuevo_{iid}"

            def paso(i=0):
                if not self.tree.exists(iid) or i >= len(colores):
                    return
                self.tree.item(iid, tags=(tag,))
                self.tree.tag_configure(tag, background=colores[i])
                self.after(110, lambda: paso(i + 1))
            paso()

        def _mover(self, direccion):
            """Sube/baja el material seleccionado un puesto DENTRO DE SU MISMA
            categoria (ej. mover el cemento de ESTR04 a ESTR06), permitiendo
            elegir manualmente el orden de los consecutivos por trazabilidad."""
            m = self._sel_material()
            if not m:
                return
            cat = m["categoria"]
            pos_categoria = [i for i, x in enumerate(self.materiales) if x["categoria"] == cat]
            idx = self.materiales.index(m)
            pos = pos_categoria.index(idx)
            nueva_pos = pos + direccion
            if not (0 <= nueva_pos < len(pos_categoria)):
                return
            otro_idx = pos_categoria[nueva_pos]
            self.materiales[idx], self.materiales[otro_idx] = (
                self.materiales[otro_idx], self.materiales[idx])
            self._refrescar()
            self.tree.selection_set(m["consecutivo"])
            self.tree.see(m["consecutivo"])

        def _sel_material(self):
            sel = self.tree.selection()
            if not sel:
                return None
            return next((m for m in self.materiales if m["consecutivo"] == sel[0]), None)

        def _editar(self):
            m = self._sel_material()
            if not m:
                return
            ficha = self.bd.obtener_ficha(m.get("id_ficha_bd")) or {}
            # ``marcas_alternativas`` acepta el formato viejo (string suelto,
            # solo texto) y el nuevo (dict vinculado a una ficha real de la BD,
            # cuyo PDF se adjunta de verdad al compilado). Se normaliza al abrir
            # para poder editar ambos casos con la misma lista.
            alternativas = []
            for a in m.get("marcas_alternativas", []) or []:
                if isinstance(a, dict):
                    alternativas.append(dict(a))
                else:
                    alternativas.append({"id_ficha_bd": None, "marca": str(a)})

            # Valores por defecto (los que trae la ficha del catalogo, o el
            # texto auto-generado de "aspectos"), para poder detectar al
            # guardar si el usuario en verdad escribio algo distinto (override
            # para ESTE submittal) o si dejo el campo tal cual venia.
            desc_defecto = ficha.get("descripcion_corta", "")
            norm_defecto = ficha.get("normativa", "")

            def _aspectos_auto():
                temp = {"marcas_alternativas": alternativas, "justificacion_stock": v_s.get()}
                return bd_manager._texto_aspectos(temp, ficha) or ficha.get("aspectos_adicionales", "")

            top = ctk.CTkToplevel(self)
            top.title("Editar material"); top.grab_set()
            top.configure(fg_color=GRIS_BG, padx=14, pady=14)
            top.geometry("820x820")
            tarjeta = _tarjeta(top); tarjeta.pack(fill="both", expand=True)
            tarjeta.grid_columnconfigure(1, weight=1)
            # ``v_s`` (justificar por stock) se define ACA, antes de armar los
            # campos: ``_aspectos_auto()`` la lee al pre-llenar "Aspectos
            # adicionales" (mas abajo), y crearla despues provocaba un
            # NameError al abrir el diálogo sobre un material sin override de
            # aspectos. Su checkbox se dibuja mas abajo reutilizando esta misma
            # variable.
            v_s = tk.BooleanVar(value=m.get("justificacion_stock", False))
            ctk.CTkLabel(tarjeta, text=f"{m['consecutivo']} — texto de la carátula",
                        font=_fuente(11, "bold"), text_color=GRIS_TEXTO).grid(
                row=0, column=0, columnspan=3, padx=16, pady=(16, 10), sticky="w")

            ctk.CTkLabel(tarjeta, text="Nombre del material:", text_color=GRIS_TEXTO).grid(
                row=1, column=0, sticky="e", pady=4, padx=(16, 6))
            v_nombre = tk.StringVar(value=m["nombre_material"])
            ctk.CTkEntry(tarjeta, textvariable=v_nombre, height=32,
                        corner_radius=8, border_color=BORDE_SUAVE).grid(
                row=1, column=1, columnspan=2, padx=(0, 16), sticky="we")

            ctk.CTkLabel(tarjeta, text="Marca principal:", text_color=GRIS_TEXTO).grid(
                row=2, column=0, sticky="e", pady=4, padx=(16, 6))
            v_p = tk.StringVar(value=m["marca"])
            ctk.CTkEntry(tarjeta, textvariable=v_p, height=32,
                        corner_radius=8, border_color=BORDE_SUAVE).grid(
                row=2, column=1, columnspan=2, padx=(0, 16), sticky="we")

            ctk.CTkLabel(tarjeta, text="Marcas alternativas\n(con ficha adjunta):",
                        text_color=GRIS_TEXTO, justify="right").grid(
                row=3, column=0, sticky="ne", pady=4, padx=(16, 6))
            lst_alt = tk.Listbox(tarjeta, height=5, width=34, font=(FUENTE, 10),
                                 bg=SUPERFICIE, fg=TEXTO_2, selectbackground=CHIP_BG,
                                 selectforeground=ACENTO_TXT, relief="flat",
                                 borderwidth=0, highlightthickness=1,
                                 highlightbackground=BORDE, highlightcolor=ACENTO_BORDE,
                                 activestyle="none")
            lst_alt.grid(row=3, column=1, padx=(0, 6), pady=4, sticky="we")

            def _pintar_alt():
                lst_alt.delete(0, "end")
                for a in alternativas:
                    nombre = a.get("marca") or "(sin nombre)"
                    if not a.get("id_ficha_bd"):
                        nombre += "  (solo texto, sin PDF)"
                    lst_alt.insert("end", nombre)
            _pintar_alt()

            def _quitar_alt():
                sel = lst_alt.curselection()
                if sel:
                    alternativas.pop(sel[0]); _pintar_alt()
            _boton_secundario(tarjeta, "Quitar", _quitar_alt, ancho=90).grid(
                row=3, column=2, padx=(0, 16), pady=4, sticky="n")

            ctk.CTkLabel(tarjeta, text="Buscar en catálogo para\nagregar como alternativa:",
                        text_color=GRIS_TEXTO, justify="right").grid(
                row=4, column=0, sticky="ne", pady=(10, 4), padx=(16, 6))
            v_busq_alt = tk.StringVar()
            ctk.CTkEntry(tarjeta, textvariable=v_busq_alt, height=32,
                        corner_radius=8, border_color=BORDE_SUAVE).grid(
                row=4, column=1, columnspan=2, padx=(0, 16), pady=(10, 4), sticky="we")
            # Resultados de la busqueda: en un contenedor con scroll horizontal
            # (los nombres de fichas -- tubos, tuberia -- son largos y antes se
            # cortaban) y VERTICAL, mas alto, y estirandose con la ventana para
            # ver la descripcion completa de cada producto.
            cont_busq = ctk.CTkFrame(tarjeta, fg_color="transparent")
            cont_busq.grid(row=5, column=1, columnspan=2, padx=(0, 16),
                           pady=(0, 4), sticky="nsew")
            cont_busq.grid_rowconfigure(0, weight=1)
            cont_busq.grid_columnconfigure(0, weight=1)
            lst_busq_alt = tk.Listbox(cont_busq, height=8, font=(FUENTE, 10),
                                      bg=SUPERFICIE, fg=TEXTO_2, selectbackground=CHIP_BG,
                                      selectforeground=ACENTO_TXT, relief="flat",
                                      borderwidth=0, highlightthickness=1,
                                      highlightbackground=BORDE,
                                      highlightcolor=ACENTO_BORDE, activestyle="none")
            lst_busq_alt.grid(row=0, column=0, sticky="nsew")
            _sb_v = tk.Scrollbar(cont_busq, orient="vertical", command=lst_busq_alt.yview)
            _sb_v.grid(row=0, column=1, sticky="ns")
            _sb_h = tk.Scrollbar(cont_busq, orient="horizontal", command=lst_busq_alt.xview)
            _sb_h.grid(row=1, column=0, sticky="ew")
            lst_busq_alt.configure(yscrollcommand=_sb_v.set, xscrollcommand=_sb_h.set)
            # La fila de resultados es la que crece cuando la ventana se agranda.
            tarjeta.grid_rowconfigure(5, weight=1)
            sug_alt = []

            def _buscar_alt(*_ev):
                q = v_busq_alt.get().strip()
                lst_busq_alt.delete(0, "end")
                sug_alt.clear()
                if not q:
                    return
                ya_puestas = {m.get("id_ficha_bd")} | {a.get("id_ficha_bd") for a in alternativas}
                for f in self.bd.buscar(q, categoria=m.get("categoria")):
                    if f["id"] in ya_puestas:
                        continue
                    sug_alt.append(f)
                    lst_busq_alt.insert(
                        "end", f"{bd_manager.BDManager.nombre_de(f)}  ·  {f.get('marca', '')}")
            v_busq_alt.trace_add("write", lambda *_ev: _buscar_alt())

            def _agregar_alt(_ev=None):
                sel = lst_busq_alt.curselection()
                if not sel or sel[0] >= len(sug_alt):
                    return
                f = sug_alt[sel[0]]
                alternativas.append({"id_ficha_bd": f["id"], "marca": f.get("marca", "")})
                v_busq_alt.set(""); _buscar_alt(); _pintar_alt()
            lst_busq_alt.bind("<Double-Button-1>", _agregar_alt)

            ctk.CTkLabel(tarjeta, text="Descripción técnica\n(carátula):",
                        text_color=GRIS_TEXTO, justify="right").grid(
                row=6, column=0, sticky="ne", pady=(10, 4), padx=(16, 6))
            txt_desc = ctk.CTkTextbox(tarjeta, height=70, corner_radius=8,
                                      border_width=1, border_color=BORDE_SUAVE)
            txt_desc.grid(row=6, column=1, columnspan=2, padx=(0, 16), pady=(10, 4), sticky="we")
            txt_desc.insert("1.0", m["descripcion"] if "descripcion" in m else desc_defecto)

            ctk.CTkLabel(tarjeta, text="Normativa (carátula):", text_color=GRIS_TEXTO).grid(
                row=7, column=0, sticky="e", pady=4, padx=(16, 6))
            v_norm = tk.StringVar(value=m["normativa"] if "normativa" in m else norm_defecto)
            ctk.CTkEntry(tarjeta, textvariable=v_norm, height=32,
                        corner_radius=8, border_color=BORDE_SUAVE).grid(
                row=7, column=1, columnspan=2, padx=(0, 16), sticky="we")

            ctk.CTkLabel(tarjeta, text="Aspectos adicionales /\nnotas (carátula):",
                        text_color=GRIS_TEXTO, justify="right").grid(
                row=8, column=0, sticky="ne", pady=(10, 4), padx=(16, 6))
            txt_aspectos = ctk.CTkTextbox(tarjeta, height=90, corner_radius=8,
                                         border_width=1, border_color=BORDE_SUAVE)
            txt_aspectos.grid(row=8, column=1, columnspan=2, padx=(0, 16), pady=(10, 4), sticky="we")
            txt_aspectos.insert(
                "1.0", m["aspectos_adicionales"] if "aspectos_adicionales" in m else _aspectos_auto())

            def _recalcular_aspectos():
                txt_aspectos.delete("1.0", "end")
                txt_aspectos.insert("1.0", _aspectos_auto())
            _boton_secundario(tarjeta, "Recalcular automático", _recalcular_aspectos,
                              ancho=180).grid(row=9, column=1, columnspan=2, padx=(0, 16),
                                              pady=(0, 4), sticky="w")

            # (``v_s`` ya se definió arriba, antes de "Aspectos adicionales".)
            ctk.CTkCheckBox(tarjeta, text="Justificar por stock (marcas alternativas aprobadas)",
                           variable=v_s, text_color=GRIS_TEXTO,
                           fg_color=AZUL_ES, hover_color=_HOVER[AZUL_ES]).grid(
                row=10, column=0, columnspan=3, sticky="w", padx=16, pady=4)

            def _guardar_local():
                """Guarda los campos como texto/overrides de ESTE material en
                ESTE submittal unicamente (no toca la ficha del catalogo)."""
                m["nombre_material"] = v_nombre.get().strip() or m["nombre_material"]
                m["marca"] = v_p.get().strip()
                m["marcas_alternativas"] = alternativas
                m["justificacion_stock"] = v_s.get()

                desc_final = txt_desc.get("1.0", "end-1c").strip()
                if desc_final == desc_defecto.strip():
                    m.pop("descripcion", None)
                else:
                    m["descripcion"] = desc_final

                norm_final = v_norm.get().strip()
                if norm_final == norm_defecto.strip():
                    m.pop("normativa", None)
                else:
                    m["normativa"] = norm_final

                aspectos_final = txt_aspectos.get("1.0", "end-1c").strip()
                if aspectos_final == _aspectos_auto().strip():
                    m.pop("aspectos_adicionales", None)
                else:
                    m["aspectos_adicionales"] = aspectos_final

            def _guardar_solo_proyecto():
                _guardar_local()
                top.destroy(); self._refrescar()

            def _guardar_en_bd():
                if not ficha.get("id"):
                    messagebox.showerror(
                        "Sin ficha en la BD",
                        "Este material no tiene una ficha vinculada en la Base "
                        "de Datos (fue agregado de otra forma); no se puede "
                        "guardar ahi. Use 'Guardar solo este proyecto'.",
                        parent=top)
                    return
                if not messagebox.askyesno(
                        "Guardar en la Base de Datos",
                        "Esto modifica la ficha del catálogo:\n"
                        f"{bd_manager.BDManager.nombre_de(ficha)}\n\n"
                        "El cambio se aplicará a TODOS los submittals (los que "
                        "ya existen y los futuros) que usen esta ficha, no "
                        "solo a este proyecto.\n\n¿Continuar?", parent=top):
                    return
                cambios_bd = {
                    "nombre_material": v_nombre.get().strip(),
                    "marca": v_p.get().strip(),
                    "descripcion_corta": txt_desc.get("1.0", "end-1c").strip(),
                    "normativa": v_norm.get().strip(),
                    "aspectos_adicionales": txt_aspectos.get("1.0", "end-1c").strip(),
                }
                try:
                    ficha_act = self.bd.editar_ficha(ficha["id"], cambios_bd,
                                                     regenerar_nombre=False)
                except Exception as e:
                    messagebox.showerror("No se pudo guardar en la BD", str(e), parent=top)
                    return
                r = self.bd.git_push(
                    f"editar ficha {ficha_act.get('nombre_ficha', ficha['id'])} desde submittal")
                # El valor recien guardado en la ficha YA es el nuevo default:
                # se limpian los overrides locales de descripcion/normativa/
                # aspectos para que este submittal (y los futuros) sigan el
                # valor de la ficha en vez de quedar pisado con una copia vieja.
                m["nombre_material"] = cambios_bd["nombre_material"] or m["nombre_material"]
                m["marca"] = cambios_bd["marca"]
                m["marcas_alternativas"] = alternativas
                m["justificacion_stock"] = v_s.get()
                m.pop("descripcion", None)
                m.pop("normativa", None)
                m.pop("aspectos_adicionales", None)
                top.destroy(); self._refrescar()
                _avisar_resultado_git(
                    self.winfo_toplevel(), r,
                    f"Ficha actualizada en la Base de Datos:\n"
                    f"{ficha_act.get('nombre_ficha', '')}\n\n"
                    "Afecta a todos los submittals que usen esta ficha.")

            barra_guardar = ctk.CTkFrame(tarjeta, fg_color="transparent")
            barra_guardar.grid(row=11, column=0, columnspan=3, pady=(12, 16))
            _boton_secundario(barra_guardar, "Guardar solo este proyecto",
                             _guardar_solo_proyecto, ancho=220).pack(side="left", padx=6)
            _boton(barra_guardar, "Guardar en la BD (todos los proyectos)",
                  _guardar_en_bd, color=NARANJA_CTA, ancho=280).pack(side="left", padx=6)
            _dimensionar_ventana(top, 820, 820)
            _traer_al_frente(top)

        def _eliminar(self):
            m = self._sel_material()
            if m and messagebox.askyesno("Eliminar", f"¿Quitar {m['consecutivo']} — {m['nombre_material']}?", parent=self.winfo_toplevel()):
                self.materiales.remove(m); self._refrescar()

        def _vista_previa(self):
            m = self._sel_material()
            if not m:
                messagebox.showinfo("Seleccione un material",
                                    "Elija un material de la lista.",
                                    parent=self.winfo_toplevel())
                return
            ficha = self.bd.obtener_ficha(m.get("id_ficha_bd"))
            if not ficha:
                messagebox.showerror("Ficha no encontrada",
                                     "La ficha original ya no está en la Base de Datos.",
                                     parent=self.winfo_toplevel())
                return
            _vista_previa_ficha(self, self.bd, ficha)

        def _cargar_ficha(self):
            # Al terminar se refresca el catalogo para que la ficha recien
            # cargada aparezca sin cerrar y reabrir la pantalla. (Antes esto
            # llamaba a ``self._sugerir()``, un metodo que no existe: cargar una
            # ficha desde aca lanzaba AttributeError al cerrar la ventana.)
            VentanaCargarFicha(self.winfo_toplevel(), self.bd,
                               al_terminar=self.buscador.refrescar)


    class VentanaCargarFicha(ctk.CTkToplevel):
        """Flujo 3: cargar una o varias fichas a la BD (con extraccion OCR/IA)."""

        def __init__(self, master, bd, al_terminar=None):
            super().__init__(master)
            self.bd = bd; self.al_terminar = al_terminar
            self._cancelado = False
            self._cerrada = False
            self._procesando = False
            self._prog_max = 1
            self.title("Cargar ficha(s) a la BD")
            self.configure(fg_color=GRIS_BG, padx=16, pady=16)
            self.grab_set()
            tarjeta = _tarjeta(self); tarjeta.pack(fill="both", expand=True, padx=4, pady=4)
            ctk.CTkLabel(tarjeta, text="Cargar fichas técnicas a la Base de Datos",
                        font=_fuente(13, "bold"), text_color=GRIS_TEXTO).pack(
                anchor="w", padx=18, pady=(18, 0))
            botones = ctk.CTkFrame(tarjeta, fg_color="transparent")
            botones.pack(anchor="w", padx=18, pady=8)
            # Una sola accion primaria (solida) por dialogo: las otras dos van
            # como secundaria y como destructiva de contorno.
            self.btn_archivos = _boton(botones, "Seleccionar archivo(s)…",
                                       self._seleccionar, ancho=180)
            self.btn_archivos.pack(side="left")
            self.btn_carpetas = _boton_secundario(botones, "Seleccionar carpeta(s)…",
                                                  self._seleccionar_carpetas,
                                                  ancho=180, alto=34)
            self.btn_carpetas.pack(side="left", padx=(8, 0))
            self.btn_cancelar = _boton_peligro(botones, "Cancelar extracción",
                                               self._cancelar, ancho=160, alto=34)
            self.btn_cancelar.configure(state="disabled")
            self.btn_cancelar.pack(side="left", padx=(8, 0))
            self.prog = ctk.CTkProgressBar(tarjeta, height=10, corner_radius=5,
                                           progress_color=AZUL_ES)
            self.prog.set(0)
            self.prog.pack(fill="x", padx=18, pady=4)
            self.txt = _consola(tarjeta, height=280)
            self.txt.pack(fill="both", expand=True, padx=18, pady=(4, 8))
            self.btn_cerrar = _boton_secundario(tarjeta, "Cerrar", self._on_close, ancho=120)
            self.btn_cerrar.pack(pady=(0, 18))
            self.protocol("WM_DELETE_WINDOW", self._on_close)
            _dimensionar_ventana(self, 760, 620)
            _traer_al_frente(self)

        def _on_close(self):
            if self._procesando:
                if not messagebox.askyesno(
                        "Cancelar extracción",
                        "Hay una extracción en curso. ¿Cancelarla y cerrar la ventana?", parent=self.winfo_toplevel()):
                    return
                self._cancelado = True
            self._cerrada = True
            self.destroy()

        def _cancelar(self):
            self._cancelado = True
            self.btn_cancelar.configure(state="disabled")
            self._log("\n⛔ Cancelando… se detendrá después del archivo en curso.")

        def _log(self, msg):
            self.txt.insert("end", msg + "\n"); self.txt.see("end"); self.update_idletasks()

        EXTENSIONES = (".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff")

        def _seleccionar(self):
            rutas = filedialog.askopenfilenames(
                title="Fichas técnicas",
                filetypes=[("Fichas", "*.pdf *.png *.jpg *.jpeg *.tif *.tiff"), ("Todos", "*.*")])
            if not rutas:
                return
            self._procesar_archivos(rutas)

        def _seleccionar_carpetas(self):
            """Elige varias carpetas (una a la vez) y extrae todas las fichas
            técnicas que encuentre dentro de ellas, incluidas subcarpetas."""
            carpetas = []
            while True:
                titulo = "Seleccione una carpeta con fichas técnicas"
                if carpetas:
                    titulo += f"  (ya elegidas: {len(carpetas)} — Cancelar para terminar)"
                c = filedialog.askdirectory(title=titulo)
                if not c:
                    break
                carpetas.append(c)

            if not carpetas:
                return

            rutas = []
            for c in carpetas:
                rutas.extend(str(p) for p in Path(c).rglob("*")
                             if p.is_file() and p.suffix.lower() in self.EXTENSIONES)

            if not rutas:
                messagebox.showinfo(
                    "Sin fichas",
                    f"No se encontraron archivos de ficha (PDF/imagen) en "
                    f"{len(carpetas)} carpeta(s) seleccionada(s).", parent=self.winfo_toplevel())
                return

            self._log(f"📁 {len(carpetas)} carpeta(s) escaneada(s) — "
                      f"{len(rutas)} ficha(s) encontrada(s).")
            self._procesar_archivos(rutas)

        def _procesar_archivos(self, rutas):
            self._rutas = list(rutas)
            self._indice = 0
            self._ok = self._adv = self._fallo = 0
            self._cancelado = False
            self._procesando = True
            self._prog_max = len(self._rutas) or 1
            self.prog.set(0)
            self.btn_archivos.configure(state="disabled")
            self.btn_carpetas.configure(state="disabled")
            self.btn_cancelar.configure(state="normal")
            self._procesar_siguiente()

        def _procesar_siguiente(self):
            if self._cerrada:
                return
            if self._cancelado or self._indice >= len(self._rutas):
                self._finalizar_lote()
                return

            r = self._rutas[self._indice]
            i = self._indice + 1
            self._indice += 1
            self._log(f"[{i}/{len(self._rutas)}] {Path(r).name} — extrayendo…")

            def trabajo():
                try:
                    datos, error = ocr_extractor.extraer([r]), None
                except Exception as e:
                    datos, error = None, e
                if self._cerrada:
                    return
                try:
                    self.after(0, lambda: self._tras_extraer(r, i, datos, error))
                except tk.TclError:
                    pass

            threading.Thread(target=trabajo, daemon=True).start()

        def _tras_extraer(self, r, i, datos, error):
            if self._cerrada:
                return
            if self._cancelado:
                self._finalizar_lote()
                return

            if error is not None:
                self._fallo += 1
                self._log(f"   ❌ {error}")
            else:
                res = self._preview(r, datos)
                accion = res.get("accion", "omitir")

                if accion == "omitir":
                    self._log("   (omitido)")
                elif accion == "usar_existente":
                    f = res["ficha"]
                    self._log(f"   ↩️ ya existía: {self.bd.nombre_de(f)}")
                elif accion == "reemplazar":
                    f = res["ficha"]
                    try:
                        self.bd.reemplazar_pdf_ficha(f["id"], r)
                        self._ok += 1
                        self._log(f"   🔁 PDF reemplazado en {self.bd.nombre_de(f)}")
                    except Exception as e:
                        self._fallo += 1; self._log(f"   ❌ {e}")
                else:
                    d = res.get("datos") or {}
                    try:
                        ficha = self.bd.agregar_ficha(r, d)
                        if d.get("_requiere_manual"):
                            self._adv += 1
                            self._log(f"   ⚠️ cargada (revisada a mano): {ficha['nombre_ficha']}")
                        else:
                            self._ok += 1
                            self._log(f"   ✅ {ficha['nombre_ficha']}")
                    except Exception as e:
                        self._fallo += 1; self._log(f"   ❌ {e}")

            self.prog.set(i / self._prog_max)
            self._procesar_siguiente()

        def _finalizar_lote(self):
            self._procesando = False
            self.btn_archivos.configure(state="normal")
            self.btn_carpetas.configure(state="normal")
            self.btn_cancelar.configure(state="disabled")
            if self._cancelado:
                self._log(f"\n⛔ Extracción cancelada por el usuario "
                          f"({self._indice}/{len(self._rutas)} archivo(s) revisados).")
            self._log(f"\nResultado: ✅ {self._ok} ok · ⚠️ {self._adv} advertencias · "
                      f"❌ {self._fallo} fallos")
            if self._ok + self._adv:
                self._subir(self._ok + self._adv)
            if self.al_terminar:
                self.al_terminar()

        def _subir(self, cantidad):
            """Sube las fichas nuevas a GitHub (Flujo 2, paso 6)."""
            self._log("\n🔄 Subiendo a GitHub…")
            r = self.bd.git_push(f"agregar {cantidad} ficha(s) desde "
                                 f"{bd_manager.socket.gethostname()}")
            if r.get("desactivado"):
                self._log("   (sincronización desactivada: BD local)")
            elif r.get("subido"):
                extra = (f" · {r['conflictos']} conflicto(s) resuelto(s)"
                         if r.get("conflictos") else "")
                self._log(f"   ✅ Cambios subidos a GitHub{extra}")
            elif r.get("offline"):
                self._log("   📡 Sin conexión: las fichas quedaron guardadas y se "
                          "subirán al reconectar")
            elif r.get("auth"):
                self._log("   🔑 Falta el token de GitHub: use '⚙️ Configuración' "
                          "en la ventana principal")
            elif r.get("nada_que_subir"):
                self._log("   (sin cambios por subir)")
            else:
                self._log(f"   ⚠️ No se pudo subir: {r.get('error', 'error desconocido')}")

        def _preview(self, ruta, datos):
            """Abre el formulario de revisión y devuelve la acción elegida."""
            d = DialogoRevisarFicha(self, self.bd, ruta, datos)
            self.wait_window(d)
            return d.resultado


    class DialogoRevisarFicha(ctk.CTkToplevel):
        """Revisión de una ficha antes de guardarla (Mejora 1, paso 4).

        Muestra el NOMBRE AUTO-GENERADO arriba, editable, y lo recalcula solo
        mientras el usuario no lo escriba a mano. Antes de guardar:

          * exige los campos obligatorios;
          * exige que el nombre sea distinguible (dimensiones, presentación o
            modelo), para que no vuelvan a entrar fichas como "Tubería
            Estructural" que no se pueden diferenciar;
          * si el nombre ya existe en la BD, ofrece usar la ficha existente,
            reemplazar su PDF o guardar esta como variante.
        """

        CAMPOS = [
            ("nombre_material", "Nombre del material"),
            ("marca", "Marca"),
            ("categoria", "Categoría (ARQ/ESTR/MEC/ELEC)"),
            ("dimensiones", "Dimensiones (ej: 8\"x8\"x3/16\", 60x60 cm, o "
                            "\"MULTIPLE\" si la ficha cubre varias medidas)"),
            ("tipo_producto", "Tipo / forma (ej: cuadrado, rectangular)"),
            ("especificacion", "Especificación / modelo (ej: CH 13, QO260)"),
            ("normativa", "Normativa (no entra en el nombre)"),
            ("descripcion_corta", "Descripción corta / presentación"),
            ("aspectos_adicionales", "Notas adicionales para la carátula (opcional; "
                                     "ej: ficha de un sistema completo)"),
            ("sinonimos", "Sinónimos para la búsqueda (coma; ej: interruptor "
                          "-> \"apagador, breaker\")"),
        ]

        def __init__(self, master, bd, ruta, datos, titulo=None, es_edicion=False):
            super().__init__(master)
            self.bd = bd
            self.datos_origen = dict(datos or {})
            self.es_edicion = es_edicion
            self.resultado = {"accion": "omitir"}
            self._nombre_auto = ""
            self._nombre_editado = bool(self.datos_origen.get("nombre_ficha_manual"))
            self.title(titulo or f"Revisar: {Path(ruta).name if ruta else 'ficha'}")
            self.configure(fg_color=GRIS_BG, padx=16, pady=14)
            self.grab_set()
            self._construir()
            self._recalcular()
            _dimensionar_ventana(self, 720, 740)
            _traer_al_frente(self)

        # ------------------------------------------------------------ armado
        def _construir(self):
            tarjeta = _tarjeta(self)
            tarjeta.pack(fill="both", expand=True, padx=4, pady=4)
            tarjeta.grid_columnconfigure(1, weight=1)
            pad = dict(padx=16)
            fila = 0
            if not self.es_edicion:
                metodo = self.datos_origen.get("_metodo", "?")
                manual = self.datos_origen.get("_requiere_manual")
                ctk.CTkLabel(tarjeta, text=f"Método de extracción: {metodo}"
                            + ("  ·  revise los datos" if manual else ""),
                            justify="left",
                            text_color=(ROJO_ES if manual else VERDE_OK)).grid(
                    row=fila, column=0, columnspan=3, sticky="w", **pad, pady=(16, 0))
                fila += 1

            ctk.CTkLabel(tarjeta, text="Nombre de la ficha (se genera solo):",
                        font=_fuente(11, "bold"), text_color=GRIS_TEXTO).grid(
                row=fila, column=0, columnspan=3, sticky="w", **pad, pady=(16, 2))
            fila += 1
            # Si la ficha ya trae un nombre escrito a mano, se muestra tal cual
            # (no se regenera solo: para eso está el botón Regenerar).
            self.v_nombre = tk.StringVar(
                value=str(self.datos_origen.get("nombre_ficha", "") or ""))
            e = ctk.CTkEntry(tarjeta, textvariable=self.v_nombre, height=34,
                             corner_radius=8, border_color=BORDE_SUAVE,
                             font=_fuente(11, "bold"), text_color=AZUL_ES)
            e.grid(row=fila, column=0, columnspan=2, sticky="we", padx=(16, 6))
            e.bind("<KeyRelease>", self._nombre_a_mano)
            _boton_secundario(tarjeta, "Regenerar", self._regenerar, ancho=120).grid(
                row=fila, column=2, padx=(0, 16))
            fila += 1

            self.lbl_aviso = ctk.CTkLabel(tarjeta, text="", justify="left",
                                          wraplength=560, text_color=GRIS_TEXTO)
            self.lbl_aviso.grid(row=fila, column=0, columnspan=3, sticky="w",
                                padx=16, pady=(4, 8))
            fila += 1

            self.vars = {}
            for clave, etiqueta in self.CAMPOS:
                ctk.CTkLabel(tarjeta, text=etiqueta + ":", text_color=GRIS_TEXTO).grid(
                    row=fila, column=0, sticky="e", pady=3, padx=(16, 6))
                v = tk.StringVar(value=str(self.datos_origen.get(clave, "") or ""))
                ent = ctk.CTkEntry(tarjeta, textvariable=v, height=32,
                                   corner_radius=8, border_color=BORDE_SUAVE)
                ent.grid(row=fila, column=1, columnspan=2, sticky="we", pady=3,
                        padx=(0, 16))
                ent.bind("<KeyRelease>", lambda _ev: self._recalcular())
                self.vars[clave] = v
                fila += 1
                if clave == "dimensiones":
                    self.v_sin_medidas = tk.BooleanVar(
                        value=bool(self.datos_origen.get("sin_medidas")))
                    ctk.CTkCheckBox(
                        tarjeta, variable=self.v_sin_medidas, command=self._recalcular,
                        text="Este material no tiene medidas/dimensiones "
                             "(ej: sacos, cubetas, unidades).",
                        text_color=GRIS_TEXTO, fg_color=AZUL_ES,
                        hover_color=_HOVER[AZUL_ES]
                    ).grid(row=fila, column=1, columnspan=2, sticky="w", padx=(0, 16))
                    fila += 1

            barra = ctk.CTkFrame(tarjeta, fg_color="transparent")
            barra.grid(row=fila, column=0, columnspan=3, pady=(12, 16))
            self.btn_ok = _boton(barra, "Confirmar y guardar", self._confirmar, ancho=190)
            self.btn_ok.pack(side="left", padx=6)
            _boton_secundario(barra, "Cancelar" if self.es_edicion else "Omitir",
                             self.destroy, ancho=120).pack(side="left", padx=6)

        # -------------------------------------------------------- nomenclatura
        def _datos_actuales(self):
            d = dict(self.datos_origen)
            d.update({k: v.get().strip() for k, v in self.vars.items()})
            d["categoria"] = d.get("categoria", "").strip().upper()
            d["sin_medidas"] = bool(self.v_sin_medidas.get())
            return d

        def _nombre_a_mano(self, _ev=None):
            """Si el usuario CAMBIA el nombre, deja de regenerarse solo.

            Se compara contra el último valor autogenerado en vez de reaccionar a
            cualquier tecla: con una flecha, un Tab o un Ctrl se marcaba como
            "manual" un nombre que en realidad no se había tocado, y entonces ya
            no se volvía a regenerar nunca.
            """
            self._nombre_editado = self.v_nombre.get().strip() != self._nombre_auto
            self._recalcular(solo_aviso=True)

        def _regenerar(self):
            self._nombre_editado = False
            self._recalcular()

        def _recalcular(self, solo_aviso=False):
            d = self._datos_actuales()
            self.analisis = nomenclatura.analizar(d)
            self._nombre_auto = self.analisis["nombre"]
            if not solo_aviso and not self._nombre_editado:
                self.v_nombre.set(self._nombre_auto)
            self._pintar_aviso()

        def _pintar_aviso(self):
            if self._suficiente():
                if self.v_sin_medidas.get() and not self.analisis["distintivos"]:
                    texto = ("✔ Marcado como material sin medidas/dimensiones: "
                             "se guardará así.")
                else:
                    texto = "✔ El nombre distingue esta ficha de otras similares."
                self.lbl_aviso.configure(text=texto, text_color=VERDE_OK)
                self.btn_ok.configure(state="normal")
                return
            self.lbl_aviso.configure(
                text="⚠ " + " ".join(self.analisis["faltantes"])
                     + "\nComplete el dato que falta (o escriba el nombre a mano) "
                       "para poder guardar.",
                text_color=ROJO_ES)
            self.btn_ok.configure(state="disabled")

        def _suficiente(self):
            """¿El nombre permite distinguir la ficha?

            Vale si el análisis encontró algo distintivo (dimensiones,
            presentación, designación o modelo) o si el usuario escribió un
            nombre a mano que sí lo tiene: la persona sabe más que la heurística.
            """
            if self.analisis["suficiente"]:
                return True
            nombre = self.v_nombre.get().strip()
            if not nombre or not self._nombre_editado:
                return False
            if not self.analisis["base"] or not self.analisis["marca"]:
                return False
            return bool(re.search(r"\d", nombre)) or len(nombre.split()) >= 5

        # ------------------------------------------------------------ guardar
        def _confirmar(self):
            d = self._datos_actuales()
            faltan = [c for c in bd_manager.CAMPOS_OBLIGATORIOS_FICHA if not d.get(c)]
            if faltan:
                messagebox.showwarning("Faltan datos",
                                       "Campos obligatorios: " + ", ".join(faltan), parent=self.winfo_toplevel())
                return
            if d["categoria"] not in bd_manager.CATEGORIAS:
                messagebox.showwarning("Categoría", "Use ARQ, ESTR, MEC o ELEC", parent=self.winfo_toplevel())
                return
            if not self._suficiente():
                messagebox.showwarning("Ficha indistinguible",
                                       "\n".join(self.analisis["faltantes"]), parent=self.winfo_toplevel())
                return

            escrito = self.v_nombre.get().strip()
            if not escrito:
                # Campo vacío: vuelve a mandar el nombre automático (y NO se
                # marca como manual, que era el error).
                escrito = self.analisis["nombre"]
                self._nombre_editado = False
            nombre = escrito
            d["nombre_ficha"] = nombre if self._nombre_editado else ""
            d["_requiere_manual"] = self.datos_origen.get("_requiere_manual", False)

            existente = self.bd.buscar_por_nombre(
                nombre, excluir_id=self.datos_origen.get("id"))
            if existente is not None:
                accion = self._preguntar_duplicado(existente, nombre)
                if accion is None:
                    return                       # el usuario canceló
                if accion != "guardar":
                    self.resultado = {"accion": accion, "ficha": existente}
                    self.destroy()
                    return
                d["nombre_ficha"] = nombre       # variante: nombre explícito
            self.resultado = {"accion": "guardar", "datos": d, "nombre_ficha": nombre}
            self.destroy()

        def _preguntar_duplicado(self, existente, nombre):
            """Tres salidas ante un nombre repetido; ``None`` si cancela."""
            top = ctk.CTkToplevel(self)
            top.title("Ficha repetida")
            top.configure(fg_color=GRIS_BG, padx=16, pady=14)
            top.grab_set()
            estado = existente.get("estado", "activo")
            tarjeta = _tarjeta(top); tarjeta.pack(fill="both", expand=True)
            ctk.CTkLabel(tarjeta, text="Ya existe una ficha con este nombre:",
                        font=_fuente(11, "bold"), text_color=GRIS_TEXTO).pack(
                anchor="w", padx=16, pady=(16, 0))
            ctk.CTkLabel(tarjeta, text=self.bd.nombre_de(existente), text_color=AZUL_ES,
                        wraplength=520, justify="left").pack(anchor="w", padx=16,
                                                             pady=(2, 6))
            ctk.CTkLabel(tarjeta, text=f"Cargada el {existente.get('fecha_carga', '?')}"
                               f"  ·  estado: {estado}", text_color=GRIS_TEXTO_SUAVE).pack(
                anchor="w", padx=16)
            ctk.CTkLabel(tarjeta, text="¿Qué desea hacer?", justify="left",
                        text_color=GRIS_TEXTO).pack(anchor="w", padx=16, pady=(10, 6))
            eleccion = {"v": None}

            def _elegir(valor):
                eleccion["v"] = valor
                top.destroy()

            if self.es_edicion:
                # Editando NO hay archivo nuevo que usar ni ficha que reemplazar:
                # las otras dos salidas no tienen sentido y antes descartaban la
                # edición en silencio.
                opciones = [
                    ("Guardar con este nombre igual", "guardar",
                     "Quedarán dos fichas con el mismo nombre."),
                ]
            else:
                opciones = [
                    ("Usar la ficha que ya existe", "usar_existente",
                     "No se carga nada nuevo."),
                    ("Reemplazar su archivo por este", "reemplazar",
                     "Corrige el archivo y conserva el nombre y las referencias."),
                    ("Guardar como variante", "guardar",
                     "Se agrega otra ficha con el mismo nombre."),
                ]
            for texto, valor, ayuda in opciones:
                f = ctk.CTkFrame(tarjeta, fg_color="transparent")
                f.pack(fill="x", pady=3, padx=16)
                _boton_secundario(f, texto, lambda v=valor: _elegir(v), ancho=220).pack(
                    side="left")
                ctk.CTkLabel(f, text=ayuda, text_color=GRIS_TEXTO_SUAVE).pack(
                    side="left", padx=8)
            _boton_secundario(tarjeta, "Cancelar", top.destroy, ancho=120).pack(
                pady=(10, 16))
            _dimensionar_ventana(top, 560, 320)
            _traer_al_frente(top)
            self.wait_window(top)
            return eleccion["v"]


    class _SelectorProyectoNube(ctk.CTkToplevel):
        """Lista los submittals ya guardados en la BD (``Proyectos/`` en
        GitHub, via ``BDManager.listar_proyectos()``) para continuarlos desde
        cualquier PC sin necesitar la carpeta local de quien los creo -- solo
        hace falta que ambas PCs sincronicen con la misma BD.

        ``self.resultado`` queda con el ``dict`` del proyecto elegido (o
        ``None`` si se cancela) para que quien la abre decida que hacer.
        """

        def __init__(self, master, bd):
            super().__init__(master)
            self.bd = bd
            self.resultado = None
            self._proyectos = []
            self.title("Abrir submittal existente")
            self.configure(fg_color=GRIS_BG, padx=14, pady=14)
            tarjeta = _tarjeta(self)
            tarjeta.pack(fill="both", expand=True, padx=4, pady=4)
            ctk.CTkLabel(tarjeta, text="Submittals guardados en la nube",
                        font=_fuente(14, "bold"), text_color=GRIS_TEXTO).pack(
                anchor="w", padx=18, pady=(16, 4))
            self.lbl_estado = ctk.CTkLabel(tarjeta, text="Sincronizando…",
                                           text_color=TEXTO_SUAVE)
            self.lbl_estado.pack(anchor="w", padx=18)

            barra = ctk.CTkFrame(tarjeta, fg_color="transparent")
            barra.pack(fill="x", padx=18, pady=(6, 8))
            _boton_secundario(barra, "Actualizar", self._sincronizar_y_listar,
                              ancho=110).pack(side="left")
            _boton_secundario(barra, "Buscar carpeta en esta PC…", self._buscar_local,
                              ancho=230).pack(side="left", padx=(8, 0))

            self.tree = ttk.Treeview(
                tarjeta, columns=("nombre", "materiales", "actualizado", "por"),
                show="headings", height=10)
            for c, t, w in (("nombre", "Proyecto", 260), ("materiales", "Materiales", 90),
                            ("actualizado", "Última actualización", 160),
                            ("por", "Actualizado por", 150)):
                self.tree.heading(c, text=t); self.tree.column(c, width=w)
            self.tree.pack(fill="both", expand=True, padx=18, pady=4)
            self.tree.bind("<Double-1>", self._abrir_seleccion)

            barra_final = ctk.CTkFrame(tarjeta, fg_color="transparent")
            barra_final.pack(pady=(4, 18))
            _boton(barra_final, "Abrir proyecto seleccionado", self._abrir_seleccion,
                  color=NARANJA_CTA, ancho=280, alto=42).pack(side="left", padx=6)
            _boton_secundario(barra_final, "Cancelar", self.destroy,
                              ancho=120, alto=42).pack(side="left", padx=6)

            _dimensionar_ventana(self, 760, 560)
            _traer_al_frente(self)
            self.after(100, self._sincronizar_y_listar)

        def _sincronizar_y_listar(self):
            self.lbl_estado.configure(text="Sincronizando con GitHub…",
                                      text_color=TEXTO_SUAVE)

            def trabajo():
                try:
                    self.bd.sincronizar()
                except Exception:
                    pass  # sin conexion: se sigue con la copia local de la BD
                proyectos = self.bd.listar_proyectos()
                try:
                    self.after(0, lambda: self._pintar(proyectos))
                except tk.TclError:
                    pass  # la ventana se cerro mientras sincronizaba
            threading.Thread(target=trabajo, daemon=True).start()

        def _pintar(self, proyectos):
            if not self.winfo_exists():
                return
            self._proyectos = proyectos
            for i in self.tree.get_children():
                self.tree.delete(i)
            for p in proyectos:
                fecha = (p.get("ultima_actualizacion") or "")[:16].replace("T", " ")
                self.tree.insert("", "end", iid=p["carpeta_bd"],
                                 values=(p["nombre_proyecto"], p["materiales"], fecha,
                                         p.get("actualizado_por", "")))
            n = len(proyectos)
            self.lbl_estado.configure(
                text=(f"{n} submittal(s) sincronizado(s)" if n else
                      "Todavía no hay submittals guardados en la BD"),
                text_color=GRIS_TEXTO_SUAVE)

        def _abrir_seleccion(self, _ev=None):
            sel = self.tree.selection()
            if not sel:
                messagebox.showinfo("Seleccione un proyecto",
                                    "Elija un proyecto de la lista.", parent=self)
                return
            try:
                proyecto = bd_manager.BDManager.cargar_submittal(sel[0])
            except Exception as e:
                messagebox.showerror("No se pudo abrir", str(e), parent=self); return
            self.resultado = proyecto
            self.destroy()

        def _buscar_local(self):
            """Alternativa para carpetas que todavia no estan en la BD (ej.
            submittals viejos de v2.6 nunca guardados con 'Guardar avance')."""
            carpeta = filedialog.askdirectory(
                title="Carpeta del submittal (con submittal_proyecto.json)", parent=self)
            if not carpeta:
                return
            try:
                proyecto = bd_manager.BDManager.cargar_submittal(carpeta)
            except Exception as e:
                messagebox.showerror("No se pudo abrir", str(e), parent=self); return
            self.resultado = proyecto
            self.destroy()

    class PantallaSubmittal(ctk.CTkFrame):
        """Pantalla 'Submittal activo': arma el submittal (tanto el flujo
        'Generar desde BD' como 'Abrir existente').

        v3.4.0: era una ventana aparte (``_VentanaSubmittal``); ahora vive
        dentro de la ventana principal, detras del item 'Submittal activo' del
        menu lateral. La logica de guardado/generado no cambio: lo que antes
        pasaba al cerrar la ventana ahora lo dispara ``guardar_al_salir()``,
        que la ventana principal llama al cerrarse o al cambiar de submittal.
        """

        _ETIQ_ES = "ES (clásica)"
        _ETIQ_MINSAL = "Ministerio de Salud"

        def __init__(self, master, bd, proyecto, destino, titulo, app=None):
            super().__init__(master, fg_color="transparent")
            self.bd = bd; self.proyecto = proyecto; self.destino = destino
            self.app = app
            self.titulo = titulo
            self.grid_rowconfigure(1, weight=1)
            self.grid_columnconfigure(0, weight=1)

            # ------------------------------------------------- encabezado
            cab = ctk.CTkFrame(self, fg_color=SUPERFICIE_2, corner_radius=0)
            cab.grid(row=0, column=0, sticky="ew")
            ctk.CTkFrame(self, height=1, fg_color=BORDE_FUERTE).grid(
                row=0, column=0, sticky="sew")

            fila1 = ctk.CTkFrame(cab, fg_color="transparent")
            fila1.pack(fill="x", padx=26, pady=(20, 0))
            izq = ctk.CTkFrame(fila1, fg_color="transparent")
            izq.pack(side="left")
            _etiqueta_seccion(izq, "editando", color=ACENTO_TXT).pack(
                side="left", padx=(0, 10))
            ctk.CTkLabel(izq, text=proyecto.get("nombre_proyecto", "Submittal"),
                        font=_fuente(19, "bold"), text_color=TEXTO).pack(side="left")
            acciones = ctk.CTkFrame(fila1, fg_color="transparent")
            acciones.pack(side="right")
            _boton_secundario(acciones, "Datos del proyecto", self._datos,
                             ancho=150).pack(side="left")
            _boton_secundario(acciones, "Guardar avance", self._guardar_avance,
                             ancho=132).pack(side="left", padx=(8, 0))
            _boton(acciones, "Generar submittal", self._generar, ancho=158).pack(
                side="left", padx=(8, 0))

            fila2 = ctk.CTkFrame(cab, fg_color="transparent")
            fila2.pack(fill="x", padx=26, pady=(14, 16))
            dest = ctk.CTkFrame(fila2, fg_color="transparent")
            dest.pack(side="left", fill="x", expand=True)
            _etiqueta_seccion(dest, "carpeta destino", color=TEXTO_SUAVE).pack(
                side="left", padx=(0, 10))
            self.var_dest = tk.StringVar(value=destino or "")
            ctk.CTkEntry(dest, textvariable=self.var_dest, height=30,
                        corner_radius=6, border_color=BORDE, fg_color=SUPERFICIE,
                        text_color=TEXTO_2, font=_mono(11)).pack(
                side="left", fill="x", expand=True)
            _boton_secundario(dest, "Examinar…", self._elegir_destino, ancho=92,
                              alto=30).pack(side="left", padx=(6, 0))

            car = ctk.CTkFrame(fila2, fg_color="transparent")
            car.pack(side="left", padx=(26, 0))
            _etiqueta_seccion(car, "carátula", color=TEXTO_SUAVE).pack(
                side="left", padx=(0, 10))
            self.var_tipo_caratula = tk.StringVar(value=self._ETIQ_MINSAL
                if proyecto.get("tipo_caratula", "clasica") == "ministerio_salud"
                else self._ETIQ_ES)
            ctk.CTkSegmentedButton(car, values=[self._ETIQ_ES, self._ETIQ_MINSAL],
                                   variable=self.var_tipo_caratula, height=30,
                                   corner_radius=6, font=_fuente(12),
                                   fg_color=BORDE_TENUE,
                                   selected_color=SUPERFICIE,
                                   selected_hover_color=SUPERFICIE,
                                   unselected_color=BORDE_TENUE,
                                   unselected_hover_color=BORDE,
                                   text_color=TEXTO, text_color_disabled=TEXTO_TENUE,
                                   border_width=2,
                                   command=self._cambiar_tipo_caratula).pack(side="left")
            self._cambiar_tipo_caratula(self.var_tipo_caratula.get())

            # ------------------------------------------------------ cuerpo
            cuerpo = ctk.CTkFrame(self, fg_color="transparent")
            cuerpo.grid(row=1, column=0, sticky="nsew", padx=16, pady=14)
            cuerpo.grid_rowconfigure(0, weight=1)
            cuerpo.grid_columnconfigure(0, weight=1)
            self.tabla = TablaMateriales(cuerpo, bd,
                                         proyecto.get("materiales_seleccionados", []))
            self.tabla.grid(row=0, column=0, sticky="nsew")

            # El registro va debajo de los materiales (columna derecha), como en
            # el diseno: lo que se acaba de hacer queda al lado de la lista que
            # cambio, no al final de toda la pantalla.
            registro = ctk.CTkFrame(self.tabla.der, fg_color=LOG_BG, corner_radius=8,
                                    height=108)
            registro.pack(fill="x", pady=(10, 0))
            registro.pack_propagate(False)
            _etiqueta_seccion(registro, "registro", color=LOG_LABEL,
                              height=14).pack(anchor="w", padx=12, pady=(8, 0))
            self.txt = _consola(registro, fg_color=LOG_BG, height=68)
            self.txt.pack(fill="both", expand=True, padx=8, pady=(2, 8))
            self._preparar_tags_log()

        # ------------------------------------------------------------- log
        def _preparar_tags_log(self):
            """Colorea el registro por tipo de linea (ok / aviso / error), como
            en el diseno. Si la version de CustomTkinter no expone el ``Text``
            interno, se degrada a texto plano sin fallar."""
            self._tags_log = False
            try:
                caja = self.txt._textbox
                caja.tag_config("ok", foreground=LOG_OK)
                caja.tag_config("warn", foreground=LOG_WARN)
                caja.tag_config("err", foreground=ROJO_CLARO)
                self._tags_log = True
            except Exception:
                pass

        def _log(self, m):
            texto = str(m)
            tag = None
            if any(s in texto for s in ("❌", "⚠️", "No se pudo", "Error")):
                tag = "err" if ("❌" in texto or "Error" in texto) else "warn"
            elif any(s in texto for s in ("✅", "☁️", "Listo")):
                tag = "ok"
            if self._tags_log and tag:
                self.txt._textbox.insert("end", texto + "\n", tag)
            else:
                self.txt.insert("end", texto + "\n")
            self.txt.see("end")
            self.update_idletasks()

        def _datos(self):
            d = DatosProyectoDialog(self, self.proyecto.get("datos_procedimiento", {}))
            self.wait_window(d)
            if d.resultado:
                self.proyecto["datos_procedimiento"] = d.resultado

        def _elegir_destino(self):
            c = filedialog.askdirectory(title="Carpeta destino del submittal")
            if c:
                self.var_dest.set(c)

        def _cambiar_tipo_caratula(self, valor):
            self.proyecto["tipo_caratula"] = (
                "ministerio_salud" if valor == self._ETIQ_MINSAL else "clasica")

        def _guardar_avance(self, silencioso=False):
            """Guarda el submittal (materiales + datos del proyecto) en la BD
            SIN generar caratulas/compilados/Excel, para seguir despues desde
            'Abrir existente'. No necesita carpeta destino.

            Con ``silencioso=True`` (usado al salir de la pantalla) no muestra
            avisos ni escribe en el log -- solo guarda."""
            self.proyecto["materiales_seleccionados"] = self.tabla.materiales
            destino = self.var_dest.get().strip()
            self.bd.guardar_submittal(self.proyecto, destino=destino or None)
            r = self.bd.git_push(
                f"guardar avance de {self.proyecto.get('nombre_proyecto', 'submittal')}")
            if silencioso:
                return
            self._log("💾 Avance guardado (sin generar carátulas/compilados).")
            if self.app is not None:
                self.app._actualizar_estado()
            _avisar_resultado_git(
                self.winfo_toplevel(), r,
                "Avance guardado. Podés seguir en otra pantalla y volver después "
                "a 'Submittal activo' (o abrirlo con 'Abrir submittal existente').")

        def guardar_al_salir(self):
            """Guarda el avance SIEMPRE al dejar de trabajar en este submittal
            (al cerrar la aplicacion o al abrir otro), aunque no se haya tocado
            "Guardar avance" ni "Generar". Antes, cerrar sin guardar perdia en
            silencio los materiales agregados en la sesion; al reabrir se
            continuaba desde el ultimo guardado real (menos materiales de los
            que la persona alcanzo a ver en pantalla), lo que se percibia como
            "el contador de consecutivos se resetea" -- la numeracion en si
            nunca estuvo mal, era el progreso el que no se habia guardado.

            Devuelve ``True`` si se puede continuar (guardo bien, o el usuario
            eligio seguir igual) y ``False`` si el usuario prefiere quedarse.
            """
            try:
                self._guardar_avance(silencioso=True)
            except Exception as e:
                return bool(messagebox.askyesno(
                    "No se pudo guardar",
                    f"No se pudo guardar el avance del submittal:\n{e}\n\n"
                    "¿Continuar de todas formas? Se perderían los cambios sin "
                    "guardar de esta sesión.", parent=self.winfo_toplevel()))
            return True

        def _generar(self):
            self.proyecto["materiales_seleccionados"] = self.tabla.materiales
            destino = self.var_dest.get().strip()
            if not destino:
                messagebox.showwarning("Destino", "Elija una carpeta destino", parent=self.winfo_toplevel()); return
            ok, errores = self.bd.validar_proyecto(self.proyecto)
            if not ok:
                messagebox.showerror("No se puede generar", "\n".join(errores), parent=self.winfo_toplevel()); return
            if not messagebox.askyesno("Confirmar",
                                       "Se regenerarán carátulas, compilados y Excel. ¿Continuar?", parent=self.winfo_toplevel()):
                return
            tipo = self.proyecto.get("tipo_caratula", "clasica")
            try:
                res = generar_entregables(self.bd, self.proyecto, destino, tipo=tipo, log=self._log)
                self._log(f"✅ Listo: {res['materiales']} materiales en {res['destino']}")
                self._subir_metadatos()
                if self.app is not None:
                    # Refresca conteos del menu lateral y "Actividad reciente".
                    self.app._actualizar_estado()
                try:
                    os.startfile(destino)  # Windows: abre el explorador
                except Exception:
                    pass
            except Exception as e:
                self._log(f"\n❌ {e}")
                messagebox.showerror("Error al generar", str(e), parent=self.winfo_toplevel())

        def _subir_metadatos(self):
            """Sube a GitHub el ``submittal_proyecto.json`` del proyecto.

            Solo metadatos: las carátulas, los CMP y los Excel se quedan en la
            carpeta local y se regeneran cuando se necesiten.
            """
            nombre = self.proyecto.get("nombre_proyecto", "proyecto")
            self._log("🔄 Sincronizando con GitHub…")
            r = self.bd.git_push(f"submittal {nombre}")
            if r.get("desactivado"):
                return
            if r.get("subido"):
                extra = (f" · {r['conflictos']} conflicto(s) resuelto(s)"
                         if r.get("conflictos") else "")
                self._log(f"☁️ Metadatos del submittal subidos a GitHub{extra}")
            elif r.get("offline"):
                self._log("📡 Sin conexión: los entregables están listos; los "
                          "metadatos se subirán al reconectar")
            elif r.get("auth"):
                self._log("🔑 Falta el token de GitHub para subir los metadatos")
            elif not r.get("nada_que_subir"):
                self._log(f"⚠️ No se pudo subir: {r.get('error', '')}")


    def _probar_openai_key(api_key):
        """Verifica una API key contra OpenAI. Devuelve ``(ok, mensaje)``.

        Replica la comprobacion de v2.6 (``test_openai``) para no depender de que
        submitals_gui.py este disponible: aqui v3 es autonomo.
        """
        api_key = (api_key or "").strip()
        if not api_key:
            return (False, "Ingrese una API Key.")
        try:
            from openai import OpenAI
        except Exception:
            return (False, "Falta la librería 'openai' (pip install openai).")
        try:
            OpenAI(api_key=api_key, timeout=15).models.list()
            return (True, "Conexión exitosa: la API key funciona.")
        except Exception as e:
            try:
                import openai
                if isinstance(e, openai.AuthenticationError):
                    return (False, "API key inválida o sin créditos.")
            except Exception:
                pass
            return (False, f"No se pudo conectar: {str(e)[:150]}")

    class PantallaConfig(ctk.CTkFrame):
        """Pantalla 'Configuración': ajustes del usuario, agrupados en un menu
        propio a la izquierda (como en el diseno) en vez de pestañas:

          * Lectura con IA: la API key que usa la lectura de fichas con IA.
          * Sincronización: repositorio, rama y token (PAT) de la BD.
          * Rutas y carpetas: donde vive la BD, la caché y el registro.
          * Acerca de: versión y estado del programa.

        Ambos secretos se guardan (ofuscados en base64, igual que v2.6) en
        ``%APPDATA%/GeneradorSubmittals/config.json``.

        v3.4.0: era ``DialogoConfiguracion`` (una ventana modal con pestañas);
        pasa a ser una pantalla de la ventana principal. La logica de guardado
        es la misma.
        """

        _SECCIONES = [("openai", "Lectura con IA"),
                      ("github", "Sincronización"),
                      ("rutas", "Rutas y carpetas"),
                      ("acerca", "Acerca de")]

        def __init__(self, master, bd, seccion_inicial="openai", app=None):
            super().__init__(master, fg_color="transparent")
            self.bd = bd
            self.app = app
            self.cambio_github = False      # solo si cambio algo de GitHub -> resync
            self._probando = False
            self.grid_rowconfigure(0, weight=1)
            self.grid_columnconfigure(0, minsize=208)
            self.grid_columnconfigure(1, weight=1)

            # ------------------------------------------- menu de secciones
            lateral = ctk.CTkFrame(self, fg_color=SUPERFICIE_2, corner_radius=0,
                                   width=208)
            lateral.grid(row=0, column=0, sticky="nsw")
            lateral.grid_propagate(False)
            ctk.CTkFrame(self, width=1, fg_color=BORDE_FUERTE).grid(
                row=0, column=0, sticky="nse")
            _etiqueta_seccion(lateral, "ajustes", color=TEXTO_SUAVE).pack(
                anchor="w", padx=24, pady=(22, 12))
            self._btn_seccion = {}
            for clave, etiqueta in self._SECCIONES:
                b = ctk.CTkButton(
                    lateral, text=etiqueta, anchor="w", height=32, corner_radius=6,
                    fg_color="transparent", hover_color=BORDE_TENUE,
                    text_color=TEXTO_2, font=_fuente(12),
                    command=lambda c=clave: self.mostrar_seccion(c))
                b.pack(fill="x", padx=14, pady=1)
                self._btn_seccion[clave] = b

            # ------------------------------------------------- contenido
            derecha = ctk.CTkFrame(self, fg_color="transparent")
            derecha.grid(row=0, column=1, sticky="nsew")
            derecha.grid_rowconfigure(0, weight=1)
            derecha.grid_columnconfigure(0, weight=1)
            self._cuerpo = ctk.CTkScrollableFrame(derecha, fg_color="transparent",
                                                  scrollbar_button_color=BORDE,
                                                  scrollbar_button_hover_color=BORDE_FUERTE)
            self._cuerpo.grid(row=0, column=0, sticky="nsew", padx=(30, 20), pady=(26, 0))

            self._paneles = {}
            for clave, _e in self._SECCIONES:
                p = ctk.CTkFrame(self._cuerpo, fg_color="transparent")
                self._paneles[clave] = p
            self._build_openai(self._paneles["openai"])
            self._build_github(self._paneles["github"])
            self._build_rutas(self._paneles["rutas"])
            self._build_acerca(self._paneles["acerca"])

            # -------------------------------------------------- pie fijo
            ctk.CTkFrame(derecha, height=1, fg_color=BORDE_FUERTE).grid(
                row=1, column=0, sticky="ew")
            pie = ctk.CTkFrame(derecha, fg_color=SUPERFICIE, corner_radius=0, height=58)
            pie.grid(row=2, column=0, sticky="ew")
            pie.grid_propagate(False)
            _boton(pie, "Guardar", self._guardar, ancho=110).pack(
                side="right", padx=(0, 30), pady=12)
            _boton_secundario(pie, "Cancelar", self._cancelar, ancho=104).pack(
                side="right", padx=(0, 8), pady=12)

            self.mostrar_seccion(seccion_inicial)

        # ---------------------------------------------------- navegacion
        def mostrar_seccion(self, clave):
            if clave not in self._paneles:
                clave = "openai"
            for c, panel in self._paneles.items():
                panel.pack_forget()
                self._btn_seccion[c].configure(fg_color="transparent",
                                               text_color=TEXTO_2,
                                               font=_fuente(12))
            self._paneles[clave].pack(fill="both", expand=True)
            self._btn_seccion[clave].configure(fg_color=CHIP_BG,
                                               text_color=ACENTO_TXT,
                                               font=_fuente(12, "bold"))

        def _cancelar(self):
            """No hay nada que revertir: los campos solo se aplican al guardar.
            Simplemente se vuelve al inicio."""
            if self.app is not None:
                self.app._ir("inicio")

        # ------------------------------------------------------ ayudantes
        def _titulo(self, padre, texto, ayuda=None):
            ctk.CTkLabel(padre, text=texto, font=_fuente(19, "bold"),
                        text_color=TEXTO, anchor="w").pack(anchor="w")
            if ayuda:
                ctk.CTkLabel(padre, text=ayuda, text_color=TEXTO_SUAVE,
                            font=_fuente(12), justify="left", wraplength=560,
                            anchor="w").pack(anchor="w", pady=(6, 0))

        def _panel(self, padre, titulo, insignia=None, color_insignia=VERDE):
            """Tarjeta con cabecera (y opcionalmente una insignia de estado a la
            derecha, como los 'Configurada · 220 ms' del diseno)."""
            tarjeta = _tarjeta(padre)
            tarjeta.pack(fill="x", pady=(18, 0))
            # padx/pady de 1: si la cabecera se pega al borde, tapa la linea de
            # 1px del marco y la tarjeta parece abierta por arriba.
            cab = ctk.CTkFrame(tarjeta, fg_color=SUPERFICIE_2, corner_radius=0)
            cab.pack(fill="x", padx=1, pady=(1, 0))
            ctk.CTkLabel(cab, text=titulo, font=_fuente(13, "bold"),
                        text_color=TEXTO_2).pack(side="left", padx=18, pady=13)
            if insignia:
                caja = ctk.CTkFrame(cab, fg_color="transparent")
                caja.pack(side="right", padx=18)
                ctk.CTkLabel(caja, text="●", text_color=color_insignia,
                            font=_fuente(11)).pack(side="left", padx=(0, 5))
                ctk.CTkLabel(caja, text=insignia, text_color=color_insignia,
                            font=_mono(11)).pack(side="left")
            ctk.CTkFrame(tarjeta, height=1, fg_color=BORDE_TENUE).pack(fill="x")
            interior = ctk.CTkFrame(tarjeta, fg_color="transparent")
            interior.pack(fill="x", padx=18, pady=16)
            return interior

        def _fila_dato(self, padre, etiqueta, valor, primera=False):
            """Fila 'etiqueta ......... valor' con separador, como en el diseno."""
            if not primera:
                ctk.CTkFrame(padre, height=1, fg_color=BORDE_TENUE).pack(
                    fill="x", pady=11)
            f = ctk.CTkFrame(padre, fg_color="transparent")
            f.pack(fill="x")
            ctk.CTkLabel(f, text=etiqueta, text_color=TEXTO_2,
                        font=_fuente(12)).pack(side="left")
            ctk.CTkLabel(f, text=valor, text_color=TEXTO_SUAVE, font=_mono(11),
                        anchor="e").pack(side="right")
            return f

        # -------------------------------------------- seccion: lectura IA
        def _build_openai(self, f):
            self._titulo(
                f, "Lectura de fichas con IA",
                "La API se usa para leer fichas técnicas en PDF o imagen. Sin "
                "ella, la extracción cae a OCR local y revisión manual.")

            # El entorno TIENE PRIORIDAD sobre la config (ver obtener_api_key):
            # se distingue la fuente para no mostrar "configurada" por una key
            # guardada que en realidad no se usa porque la enmascara el entorno.
            env_key = os.environ.get("OPENAI_API_KEY", "").strip()
            guardada = bd_manager.descifrar_api_key(
                self.bd.cfg.get("api", {}).get("openai_key_encrypted", ""))
            if env_key:
                insignia, color = "Configurada por entorno", AMBAR
                nota = ("La variable de entorno OPENAI_API_KEY tiene PRIORIDAD "
                        "sobre la clave guardada acá.")
            elif guardada:
                insignia, color = "Configurada", VERDE
                nota = "Deje el campo vacío para conservar la clave actual."
            else:
                insignia, color = "Sin configurar", ROJO
                nota = "Cree su API key en platform.openai.com/api-keys"

            caja = self._panel(f, "Proveedor de IA", insignia=insignia,
                               color_insignia=color)
            _etiqueta_seccion(caja, "api key", color=TEXTO_SUAVE).pack(anchor="w")
            fila = ctk.CTkFrame(caja, fg_color="transparent")
            fila.pack(fill="x", pady=(7, 0))
            self.v_openai = tk.StringVar(value="")
            self.e_openai = ctk.CTkEntry(
                fila, textvariable=self.v_openai, height=34, corner_radius=6,
                border_color=BORDE, fg_color=SUPERFICIE_3, text_color=TEXTO,
                font=_mono(12), show="•")
            _pista(self.e_openai, self.v_openai,
                   "sk-•••• •••• •••• (guardada — dejar vacío para conservarla)"
                   if (guardada or env_key) else "sk-… (pegá acá la API key)",
                   size=11)
            self.e_openai.pack(side="left", fill="x", expand=True)
            self.v_mostrar = tk.BooleanVar(value=False)
            self.btn_mostrar = _boton_secundario(
                fila, "Mostrar", self._alternar_mostrar, ancho=84, alto=34)
            self.btn_mostrar.pack(side="left", padx=(8, 0))
            ctk.CTkLabel(caja, text=nota, text_color=TEXTO_SUAVE,
                        font=_fuente(11), justify="left", wraplength=520).pack(
                anchor="w", pady=(7, 0))

            prueba = ctk.CTkFrame(caja, fg_color="transparent")
            prueba.pack(fill="x", pady=(16, 0))
            self.btn_probar = _boton_secundario(prueba, "Probar conexión",
                                                self._probar_openai, ancho=140,
                                                alto=34)
            self.btn_probar.pack(side="left")
            self.lbl_openai_estado = ctk.CTkLabel(prueba, text="", text_color=TEXTO_SUAVE,
                                                  font=_mono(11), justify="left",
                                                  wraplength=380)
            self.lbl_openai_estado.pack(side="left", padx=(12, 0))

        def _alternar_mostrar(self):
            """Mostrar/ocultar la API key. El diseno usa un boton "Mostrar" en
            vez de una casilla: ocupa menos y dice mejor que va a pasar."""
            self.v_mostrar.set(not self.v_mostrar.get())
            visible = self.v_mostrar.get()
            self.e_openai.configure(show="" if visible else "•")
            self.btn_mostrar.configure(text="Ocultar" if visible else "Mostrar")

        def _probar_openai(self):
            if self._probando:
                return
            key = (self.v_openai.get().strip()
                   or bd_manager.obtener_api_key(cfg=self.bd.cfg,
                                                 config_dir=self.bd.config_dir))
            if not key:
                self.lbl_openai_estado.configure(
                    text="Ingrese una API Key para probar.", text_color=ROJO_ES)
                return
            self._probando = True
            self.btn_probar.configure(state="disabled")
            self.lbl_openai_estado.configure(text="Probando conexión…",
                                             text_color=TEXTO_SUAVE)

            def trabajo():
                ok, msg = _probar_openai_key(key)
                try:
                    self.after(0, lambda: self._fin_probar(ok, msg))
                except tk.TclError:
                    pass

            threading.Thread(target=trabajo, daemon=True).start()

        def _fin_probar(self, ok, msg):
            # La prueba corre en un hilo (hasta 15 s). Si el usuario cerró la
            # ventana mientras tanto, los widgets ya no existen: winfo_exists()
            # devuelve 0 sin lanzar, y así evitamos un TclError al tocarlos.
            if not self.winfo_exists():
                return
            self._probando = False
            self.btn_probar.configure(state="normal")
            self.lbl_openai_estado.configure(text=("✅ " if ok else "❌ ") + msg,
                                             text_color=(VERDE_OK if ok else ROJO_ES))

        # ------------------------------------------ seccion: sincronizacion
        def _build_github(self, f):
            gh = self.bd.cfg.get("github", {}) or {}
            est = self.bd.git_status()
            self._titulo(
                f, "Sincronización de la Base de Datos",
                "La BD de fichas vive en un repositorio de GitHub: así todas "
                "las computadoras trabajan sobre el mismo catálogo.")

            autenticado = bool(est.get("autenticado"))
            caja = self._panel(
                f, "Repositorio",
                insignia=("Token configurado" if autenticado else "Sin token"),
                color_insignia=(VERDE if autenticado else ROJO))
            self.v_repo = tk.StringVar(value=gh.get("repo", ""))
            self.v_rama = tk.StringVar(value=gh.get("branch", "main"))
            self.v_token = tk.StringVar(value="")
            campos = [("repositorio (usuario/repo)", self.v_repo, False,
                       "es-constructora/submittals-bd"),
                      ("rama", self.v_rama, False, "main"),
                      ("token (pat)", self.v_token, True,
                       "ghp_… (vacío = conservar el actual)")]
            for i, (etiqueta, var, secreto, pista) in enumerate(campos):
                _etiqueta_seccion(caja, etiqueta, color=TEXTO_SUAVE).pack(
                    anchor="w", pady=((0 if i == 0 else 14), 6))
                e = ctk.CTkEntry(caja, textvariable=var, height=34, corner_radius=6,
                                border_color=BORDE, fg_color=SUPERFICIE_3,
                                text_color=TEXTO, font=_mono(12),
                                show="•" if secreto else "")
                _pista(e, var, pista)
                e.pack(fill="x")

            modo = {"git": "git instalado", "rest": "API REST (sin git)",
                    "local": "BD local (sin sincronización)"}.get(
                est.get("backend"), est.get("backend", "?"))
            info = self._panel(f, "Estado")
            self._fila_dato(info, "Método de transporte", modo, primera=True)
            self._fila_dato(info, "Cambios sin subir",
                            str(est.get("pendientes", 0)))
            self._fila_dato(info, "Última sincronización",
                            self.bd.texto_estado_sync())
            ctk.CTkLabel(
                f, text="Cree el token en github.com/settings/tokens con permiso "
                        "Contents: write SOLO sobre este repositorio.",
                text_color=TEXTO_SUAVE, font=_fuente(11), justify="left",
                wraplength=560).pack(anchor="w", pady=(14, 24))

        # ------------------------------------------ seccion: rutas/carpetas
        def _build_rutas(self, f):
            self._titulo(
                f, "Rutas y carpetas",
                "Dónde guarda el programa la base de datos, la caché y el "
                "registro de esta computadora. Son de solo lectura.")
            caja = self._panel(f, "Ubicaciones en esta PC")
            rutas = [("Base de datos (copia local)",
                      getattr(self.bd, "bd_root", "") or getattr(self.bd, "cache_dir", "")),
                     ("Submittals guardados", getattr(self.bd, "proyectos_dir", "")),
                     ("Caché", getattr(self.bd, "cache_dir", "")),
                     ("Configuración", getattr(self.bd, "config_dir", "")),
                     ("Registro (app.log)", LOG_PATH)]
            for i, (etiqueta, ruta) in enumerate(rutas):
                fila = self._fila_dato(caja, etiqueta,
                                       self._acortar(str(ruta)), primera=(i == 0))
                _enlace(fila, "abrir", lambda r=ruta: self._abrir_carpeta(r)).pack(
                    side="right", padx=(0, 10))

        @staticmethod
        def _acortar(ruta, tope=46):
            return ruta if len(ruta) <= tope else "…" + ruta[-(tope - 1):]

        def _abrir_carpeta(self, ruta):
            """Abre la carpeta en el explorador (si el destino es un archivo, se
            abre la carpeta que lo contiene)."""
            try:
                p = Path(str(ruta))
                destino = p if p.is_dir() else p.parent
                os.startfile(destino)
            except Exception as e:
                messagebox.showerror("No se pudo abrir", str(e),
                                     parent=self.winfo_toplevel())

        # ----------------------------------------------- seccion: acerca de
        def _build_acerca(self, f):
            self._titulo(f, "Acerca de",
                         "Generador de Submittals ES — plataforma de armado de "
                         "submittals sobre una BD central de fichas técnicas.")
            caja = self._panel(f, "Versión y entorno")
            self._fila_dato(caja, "Versión del programa", f"v{VERSION}", primera=True)
            self._fila_dato(caja, "Modo de ejecución",
                            "empaquetado (.exe)" if getattr(sys, "frozen", False)
                            else "código fuente (.py)")
            r = self.bd.resumen_por_categoria()
            self._fila_dato(caja, "Fichas activas en la BD", str(r["TOTAL"]))
            self._fila_dato(caja, "Por especialidad",
                            f"ARQ {r['ARQ']} · ESTR {r['ESTR']} · "
                            f"MEC {r['MEC']} · ELEC {r['ELEC']}")
            acciones = ctk.CTkFrame(f, fg_color="transparent")
            acciones.pack(anchor="w", pady=(18, 24))
            _boton_secundario(acciones, "Buscar actualización",
                              lambda: self.app._buscar_update() if self.app else None,
                              ancho=160, alto=34).pack(side="left")
            _boton_secundario(acciones, "Ver registro",
                              lambda: self._abrir_registro(), ancho=124,
                              alto=34).pack(side="left", padx=(8, 0))

        def _abrir_registro(self):
            try:
                os.startfile(LOG_PATH)
            except Exception as e:
                messagebox.showerror("No se pudo abrir el registro", str(e),
                                     parent=self.winfo_toplevel())

        # -------------------------------------------------- guardado
        def _guardar(self):
            cfg = self.bd.cfg

            # --- OpenAI: solo se toca si el usuario escribio algo ---
            key = self.v_openai.get().strip()
            if key:
                cfg.setdefault("api", {})["openai_key_encrypted"] = \
                    bd_manager.cifrar_api_key(key)

            # --- GitHub ---
            cfg.setdefault("github", {})
            old_repo = cfg["github"].get("repo")
            old_rama = cfg["github"].get("branch")
            repo_nuevo = self.v_repo.get().strip() or old_repo or ""
            rama_nueva = self.v_rama.get().strip() or "main"
            token = self.v_token.get().strip()
            repo_o_rama_cambio = (repo_nuevo != old_repo) or (rama_nueva != old_rama)
            if repo_o_rama_cambio or token:
                self.cambio_github = True
            cfg["github"]["repo"] = repo_nuevo
            cfg["github"]["branch"] = rama_nueva
            if token:
                cfg["github"]["token_encrypted"] = bd_manager.cifrar_secreto(token)

            bd_manager.guardar_config(cfg, self.bd.config_dir)

            # Aplicar los cambios al objeto de sincronización EN VIVO (sin reiniciar):
            #  * repo/rama quedan fijados al CONSTRUIR el transporte, así que un
            #    cambio de repo/rama exige reconstruir el sync (set_token no basta);
            #    crear_sync toma repo/rama/token del cfg ya actualizado.
            #  * si solo cambió el token, basta set_token (camino ya probado).
            if self.bd.sync is not None:
                if repo_o_rama_cambio:
                    self.bd.sync = bd_manager.crear_sync(
                        cfg, self.bd.config_dir, self.bd.cache_dir, self.bd.log)
                elif token:
                    self.bd.sync.set_token(token)

            # Aviso: una variable de entorno OPENAI_API_KEY tiene PRIORIDAD sobre la
            # key guardada, así que la lectura de fichas la ignoraría en silencio.
            if key and os.environ.get("OPENAI_API_KEY", "").strip():
                messagebox.showwarning(
                    "Variable de entorno detectada",
                    "Guardé la API key, pero esta computadora tiene una variable de "
                    "entorno OPENAI_API_KEY que TIENE PRIORIDAD sobre ella.\n\n"
                    "La lectura de fichas seguirá usando la variable de entorno hasta "
                    "que la elimine.", parent=self.winfo_toplevel())
            else:
                messagebox.showinfo("Configuración", "Configuración guardada.", parent=self.winfo_toplevel())
            # Antes esto cerraba la ventana modal; ahora vuelve al inicio y, si
            # cambio algo de GitHub, la ventana principal vuelve a sincronizar.
            if self.app is not None:
                self.app._tras_guardar_config(self.cambio_github)


    class PantallaInicio(ctk.CTkFrame):
        """Pantalla de inicio: las dos acciones grandes (generar / abrir), el
        mantenimiento del catálogo y la actividad reciente.

        Reemplaza al menu 2x2 de v3.3.x sin quitar ninguna de sus cuatro
        entradas: 'Generar desde BD' y 'Abrir submittal existente' pasan a ser
        las tarjetas grandes, y 'Cargar ficha a BD' y 'Gestionar BD' quedan en
        el bloque de mantenimiento, con el resto de las acciones de catalogo.
        """

        def __init__(self, master, app, bd=None):
            super().__init__(master, fg_color="transparent")
            self.app = app
            self._build()

        @property
        def bd(self):
            # La BD se crea DESPUES de armar la ventana (primero hay que
            # verificar git), asi que se busca al usarla, no al construir.
            return getattr(self.app, "bd", None)

        def _build(self):
            # Con scroll: en un monitor de 1366x768 el area de contenido queda
            # en ~660 px de alto y, sin scroll, la actividad reciente quedaria
            # fuera de alcance.
            cuerpo = ctk.CTkScrollableFrame(
                self, fg_color="transparent", height=380, width=600,
                scrollbar_button_color=BORDE,
                scrollbar_button_hover_color=BORDE_FUERTE)
            cuerpo.pack(fill="both", expand=True, padx=20, pady=(24, 16))

            enc = ctk.CTkFrame(cuerpo, fg_color="transparent")
            enc.pack(fill="x")
            ctk.CTkLabel(enc, text="¿Qué vamos a hacer hoy?", font=_fuente(23, "bold"),
                        text_color=TEXTO, anchor="w").pack(anchor="w")
            ctk.CTkLabel(enc, text="Generá un submittal nuevo desde la base de datos "
                                   "o continuá con uno existente.",
                        font=_fuente(13), text_color=TEXTO_SUAVE, anchor="w").pack(
                anchor="w", pady=(4, 0))

            # ------------------------------------------ acciones principales
            acc = ctk.CTkFrame(cuerpo, fg_color="transparent")
            acc.pack(fill="x", pady=(18, 0))
            acc.grid_columnconfigure(0, weight=1, uniform="acc")
            acc.grid_columnconfigure(1, weight=1, uniform="acc")
            self._tarjeta_grande(
                acc, 0, "Generar desde BD",
                "Seleccioná fichas por especialidad y armá el paquete en PDF.",
                "ACCIÓN PRINCIPAL", self.app._generar_desde_bd, destacada=True)
            self.lbl_reciente = self._tarjeta_grande(
                acc, 1, "Abrir submittal existente",
                "Retomá un paquete guardado y actualizá sus fichas.",
                "SIN SUBMITTALS GUARDADOS", self.app._abrir_existente)

            # ------------------------------------------------ mantenimiento
            _etiqueta_seccion(cuerpo, "mantenimiento del catálogo",
                              color=TEXTO_SUAVE, height=16).pack(anchor="w",
                                                                 pady=(20, 8))
            mant = ctk.CTkFrame(cuerpo, fg_color="transparent")
            mant.pack(fill="x")
            mant.grid_columnconfigure(0, weight=1, uniform="mant")
            mant.grid_columnconfigure(1, weight=1, uniform="mant")
            self.lbl_total_bd = None
            fichas = [
                ("Cargar ficha a BD", "+", self.app._cargar_ficha, 0, 0),
                ("Gestionar BD", "—", self.app._gestionar_bd, 0, 1),
                ("Generar desde carpetas (v2.6)", "v2.6", self.app._lanzar_v26, 1, 0),
                ("Cargar carpeta completa de fichas", "lote",
                 lambda: self.app._cargar_ficha(por_carpetas=True), 1, 1),
            ]
            for titulo, valor, cmd, fila, col in fichas:
                lbl = self._tarjeta_chica(mant, fila, col, titulo, valor, cmd)
                if titulo == "Gestionar BD":
                    self.lbl_total_bd = lbl

            # ------------------------------------------- actividad reciente
            _etiqueta_seccion(cuerpo, "actividad reciente", color=TEXTO_SUAVE,
                              height=16).pack(anchor="w", pady=(20, 8))
            marco, self.tree, pie = _tabla_ttk(
                cuerpo, ("nombre", "fichas", "fecha", "estado"),
                {"nombre": "Submittal", "fichas": "Fichas", "fecha": "Actualizado",
                 "estado": "Estado"},
                {"nombre": 300, "fichas": 66, "fecha": 130, "estado": 96},
                alineados=("fichas", "estado"), alto=4)
            marco.pack(fill="both", expand=True)
            _hover_filas(self.tree)
            self.tree.bind("<Double-Button-1>", lambda _ev: self._abrir_reciente())
            self.lbl_pie = ctk.CTkLabel(pie, text="", text_color=TEXTO_SUAVE,
                                        font=_mono(10))
            self.lbl_pie.pack(anchor="w", padx=12, pady=4)
            self._recientes = []

        # ------------------------------------------------------- tarjetas
        def _tarjeta_grande(self, padre, col, titulo, desc, meta, command,
                            destacada=False):
            fondo = ACENTO_SUAVE if destacada else SUPERFICIE
            hover = ACENTO_SUAVE_H if destacada else SUPERFICIE_2
            borde = ACENTO_BORDE if destacada else BORDE_FUERTE
            t = ctk.CTkFrame(padre, fg_color=fondo, corner_radius=9, border_width=1,
                             border_color=borde)
            t.grid(row=0, column=col, sticky="nsew", padx=(0, 7) if col == 0 else (7, 0))
            fila = ctk.CTkFrame(t, fg_color="transparent")
            fila.pack(fill="x", padx=18, pady=(16, 0))
            ctk.CTkLabel(fila, text=titulo, font=_fuente(16, "bold"),
                        text_color=ACENTO_TXT if destacada else TEXTO_2).pack(side="left")
            ctk.CTkLabel(fila, text="→", font=_fuente(15),
                        text_color=ACENTO if destacada else TEXTO_SUAVE).pack(side="right")
            ctk.CTkLabel(t, text=desc, font=_fuente(12), justify="left",
                        wraplength=330, anchor="w", height=34,
                        text_color=TEXTO_SUAVE).pack(anchor="w", padx=18, pady=(4, 0))
            lbl_meta = ctk.CTkLabel(t, text=meta, font=_mono(10), anchor="w",
                                    height=14,
                                    text_color=ACENTO if destacada else TEXTO_TENUE)
            lbl_meta.pack(anchor="w", padx=18, pady=(6, 16))
            _clicable(t, command, normal=fondo, hover=hover,
                      borde_normal=borde,
                      borde_hover=ACENTO_BORDE_FUERTE if destacada else BORDE_FUERTE)
            return lbl_meta

        def _tarjeta_chica(self, padre, fila, col, titulo, valor, command):
            t = ctk.CTkFrame(padre, fg_color=SUPERFICIE, corner_radius=8,
                             border_width=1, border_color=BORDE)
            t.grid(row=fila, column=col, sticky="nsew",
                   padx=(0, 5) if col == 0 else (5, 0), pady=(0, 10))
            interior = ctk.CTkFrame(t, fg_color="transparent")
            interior.pack(fill="x", padx=14, pady=10)
            ctk.CTkLabel(interior, text=titulo, font=_fuente(13), height=20,
                        text_color=TEXTO_2).pack(side="left")
            lbl = ctk.CTkLabel(interior, text=valor, font=_mono(12),
                               text_color=TEXTO_SUAVE)
            lbl.pack(side="right")
            _clicable(t, command, normal=SUPERFICIE, hover=SUPERFICIE_3,
                      borde_normal=BORDE, borde_hover=BORDE_FUERTE)
            return lbl

        # ------------------------------------------------------- refresco
        def refrescar(self):
            """Repinta conteos y actividad reciente. Tolera que la BD todavia
            no exista (se llama tambien al armar la ventana)."""
            bd = self.bd
            if bd is None:
                return
            try:
                total = bd.resumen_por_categoria()["TOTAL"]
                if self.lbl_total_bd is not None:
                    self.lbl_total_bd.configure(text=str(total))
            except Exception:
                pass
            try:
                proyectos = bd.listar_proyectos()
            except Exception:
                proyectos = []
            self._recientes = proyectos[:6]
            for i in self.tree.get_children():
                self.tree.delete(i)
            for p in self._recientes:
                fecha = (p.get("ultima_actualizacion") or "")[:16].replace("T", " ")
                self.tree.insert("", "end", iid=p["carpeta_bd"],
                                 values=(p.get("nombre_proyecto", ""),
                                         p.get("materiales", 0), fecha or "—",
                                         self._estado_proyecto(p)))
            n = len(proyectos)
            self.lbl_pie.configure(
                text=(f"{n} submittal(s) guardado(s) · doble clic para continuar"
                      if n else "Todavía no hay submittals guardados en la BD"))
            if self._recientes:
                self.lbl_reciente.configure(
                    text="RECIENTE · " +
                         self._recientes[0].get("nombre_proyecto", "")[:34].upper())
            else:
                self.lbl_reciente.configure(text="SIN SUBMITTALS GUARDADOS")

        @staticmethod
        def _estado_proyecto(p):
            """'generado' si ya se armaron los entregables, 'en edición' si no.

            El dato vive en el ``submittal_proyecto.json`` del proyecto (que
            ``listar_proyectos()`` no devuelve), asi que se lee de ahi.
            """
            try:
                ruta = Path(p["carpeta_bd"]) / bd_manager.NOMBRE_SUBMITTAL_JSON
                d = json.loads(ruta.read_text(encoding="utf-8"))
                return "generado" if d.get("entregables_generados") else "en edición"
            except Exception:
                return "—"

        def _abrir_reciente(self):
            sel = self.tree.selection()
            if not sel:
                return
            try:
                proyecto = bd_manager.BDManager.cargar_submittal(sel[0])
            except Exception as e:
                messagebox.showerror("No se pudo abrir", str(e),
                                     parent=self.winfo_toplevel())
                return
            ruta_previa = proyecto.get("ruta_entregables", "")
            destino = ruta_previa if ruta_previa and Path(ruta_previa).is_dir() else ""
            self.app._abrir_submittal(
                proyecto, destino,
                f"Editando: {proyecto.get('nombre_proyecto', '')}")


    class App(ctk.CTk):
        """Ventana principal: shell con menu lateral oscuro fijo y las cuatro
        pantallas del programa adentro (Inicio, Submittal activo, Base de datos,
        Configuración).

        v3.4.0: antes era un menu 2x2 que abria una ventana por flujo. Ahora
        todo pasa en una sola ventana; solo siguen siendo modales los dialogos
        puntuales (cargar ficha, revisar ficha, datos del proyecto, PIN, abrir
        submittal existente), que interrumpen a proposito.
        """

        def __init__(self):
            super().__init__()
            _configurar_estilo_ttk()
            self.title(f"Generador de Submittals ES v{VERSION}")
            self.configure(fg_color=FONDO)
            self._sincronizando = False
            self.modo_dev = False
            self.pantalla_actual = None
            self.submittal = None          # PantallaSubmittal activa (o None)
            self._pantallas = {}
            self._construir()
            # Tamano del diseno (1240x800) como preferido, acotado a la pantalla
            # y redimensionable: en un monitor de 1366x768 la ventana entra
            # igual en vez de quedar con los botones fuera de vista.
            _dimensionar_principal(self, 1240, 800)
            # Se revisa/instala Git ANTES de crear el BDManager: la eleccion de
            # transporte (git vs API REST) se decide una sola vez, al construir
            # el GitSync de adentro. Solo tarda si esta PC no tiene git (primera
            # vez); despues git_disponible() es instantaneo.
            self.lbl_sync.configure(text="🔧 Verificando Git…", text_color=AZUL_ES)
            self.update()
            git_bd.instalar_git_si_falta(logger.info)
            self.bd = bd_manager.BDManager(logger=logger)
            self.protocol("WM_DELETE_WINDOW", self._cerrar_seguro)
            self.after(100, lambda: self._sincronizar(inicial=True))

        # =============================================== armado del shell
        def _construir(self):
            self.grid_rowconfigure(0, weight=1)
            self.grid_columnconfigure(0, minsize=262)   # ancho del menu lateral
            self.grid_columnconfigure(1, weight=1)
            self._construir_sidebar()

            principal = ctk.CTkFrame(self, fg_color=WINDOW_BG, corner_radius=0)
            principal.grid(row=0, column=1, sticky="nsew")
            principal.grid_rowconfigure(2, weight=1)
            principal.grid_columnconfigure(0, weight=1)

            # ------------------------------------------ barra de titulo
            topbar = ctk.CTkFrame(principal, fg_color=SUPERFICIE, corner_radius=0,
                                  height=56)
            topbar.grid(row=0, column=0, sticky="ew")
            topbar.grid_propagate(False)
            self.lbl_titulo = ctk.CTkLabel(topbar, text="Inicio", font=_fuente(14, "bold"),
                                           text_color=TEXTO_2)
            self.lbl_titulo.pack(side="left", padx=(26, 8), pady=16)
            self.lbl_subtitulo = ctk.CTkLabel(topbar, text="", font=_mono(11),
                                              text_color=TEXTO_SUAVE)
            self.lbl_subtitulo.pack(side="left")
            ctk.CTkFrame(principal, height=1, fg_color=BORDE).grid(
                row=1, column=0, sticky="ew")

            # --------------------------------------- contenedor de pantallas
            self.contenedor = ctk.CTkFrame(principal, fg_color="transparent")
            self.contenedor.grid(row=2, column=0, sticky="nsew")

            # ------------------------------------------- barra de estado
            ctk.CTkFrame(principal, height=1, fg_color=BORDE).grid(
                row=3, column=0, sticky="ew")
            pie = ctk.CTkFrame(principal, fg_color=SUPERFICIE, corner_radius=0,
                               height=44)
            pie.grid(row=4, column=0, sticky="ew")
            pie.grid_propagate(False)
            self.lbl_estado = ctk.CTkLabel(pie, text="", text_color=TEXTO_SUAVE,
                                           font=_mono(11))
            self.lbl_estado.pack(side="left", padx=26)
            _boton_secundario(pie, "Cerrar", self._cerrar_seguro, ancho=88,
                              alto=28).pack(side="right", padx=26)

            # La pantalla de inicio se arma ya; las demas, cuando se usan.
            self._pantallas["inicio"] = PantallaInicio(self.contenedor, self, None)
            self._ir("inicio")

        def _construir_sidebar(self):
            lat = ctk.CTkFrame(self, fg_color=SIDEBAR, corner_radius=0, width=262)
            lat.grid(row=0, column=0, sticky="nsw")
            lat.grid_propagate(False)
            lat.grid_rowconfigure(3, weight=1)
            lat.grid_columnconfigure(0, weight=1)

            # ------------------------------------------------------- marca
            marca = ctk.CTkFrame(lat, fg_color="transparent")
            marca.grid(row=0, column=0, sticky="ew", padx=16, pady=(20, 22))
            sello = ctk.CTkFrame(marca, fg_color=ACENTO, corner_radius=7, width=30,
                                 height=30)
            sello.pack(side="left")
            sello.pack_propagate(False)
            ctk.CTkLabel(sello, text="ES", font=_mono(13, "bold"),
                        text_color="white").pack(expand=True)
            txt = ctk.CTkFrame(marca, fg_color="transparent")
            txt.pack(side="left", padx=(11, 0))
            ctk.CTkLabel(txt, text="Generador de Submittals", font=_fuente(13, "bold"),
                        text_color=SIDEBAR_TXT, anchor="w").pack(anchor="w")
            self.lbl_version = ctk.CTkLabel(txt, text=f"v{VERSION}", font=_mono(11),
                                            text_color=SIDEBAR_TXT_3, anchor="w")
            self.lbl_version.pack(anchor="w")

            # -------------------------------------------------------- menu
            menu = ctk.CTkFrame(lat, fg_color="transparent")
            menu.grid(row=1, column=0, sticky="ew", padx=16)
            _etiqueta_seccion(menu, "menú", color=SIDEBAR_LABEL).pack(
                anchor="w", padx=10, pady=(0, 8))
            self._nav = {}
            self._nav_badge = {}
            for clave, etiqueta in (("inicio", "Inicio"),
                                    ("submittal", "Submittal activo"),
                                    ("bd", "Base de datos"),
                                    ("config", "Configuración")):
                self._nav_item(menu, clave, etiqueta)

            # -------------------------------------------------- catalogo
            cat = ctk.CTkFrame(lat, fg_color="transparent")
            cat.grid(row=2, column=0, sticky="ew", padx=16, pady=(22, 0))
            _etiqueta_seccion(cat, "catálogo", color=SIDEBAR_LABEL).pack(
                anchor="w", padx=10, pady=(0, 8))
            self._lbl_cat = {}
            for clave, etiqueta in (("ARQ", "Arquitectura"), ("ESTR", "Estructura"),
                                    ("MEC", "Mecánica"), ("ELEC", "Eléctrica")):
                fila = ctk.CTkFrame(cat, fg_color="transparent")
                fila.pack(fill="x", padx=10, pady=3)
                ctk.CTkLabel(fila, text=etiqueta, font=_fuente(12),
                            text_color=SIDEBAR_TXT_2).pack(side="left")
                v = ctk.CTkLabel(fila, text="—", font=_mono(12),
                                 text_color=SIDEBAR_NUM)
                v.pack(side="right")
                self._lbl_cat[clave] = v

            # -------------------------------------------- pie: sync + extras
            pie = ctk.CTkFrame(lat, fg_color="transparent")
            pie.grid(row=4, column=0, sticky="ews", padx=16, pady=(12, 16))

            tarjeta = ctk.CTkFrame(pie, fg_color=SIDEBAR_CARD, corner_radius=7)
            tarjeta.pack(fill="x")
            cabsync = ctk.CTkFrame(tarjeta, fg_color="transparent")
            cabsync.pack(fill="x", padx=12, pady=(12, 0))
            self.punto_sync = ctk.CTkLabel(cabsync, text="●", font=_fuente(12),
                                           text_color=SIDEBAR_TXT_3)
            self.punto_sync.pack(side="left", padx=(0, 7))
            # ``lbl_sync`` conserva el nombre de v3.3.x: todos los flujos de
            # sincronizacion/actualizacion le escriben el estado ahi.
            self.lbl_sync = _EtiquetaSync(cabsync, text="Iniciando…", font=_fuente(12),
                                          text_color=SIDEBAR_TXT, anchor="w",
                                          wraplength=170, justify="left")
            self.lbl_sync._punto = self.punto_sync
            self.lbl_sync.pack(side="left", fill="x", expand=True)
            self.lbl_sync_meta = ctk.CTkLabel(tarjeta, text="", font=_mono(11),
                                              text_color=SIDEBAR_TXT_3, anchor="w",
                                              wraplength=178, justify="left")
            self.lbl_sync_meta.pack(fill="x", padx=12, pady=(4, 0))
            # La barra de progreso vive en su propio hueco: los flujos hacen
            # ``prog.pack()`` / ``prog.pack_forget()`` y asi siempre aparece en
            # el mismo lugar (entre el estado y los botones), no al final.
            hueco_prog = ctk.CTkFrame(tarjeta, fg_color="transparent", height=10)
            hueco_prog.pack(fill="x", padx=12)
            self.prog = ctk.CTkProgressBar(hueco_prog, height=4, corner_radius=2,
                                           mode="indeterminate", progress_color=ACENTO,
                                           fg_color=SIDEBAR_BTN)
            botsync = ctk.CTkFrame(tarjeta, fg_color="transparent")
            botsync.pack(fill="x", padx=12, pady=(10, 12))
            ctk.CTkButton(botsync, text="Sincronizar", command=self._sincronizar,
                          fg_color=SIDEBAR_BTN, hover_color=SIDEBAR_BTN_H,
                          text_color=SIDEBAR_TXT, font=_fuente(11), height=28,
                          corner_radius=5, width=1).pack(side="left", fill="x",
                                                         expand=True, padx=(0, 6))
            ctk.CTkButton(botsync, text="Subir", command=self._subir_pendientes,
                          fg_color="transparent", hover_color=SIDEBAR_HOVER,
                          text_color=SIDEBAR_TXT_2, font=_fuente(11), height=28,
                          corner_radius=5, width=1, border_width=1,
                          border_color=SIDEBAR_BORDE).pack(side="left", fill="x",
                                                           expand=True)

            extras = ctk.CTkFrame(pie, fg_color="transparent")
            extras.pack(fill="x", pady=(12, 0))
            # "Generar desde carpetas (v2.6)" vive en las tarjetas de Inicio
            # (como en el diseno); aca solo quedan las dos acciones que no
            # tienen otro lugar.
            ctk.CTkButton(extras, text="Buscar actualización",
                          command=self._buscar_update, anchor="w",
                          fg_color="transparent", hover_color=SIDEBAR_HOVER,
                          text_color=SIDEBAR_TXT_2, font=_fuente(12), height=28,
                          corner_radius=5).pack(fill="x")
            # ``btn_modo_dev`` conserva el nombre: ``_toggle_modo_dev`` lo repinta.
            self.btn_modo_dev = ctk.CTkButton(
                extras, text="Modo desarrollador", command=self._toggle_modo_dev,
                anchor="w", fg_color="transparent", hover_color=SIDEBAR_HOVER,
                text_color=SIDEBAR_TXT_3, font=_fuente(12), height=28,
                corner_radius=5)
            self.btn_modo_dev.pack(fill="x")

        def _nav_item(self, padre, clave, etiqueta):
            """Item del menu lateral: etiqueta a la izquierda y, a la derecha,
            un contador (fichas de la BD) o una insignia (submittal activo)."""
            fila = ctk.CTkFrame(padre, fg_color="transparent", corner_radius=6,
                                height=34)
            fila.pack(fill="x", pady=1)
            fila.pack_propagate(False)
            lbl = ctk.CTkLabel(fila, text=etiqueta, font=_fuente(13),
                               text_color=SIDEBAR_TXT_2, anchor="w")
            lbl.pack(side="left", padx=(10, 0))
            badge = ctk.CTkLabel(fila, text="", font=_mono(11),
                                 text_color=SIDEBAR_TXT_3, corner_radius=3)
            badge.pack(side="right", padx=(0, 10))
            self._nav[clave] = (fila, lbl)
            self._nav_badge[clave] = badge
            _clicable(fila, lambda: self._ir(clave),
                      normal="transparent", hover=SIDEBAR_HOVER)

        # ================================================== navegacion
        _TITULOS = {"inicio": ("Inicio", ""),
                    "submittal": ("Submittal activo", ""),
                    "bd": ("Base de datos", ""),
                    "config": ("Configuración", "")}

        def _ir(self, clave):
            """Muestra una pantalla (creandola la primera vez)."""
            if clave == "submittal" and self.submittal is None:
                # Sin submittal abierto no hay nada que mostrar: se ofrece
                # empezar uno en vez de dejar el clic sin respuesta.
                if messagebox.askyesno(
                        "Sin submittal activo",
                        "Todavía no hay un submittal abierto.\n\n"
                        "¿Generar uno nuevo desde la Base de Datos?",
                        parent=self):
                    self._generar_desde_bd()
                return
            if clave == "bd" and "bd" not in self._pantallas:
                self._pantallas["bd"] = PantallaBD(self.contenedor, self.bd, app=self,
                                                   al_cambiar=self._actualizar_estado)
            if clave == "config":
                # Se rearma cada vez: muestra el estado real de la API key, del
                # token y de los conteos en el momento de abrirla.
                anterior = self._pantallas.pop("config", None)
                if anterior is not None:
                    anterior.destroy()
                self._pantallas["config"] = PantallaConfig(
                    self.contenedor, self.bd, seccion_inicial=self._seccion_config,
                    app=self)
            pantalla = self._pantallas.get(clave)
            if pantalla is None:
                return
            if self.pantalla_actual is not None:
                self.pantalla_actual.pack_forget()
            pantalla.pack(fill="both", expand=True)
            self.pantalla_actual = pantalla
            self._clave_actual = clave
            if clave == "bd":
                pantalla.al_mostrar()
            elif clave == "inicio":
                pantalla.refrescar()
            self._pintar_nav(clave)
            self._pintar_titulo(clave)

        _seccion_config = "openai"

        def _pintar_nav(self, activa):
            for clave, (fila, lbl) in self._nav.items():
                encendido = (clave == activa)
                fila.configure(fg_color=ACENTO if encendido else "transparent")
                lbl.configure(text_color="white" if encendido else SIDEBAR_TXT_2,
                              font=_fuente(13, "bold") if encendido else _fuente(13))
                # El contador tambien tiene que leerse sobre el rojo del item
                # activo (en gris tenue quedaba por debajo de 4.5:1).
                badge = self._nav_badge[clave]
                if badge.cget("fg_color") in ("transparent", None):
                    badge.configure(text_color="white" if encendido
                                    else SIDEBAR_TXT_3)
                _clicable(fila, lambda c=clave: self._ir(c),
                          normal=ACENTO if encendido else "transparent",
                          hover=ACENTO_HOVER if encendido else SIDEBAR_HOVER)

        def _pintar_titulo(self, clave):
            titulo, sub = self._TITULOS.get(clave, (clave, ""))
            if clave == "submittal" and self.submittal is not None:
                sub = "· " + self.submittal.proyecto.get("nombre_proyecto", "")
            elif clave == "bd":
                try:
                    sub = f"· {self.bd.resumen_por_categoria()['TOTAL']} fichas"
                except Exception:
                    sub = ""
            self.lbl_titulo.configure(text=titulo)
            self.lbl_subtitulo.configure(text=sub)

        def _abrir_submittal(self, proyecto, destino, titulo):
            """Punto unico de entrada a la pantalla de submittal (la usan tanto
            'Generar desde BD' como 'Abrir existente').

            Si ya habia un submittal abierto se guarda su avance antes de
            reemplazarlo: es la misma garantia que daba el cierre de la ventana
            en v3.3.x.
            """
            if self.submittal is not None:
                if not self.submittal.guardar_al_salir():
                    return
                if self.pantalla_actual is self.submittal:
                    self.submittal.pack_forget()
                    self.pantalla_actual = None
                self.submittal.destroy()
                self._pantallas.pop("submittal", None)
            self.submittal = PantallaSubmittal(self.contenedor, self.bd, proyecto,
                                               destino, titulo, app=self)
            self._pantallas["submittal"] = self.submittal
            self._ir("submittal")
            self._actualizar_estado()

        # ------------------------------------------------ sincronizacion
        def _sincronizar(self, inicial=False):
            """Flujo 1: ``git pull`` en segundo plano para no congelar la GUI."""
            if self._sincronizando:
                return
            self._sincronizando = True
            self.lbl_sync.configure(text="🔄 Sincronizando con GitHub…", text_color=AZUL_ES)
            self.prog.pack(pady=(0, 6))
            self.prog.start()

            def trabajo():
                try:
                    _data, resumen = self.bd.sync_indice()
                except Exception as e:      # no debería ocurrir: sync_indice absorbe
                    resumen = {"error": str(e)}
                self.after(0, lambda: self._fin_sincronizar(resumen, inicial))

            threading.Thread(target=trabajo, daemon=True).start()

        def _fin_sincronizar(self, resumen, inicial=False):
            self.prog.stop()
            self.prog.pack_forget()
            self._sincronizando = False
            self._actualizar_estado()

            if resumen.get("conflictos"):
                self.lbl_sync.configure(
                    text=f"✅ Conflicto resuelto y sincronizado "
                         f"({resumen['conflictos']} archivo(s))", text_color=VERDE_OK)
                messagebox.showinfo(
                    "Conflicto resuelto",
                    "Otra computadora había subido cambios a la vez.\n\n"
                    "Se fusionaron automáticamente sin perder datos: se "
                    "conservaron las fichas de ambos lados.", parent=self.winfo_toplevel())
                return
            if resumen.get("indice_invalido"):
                self.lbl_sync.configure(text="⚠️ Índice con problemas: usando caché local",
                                        text_color=ROJO_ES)
                messagebox.showwarning(
                    "Índice inconsistente",
                    "El índice descargado no pasó la validación; se está usando la "
                    "copia local.\n\nDetalle:\n- " +
                    "\n- ".join(resumen["indice_invalido"][:6]), parent=self.winfo_toplevel())
                return
            if resumen.get("auth") and inicial:
                self.lbl_sync.configure(text="🔑 Sin token de GitHub (solo lectura)",
                                        text_color=ROJO_ES)
                return
            if resumen.get("offline"):
                self.lbl_sync.configure(text="📡 Sin conexión — trabajando con la copia local",
                                        text_color=ROJO_ES)
                return
            if resumen.get("error"):
                self.lbl_sync.configure(text=f"⚠️ {resumen['error'][:70]}", text_color=ROJO_ES)
                return
            self.lbl_sync.configure(text=self.bd.texto_estado_sync(), text_color=VERDE_OK)

        def _subir_pendientes(self):
            if not self.bd.hay_cambios_sin_subir():
                messagebox.showinfo("Sincronización", "No hay cambios pendientes de subir.", parent=self.winfo_toplevel())
                return
            self.lbl_sync.configure(text="🔄 Subiendo cambios…", text_color=AZUL_ES)
            self.update_idletasks()
            r = self.bd.git_push("subir cambios pendientes")
            self._reportar_push(r)

        def _reportar_push(self, r):
            if r.get("subido"):
                extra = (f" ({r['conflictos']} conflicto(s) resuelto(s))"
                         if r.get("conflictos") else "")
                self.lbl_sync.configure(text=f"☁️ Cambios subidos a GitHub{extra}",
                                        text_color=VERDE_OK)
            elif r.get("offline"):
                self.lbl_sync.configure(text="📡 Sin conexión — se subirán al reconectar",
                                        text_color=ROJO_ES)
            elif r.get("auth"):
                self.lbl_sync.configure(text="🔑 Falta el token de GitHub", text_color=ROJO_ES)
                if messagebox.askyesno("Token requerido",
                                       "Para subir cambios hace falta un token de "
                                       "GitHub.\n\n¿Configurarlo ahora?", parent=self.winfo_toplevel()):
                    self._config_github()
            elif r.get("nada_que_subir"):
                self.lbl_sync.configure(text="Sin cambios por subir", text_color=GRIS_TEXTO_SUAVE)
            else:
                self.lbl_sync.configure(text=f"⚠️ {str(r.get('error', ''))[:70]}",
                                        text_color=ROJO_ES)

        def _configuracion(self, tab_inicial="openai"):
            """Va a la pantalla de configuración (OpenAI + GitHub + rutas)."""
            self._seccion_config = tab_inicial
            self._ir("config")

        def _config_github(self):
            """Atajo que abre la configuración directamente en la sección de
            sincronización (lo usa el aviso de 'falta el token' al subir)."""
            self._configuracion(tab_inicial="github")

        def _tras_guardar_config(self, cambio_github):
            """Lo llama ``PantallaConfig`` al guardar: vuelve al inicio y, si se
            tocó algo de GitHub, vuelve a sincronizar (igual que cuando la
            configuración era una ventana modal)."""
            self._ir("inicio")
            self._actualizar_estado()
            if cambio_github:
                self._sincronizar()

        def _actualizar_estado(self):
            """Refresca los conteos del menu lateral, la barra de estado y la
            insignia del submittal activo."""
            res = self.bd.resumen_por_categoria()
            cache = "  ·  usando caché anterior" if self.bd.usando_cache else ""
            pend = self.bd.pendientes
            sin_subir = f"  ·  {len(pend)} cambio(s) sin subir" if pend else ""
            self.lbl_estado.configure(
                text=f"BD: {res['TOTAL']} fichas  ·  ARQ {res['ARQ']} · "
                     f"ESTR {res['ESTR']} · MEC {res['MEC']} · "
                     f"ELEC {res['ELEC']}{cache}{sin_subir}",
                text_color=(AMBAR if (cache or sin_subir) else TEXTO_SUAVE))

            for clave, lbl in self._lbl_cat.items():
                lbl.configure(text=str(res.get(clave, 0)))
            activa = getattr(self, "_clave_actual", "inicio")
            self._nav_badge["bd"].configure(
                text=str(res["TOTAL"]), fg_color="transparent",
                text_color=("white" if activa == "bd" else SIDEBAR_TXT_3))
            n = (len(self.submittal.tabla.materiales)
                 if self.submittal is not None else 0)
            self._nav_badge["submittal"].configure(
                text=(str(n) if self.submittal is not None else ""),
                fg_color=(ACENTO if n else "transparent"),
                text_color=("white" if n else SIDEBAR_TXT_3),
                width=(22 if n else 0))
            self.lbl_sync_meta.configure(
                text=f"{len(pend)} pendiente(s)" if pend else "0 pendientes")
            if "inicio" in self._pantallas:
                self._pantallas["inicio"].refrescar()

        # -------- modo desarrollador
        def _toggle_modo_dev(self):
            if self.modo_dev:
                self.modo_dev = False
                self.btn_modo_dev.configure(text="Modo desarrollador",
                                            fg_color="transparent",
                                            text_color=SIDEBAR_TXT_3)
                self.title(f"Generador de Submittals ES v{VERSION}")
                return

            d = _PinDialog(self, "Modo desarrollador", "Ingrese el PIN:")
            self.wait_window(d)
            pin = d.resultado
            if pin is None:
                return
            if pin != PIN_MODO_DEV:
                messagebox.showerror("PIN incorrecto", "El PIN ingresado no es correcto.", parent=self.winfo_toplevel())
                return

            self.modo_dev = True
            self.btn_modo_dev.configure(text="Modo desarrollador: ACTIVO",
                                        fg_color=ACENTO, text_color="white")
            self.title(f"Generador de Submittals ES v{VERSION} — MODO DESARROLLADOR")
            messagebox.showinfo(
                "Modo desarrollador activado",
                "Modo desarrollador activado.\n\n"
                "Al generar un submittal desde BD ya no se pedirán los datos del "
                "proyecto: se completan automáticamente con datos de prueba.\n\n"
                "Use este modo solo para pruebas, nunca para submittals reales.", parent=self.winfo_toplevel())

        # -------- flujos
        def _pedir_datos_proyecto(self):
            if self.modo_dev:
                return {
                    "numero_procedimiento": "PRUEBA-0000",
                    "institucion": "Institución de prueba",
                    "detalle": "Submittal de prueba (modo desarrollador)",
                    "plazo": "N/A",
                    "monto": "0",
                }
            d = DatosProyectoDialog(self)
            self.wait_window(d)
            return d.resultado

        def _generar_desde_bd(self):
            destino = filedialog.askdirectory(title="Carpeta destino del submittal")
            if not destino:
                return
            # Si esta carpeta ya tiene un submittal guardado (con materiales),
            # "Generar desde BD" NO debe pisarlo con uno vacio: eso es lo que
            # reseteaba el contador de consecutivos al "reabrir" un proyecto
            # por aca en vez de con "Abrir submittal existente". Se ofrece
            # continuar el que ya existe.
            existente_path = Path(destino) / bd_manager.NOMBRE_SUBMITTAL_JSON
            if existente_path.exists():
                try:
                    existente = bd_manager.BDManager.cargar_submittal(destino)
                except Exception:
                    existente = None
                n = len((existente or {}).get("materiales_seleccionados", []))
                if existente and n:
                    continuar = messagebox.askyesno(
                        "Ya hay un submittal en esta carpeta",
                        f"Esta carpeta ya tiene un submittal guardado con {n} "
                        "material(es).\n\n"
                        "¿Continuar ESE submittal? Es lo recomendado: si empieza "
                        "uno nuevo en esta misma carpeta, los consecutivos "
                        "arrancan de nuevo desde 1 y corre riesgo de pisar lo "
                        "ya guardado.\n\n"
                        "Sí = continuar el existente.\n"
                        "No = empezar uno nuevo vacío en esta misma carpeta.",
                        parent=self.winfo_toplevel())
                    if continuar:
                        self._abrir_submittal(
                            existente, destino,
                            f"Editando: {existente.get('nombre_proyecto', '')}")
                        return
            datos = self._pedir_datos_proyecto()
            if not datos:
                return
            proyecto = {"nombre_proyecto": Path(destino).name,
                        "datos_procedimiento": datos, "tipo_caratula": "clasica",
                        "materiales_seleccionados": []}
            self._abrir_submittal(proyecto, destino, "Generar submittal desde BD")

        def _abrir_existente(self):
            """Abre un submittal ya guardado, ya sea eligiendolo de la lista
            sincronizada en la BD (para continuar el trabajo de otra PC sin
            tener acceso a su carpeta) o buscando una carpeta local."""
            d = _SelectorProyectoNube(self, self.bd)
            self.wait_window(d)
            proyecto = d.resultado
            if not proyecto:
                return
            # La carpeta de entregables de quien guardo por ultima vez puede
            # ser de OTRA PC (no existe aqui); solo se reutiliza como destino
            # si de verdad existe en este equipo, si no se deja vacia y se
            # pide al generar (igual que "Generar desde BD").
            ruta_previa = proyecto.get("ruta_entregables", "")
            destino = ruta_previa if ruta_previa and Path(ruta_previa).is_dir() else ""
            self._abrir_submittal(proyecto, destino,
                                  f"Editando: {proyecto.get('nombre_proyecto', '')}")

        def _cargar_ficha(self, por_carpetas=False):
            """Abre el diálogo de carga de fichas. Con ``por_carpetas=True``
            arranca directo en la selección de carpetas (carga masiva)."""
            v = VentanaCargarFicha(self, self.bd, al_terminar=self._actualizar_estado)
            if por_carpetas:
                v.after(200, v._seleccionar_carpetas)

        def _gestionar_bd(self):
            self._ir("bd")

        def _lanzar_v26(self):
            import subprocess
            if getattr(sys, "frozen", False):
                # Empaquetado: no hay interprete de Python al que pasarle un
                # .py (sys.executable ES este mismo .exe). Se busca el .exe de
                # v2.6 compilado aparte, en la misma carpeta.
                exe_v26 = Path(sys.executable).resolve().parent / "GeneradorSubmittalsES.exe"
                if exe_v26.exists():
                    subprocess.Popen([str(exe_v26)], cwd=str(exe_v26.parent),
                                     creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
                else:
                    messagebox.showinfo(
                        "v2.6", "No se encontró GeneradorSubmittalsES.exe (v2.6).\n\n"
                        "Coloque ese archivo en la misma carpeta que este programa "
                        "para poder abrirlo.", parent=self.winfo_toplevel())
                return
            ruta = BASE_DIR / "submitals_gui.py"
            if ruta.exists():
                subprocess.Popen([sys.executable, str(ruta)], cwd=str(BASE_DIR),
                                 creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
            else:
                messagebox.showinfo("v2.6", "No se encontró submitals_gui.py", parent=self.winfo_toplevel())

        def _buscar_update(self):
            info = updater_gh.hay_actualizacion(logf=logger.info)
            if info.get("error"):
                messagebox.showinfo("Actualización", info["error"], parent=self.winfo_toplevel()); return
            if not info.get("disponible"):
                messagebox.showinfo("Actualización", "No hay actualizaciones.", parent=self.winfo_toplevel()); return
            if not messagebox.askyesno("Actualización disponible",
                                       f"Nueva versión {info.get('version_remota')}.\n"
                                       f"{info.get('changelog', '')}\n\n"
                                       "Se actualizará el programa y la Base de Datos.\n"
                                       "¿Aplicar ahora?", parent=self.winfo_toplevel()):
                return
            # Modo empaquetado con .exe nuevo: aplicar_y_sincronizar() ignora el
            # .exe a proposito (solo reemplaza .py/.html en vivo), asi que sin
            # esta rama el boton nunca descargaba/instalaba el ejecutable nuevo
            # y "no hay actualizaciones" seguia apareciendo para siempre.
            if info.get("requiere_exe"):
                self._aplicar_update_exe(info)
                return
            ok, msg, reinicio, _bd = updater_gh.aplicar_y_sincronizar(
                info, bd=self.bd, logf=logger.info)
            messagebox.showinfo("Actualización", msg, parent=self.winfo_toplevel())
            self._actualizar_estado()
            if ok and reinicio:
                updater_gh.reiniciar()

        def _aplicar_update_exe(self, info):
            """Descarga el .exe nuevo en segundo plano y prepara el swap: al
            cerrar la app, un .bat lo reemplaza y la reabre solo."""
            self.lbl_sync.configure(text="⬇️ Descargando actualización…", text_color=AZUL_ES)
            self.prog.pack(pady=(0, 6))
            self.prog.start()

            def _prog(_pct, texto):
                self.after(0, lambda: self.lbl_sync.configure(text=f"⬇️ {texto}"))

            def trabajo():
                ok, msg = updater_gh.preparar_exe(info, progreso=_prog, logf=logger.info)
                self.after(0, lambda: self._fin_update_exe(ok, msg))

            threading.Thread(target=trabajo, daemon=True).start()

        def _fin_update_exe(self, ok, msg):
            self.prog.stop()
            self.prog.pack_forget()
            if not ok:
                messagebox.showerror("Actualización", msg, parent=self.winfo_toplevel())
                self._actualizar_estado()
                return
            try:
                self.bd.sync_indice()
            except Exception:
                pass
            messagebox.showinfo(
                "Actualización descargada",
                msg + "\n\nEl programa se cerrará y reabrirá con la nueva versión.", parent=self.winfo_toplevel())
            updater_gh.lanzar_swap()

        def _cerrar_seguro(self):
            """Cierre seguro: ofrece subir lo que quedó pendiente.

            Ya no hay lock que liberar; lo único que puede quedar a medias es un
            cambio local sin subir (por ejemplo si se trabajó sin conexión).
            """
            # Antes de cualquier otra cosa: el submittal en pantalla se guarda
            # SIEMPRE (es lo que hacia el cierre de su ventana en v3.3.x).
            if self.submittal is not None:
                if not self.submittal.guardar_al_salir():
                    return
            try:
                pendiente = self.bd.hay_cambios_sin_subir()
            except Exception:
                pendiente = False
            if pendiente:
                r = messagebox.askyesnocancel(
                    "Cambios sin subir",
                    "Hay cambios en la BD que todavía no están en GitHub.\n\n"
                    "¿Subirlos antes de cerrar?", parent=self.winfo_toplevel())
                if r is None:
                    return
                if r:
                    self.lbl_sync.configure(text="🔄 Subiendo cambios…", text_color=AZUL_ES)
                    self.update_idletasks()
                    res = self.bd.git_push("subir cambios antes de cerrar")
                    if not res.get("subido") and not res.get("nada_que_subir"):
                        if not messagebox.askyesno(
                                "No se pudo subir",
                                f"{res.get('error', 'Error desconocido')}\n\n"
                                "Los cambios están guardados localmente y se "
                                "subirán la próxima vez.\n\n¿Cerrar de todos modos?", parent=self.winfo_toplevel()):
                            return
            self.destroy()


    class PantallaBD(ctk.CTkFrame):
        """Pantalla 'Base de datos': gestionar la BD (buscar, filtrar, editar,
        duplicar, desactivar/reactivar, reemplazar PDF, vista previa).

        v3.2.0: la columna principal es el NOMBRE DESCRIPTIVO completo, y se
        puede EDITAR una ficha (regenerando su nombre). Poder corregir es lo que
        faltaba; por eso el borrado sigue siendo lógico y reversible.

        v3.4.0: era ``VentanaGestionarBD`` (ventana aparte); ahora es una
        pantalla de la ventana principal. Las acciones y sus confirmaciones no
        cambiaron; solo se reordenaron segun el diseno (las de edicion arriba a
        la derecha, las de estado abajo a la izquierda).
        """

        def __init__(self, master, bd, al_cambiar=None, app=None):
            super().__init__(master, fg_color="transparent")
            self.bd = bd; self.al_cambiar = al_cambiar
            self.app = app
            self._migracion_ofrecida = False
            cuerpo = ctk.CTkFrame(self, fg_color="transparent")
            cuerpo.pack(fill="both", expand=True, padx=26, pady=(22, 0))

            enc = ctk.CTkFrame(cuerpo, fg_color="transparent")
            enc.pack(fill="x", pady=(0, 14))
            titulos = ctk.CTkFrame(enc, fg_color="transparent")
            titulos.pack(side="left")
            ctk.CTkLabel(titulos, text="Base de datos de fichas",
                        font=_fuente(20, "bold"), text_color=TEXTO,
                        anchor="w").pack(anchor="w")
            ctk.CTkLabel(titulos, text="Doble clic sobre una fila para editarla.",
                        font=_fuente(12), text_color=TEXTO_SUAVE,
                        anchor="w").pack(anchor="w", pady=(4, 0))
            acciones = ctk.CTkFrame(enc, fg_color="transparent")
            acciones.pack(side="right")
            _boton_secundario(acciones, "Vista previa", self._vista_previa,
                              ancho=112).pack(side="left")
            _boton_secundario(acciones, "Reemplazar PDF", self._reemplazar_pdf,
                              ancho=132).pack(side="left", padx=(8, 0))
            _boton_secundario(acciones, "Duplicar", self._duplicar, ancho=94).pack(
                side="left", padx=(8, 0))
            _boton(acciones, "Editar ficha", self._editar, ancho=124).pack(
                side="left", padx=(8, 0))

            # Mismo buscador que al armar un submittal, con la mejor coincidencia
            # de primera, filtros y scroll; aqui ademas puede mostrar las fichas
            # desactivadas (para reactivarlas).
            self.buscador = _BuscadorFichas(cuerpo, self.bd,
                                            on_activar=lambda _f: self._editar(),
                                            permitir_inactivas=True, alto_filas=14)
            self.buscador.pack(fill="both", expand=True)

            bar = ctk.CTkFrame(cuerpo, fg_color="transparent")
            bar.pack(fill="x", pady=(12, 16))
            _boton_peligro(bar, "Desactivar", self._eliminar, ancho=112).pack(side="left")
            _boton_secundario(bar, "Reactivar", self._reactivar, ancho=104).pack(
                side="left", padx=(8, 0))
            self.lbl_resumen = ctk.CTkLabel(bar, text="", text_color=TEXTO_SUAVE,
                                            font=_mono(11))
            self.lbl_resumen.pack(side="right")
            self._pintar_resumen()

        def al_mostrar(self):
            """La ventana principal llama a esto cada vez que se entra a la
            pantalla: refresca el listado y, la primera vez, ofrece generar los
            nombres de las fichas viejas (antes se preguntaba al abrir la
            ventana)."""
            self._refrescar()
            if not self._migracion_ofrecida:
                self._migracion_ofrecida = True
                self._ofrecer_migracion()

        def _pintar_resumen(self):
            r = self.bd.resumen_por_categoria()
            self.lbl_resumen.configure(
                text=f"{r['TOTAL']} activas · ARQ {r['ARQ']} · ESTR {r['ESTR']} · "
                     f"MEC {r['MEC']} · ELEC {r['ELEC']}")

        # ------------------------------------------------------------ listado
        def _refrescar(self):
            self.buscador.refrescar()
            self._pintar_resumen()

        def _sel(self):
            f = self.buscador.ficha_seleccionada()
            if not f:
                messagebox.showinfo("Seleccione una ficha",
                                    "Elija una ficha de la lista.", parent=self.winfo_toplevel())
                return None
            return f

        def _ofrecer_migracion(self):
            """Fichas de versiones anteriores sin nombre descriptivo."""
            pendientes = self.bd.fichas_sin_nombre()
            if not pendientes:
                return
            if not messagebox.askyesno(
                    "Generar nombres",
                    f"Hay {len(pendientes)} ficha(s) sin nombre descriptivo "
                    "(cargadas con una versión anterior).\n\n"
                    "¿Generar sus nombres ahora?", parent=self.winfo_toplevel()):
                return
            # Sincronizar primero: si otra PC editó esas fichas, se trabaja sobre
            # la versión al día en vez de pisarla.
            self.bd.sincronizar()
            n = self.bd.migrar_nombres_ficha()
            r = self.bd.git_push(f"generar nombre de {n} ficha(s)")
            self._refrescar()
            self._avisar_push(r, f"Nombres generados para {n} ficha(s).")

        # ------------------------------------------------------------ acciones
        def _editar(self):
            f = self._sel()
            if not f:
                return
            datos = dict(f)
            d = DialogoRevisarFicha(self, self.bd, None, datos,
                                    titulo=f"Editar: {self.bd.nombre_de(f)}",
                                    es_edicion=True)
            self.wait_window(d)
            res = d.resultado
            if res.get("accion") != "guardar":
                return
            cambios = {k: v for k, v in res["datos"].items()
                       if not k.startswith("_") and k in
                       ("nombre_material", "marca", "categoria", "dimensiones",
                        "sin_medidas", "tipo_producto", "especificacion",
                        "normativa", "descripcion_corta", "aspectos_adicionales",
                        "sinonimos")}
            manual = res["datos"].get("nombre_ficha", "")
            if manual:
                cambios["nombre_ficha"] = manual
            try:
                # Sin nombre manual se regenera SIEMPRE: es lo que hace efectivo
                # el botón "Regenerar" incluso sobre una ficha que antes tenía el
                # nombre escrito a mano.
                ficha = self.bd.editar_ficha(f["id"], cambios,
                                             regenerar_nombre=not bool(manual))
            except Exception as e:
                messagebox.showerror("No se pudo editar", str(e), parent=self.winfo_toplevel()); return
            r = self.bd.git_push(f"editar ficha {ficha['nombre_ficha']}")
            self._refrescar()
            self._avisar_push(r, f"Ficha actualizada:\n{ficha['nombre_ficha']}")
            if self.al_cambiar:
                self.al_cambiar()

        def _vista_previa(self):
            f = self._sel()
            if not f:
                return
            _vista_previa_ficha(self, self.bd, f)

        def _reemplazar_pdf(self):
            f = self._sel()
            if not f:
                return
            ruta = filedialog.askopenfilename(
                title=f"PDF correcto para {self.bd.nombre_de(f)}",
                filetypes=[("Fichas", "*.pdf *.png *.jpg *.jpeg *.tif *.tiff"),
                           ("Todos", "*.*")])
            if not ruta:
                return
            try:
                self.bd.reemplazar_pdf_ficha(f["id"], ruta)
            except Exception as e:
                messagebox.showerror("No se pudo reemplazar", str(e), parent=self.winfo_toplevel()); return
            r = self.bd.git_push(f"reemplazar archivo de {self.bd.nombre_de(f)}")
            self._avisar_push(r, "Archivo reemplazado. El nombre y las referencias "
                                 "de los submittals se conservan.")
            self._refrescar()
            if self.al_cambiar:
                self.al_cambiar()

        def _duplicar(self):
            """Crea una ficha nueva para OTRA especificacion reutilizando el PDF
            de la ficha seleccionada, sin volver a subir/leer el archivo.

            Pensado para proveedores (ej. METALCO) que documentan varias medidas
            en un solo PDF: en vez de re-subir y re-procesar el mismo archivo por
            cada medida, se duplica la ficha cambiando solo la especificacion."""
            f = self._sel()
            if not f:
                return
            if f.get("estado", "activo") != "activo":
                messagebox.showinfo(
                    "Duplicar ficha",
                    "Reactive la ficha antes de duplicarla: solo se duplican "
                    "fichas activas.", parent=self.winfo_toplevel())
                return
            # Se pre-llena el mismo diálogo de revisión con los datos de la ficha
            # original (SIN su id/ruta/hash: es una ficha nueva). El usuario
            # cambia lo que distingue la variante -- casi siempre las medidas o
            # la especificación -- y el nombre se regenera solo.
            base = {k: f.get(k, "") for k in
                    ("nombre_material", "marca", "categoria", "dimensiones",
                     "tipo_producto", "especificacion", "normativa",
                     "descripcion_corta", "aspectos_adicionales", "sinonimos")}
            base["sin_medidas"] = bool(f.get("sin_medidas"))
            d = DialogoRevisarFicha(
                self, self.bd, None, base,
                titulo=f"Duplicar con el mismo PDF de: {self.bd.nombre_de(f)}",
                es_edicion=True)
            self.wait_window(d)
            if d.resultado.get("accion") != "guardar":
                return
            try:
                nueva = self.bd.duplicar_ficha(f["id"], d.resultado["datos"])
            except Exception as e:
                messagebox.showerror("No se pudo duplicar", str(e),
                                     parent=self.winfo_toplevel()); return
            r = self.bd.git_push(
                f"duplicar ficha {nueva['nombre_ficha']} "
                f"(mismo PDF que {self.bd.nombre_de(f)})")
            self._refrescar()
            self._avisar_push(
                r, f"Ficha duplicada:\n{nueva['nombre_ficha']}\n\n"
                   "Reutiliza el PDF de la ficha original (no se subió de nuevo).")
            if self.al_cambiar:
                self.al_cambiar()

        def _eliminar(self):
            f = self._sel()
            if not f:
                return
            nombre = self.bd.nombre_de(f)
            en_uso = self.bd.proyectos_que_usan(f["id"])
            aviso = ""
            if en_uso:
                detalle = "\n".join(f"  · {p['nombre_proyecto']} ({p['consecutivo']})"
                                    for p in en_uso[:8])
                aviso = ("\n\n⚠️ Esta ficha se usa en "
                         f"{len(en_uso)} submittal(s):\n{detalle}\n"
                         "Esos submittals no podrán regenerarse hasta que se "
                         "cambie el material.")
            if not messagebox.askyesno(
                    "Desactivar ficha",
                    f"¿Desactivar '{nombre}'?\n\n"
                    "Dejará de aparecer en las búsquedas. El PDF NO se borra y la "
                    "ficha se puede reactivar después." + aviso, parent=self.winfo_toplevel()):
                return
            if not self.bd.soft_delete_ficha(f["id"]):
                messagebox.showerror(
                    "No se pudo desactivar",
                    "La ficha ya no está en el índice. Sincronice y vuelva a "
                    "intentar.", parent=self.winfo_toplevel())
                self._refrescar()
                return
            r = self.bd.git_push(f"desactivar ficha {nombre}")
            self._refrescar()
            self._avisar_push(r, "Ficha desactivada.")
            if self.al_cambiar:
                self.al_cambiar()

        def _reactivar(self):
            f = self._sel()
            if not f:
                return
            if f.get("estado", "activo") == "activo":
                messagebox.showinfo("Reactivar", "La ficha ya está activa.", parent=self.winfo_toplevel())
                return
            if not self.bd.reactivar_ficha(f["id"]):
                messagebox.showerror(
                    "No se pudo reactivar",
                    "La ficha ya no está en el índice. Sincronice y vuelva a "
                    "intentar.", parent=self.winfo_toplevel())
                self._refrescar()
                return
            r = self.bd.git_push(f"reactivar ficha {self.bd.nombre_de(f)}")
            self._refrescar()
            self._avisar_push(r, "Ficha reactivada.")
            if self.al_cambiar:
                self.al_cambiar()

        def _avisar_push(self, r, mensaje_ok):
            if r.get("offline"):
                messagebox.showinfo("Sin conexión",
                                    mensaje_ok + "\n\nEl cambio se guardó localmente "
                                    "y se subirá al reconectar.", parent=self.winfo_toplevel())
            elif r.get("auth"):
                messagebox.showwarning("Falta el token",
                                       mensaje_ok + "\n\nNo se pudo subir a GitHub: "
                                       "configure el token.", parent=self.winfo_toplevel())
            elif r.get("subido") or r.get("nada_que_subir") or r.get("desactivado"):
                messagebox.showinfo("Listo", mensaje_ok, parent=self.winfo_toplevel())
            else:
                messagebox.showwarning("No se pudo subir",
                                       mensaje_ok + f"\n\n{r.get('error', '')}", parent=self.winfo_toplevel())


def main():
    if not _TK_OK:
        print("tkinter no disponible en este entorno; la GUI no puede iniciarse.")
        return
    App().mainloop()


if __name__ == "__main__":
    main()
