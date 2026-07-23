#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
================================================================================
 GENERADOR DE SUBMITTALS - ES CONSTRUCTORA  (submitals_gui.py)  v2.6.6
 Elaborado por Adrian Castro
================================================================================
 v2.6.6 sobre v2.6.5:
   1. Fix: la columna "Descripcion" del Excel (Guia Materiales.xlsx) mostraba
      el texto largo que ChatGPT extrae de la ficha tecnica (item["descripcion"]
      en el JSON). Ahora muestra el nombre del material tal como esta en el
      nombre de la carpeta (item["nombre"]).
   2. Icono de la app: nuevo icono (Icono Generador de Submittals/) aplicado
      a la ventana/barra de tareas (assets/icono_app.ico) y como logo en el
      header de la interfaz (assets/icono_header.png).
   3. OCR: datos de idioma espanol para Tesseract ("spa.traineddata") ahora se
      guardan en una carpeta propia de la app (tessdata_es/, sin necesitar
      permisos de administrador) en vez de la carpeta de instalacion de
      Tesseract-OCR.
 v2.6.5 sobre v2.6.4:
   1. Fix: fichas tecnicas con tablas densas de especificaciones (diametros,
      tolerancias, dimensiones) fallaban en extraer_con_chatgpt() porque el
      texto plano (pypdf) llega a ChatGPT con las tablas "aplanadas" en una
      sola linea de numeros sin estructura de fila/columna, y el modelo no
      logra identificar marca/descripcion. Ahora se detecta ese caso
      (es_tabla_tecnica_densa) y, si es PDF, se relee con pdfplumber
      preservando filas/columnas (extraer_resumen_tabla_pdfplumber) antes de
      enviarlo a ChatGPT.
   2. Nueva validacion (_desc_parece_valida): si aun asi ChatGPT devuelve una
      "descripcion" que en realid/ad es ruido numerico, se reemplaza por un
      texto generico en vez de guardar basura en el submittal.
   3. Nueva dependencia opcional: pdfplumber (bootstrap la instala sola; si no
      esta disponible el codigo sigue funcionando igual que antes, sin el fix).
 v2.4 sobre v2.3:
   1. FIX CRITICO (multiples ventanas en .exe): instancia unica + control de
      hilos + NO relanzar el .exe (el bootstrap ya NO ejecuta subprocess con
      sys.executable cuando esta empaquetado) + multiprocessing.freeze_support().
   2. Seguridad: se ELIMINA el boton "Copiar" de la API key; solo se muestra su
      estado (Configurada / Sin configurar).
   3. Traduccion automatica: ChatGPT detecta el idioma de la ficha y, si no esta
      en espanol, la traduce; se registra idioma_original y fue_traducido.
   4. Fix pagina en blanco en la caratula del Ministerio (CSS de impresion).
================================================================================
"""

import os
import re
import io
import sys
import json
import time
import base64
import queue
import shutil
import socket
import difflib
import logging
import threading
import subprocess
import importlib.util
import multiprocessing
from pathlib import Path
from datetime import datetime

try:
    import auto_updater            # v2.6.7: auto-actualizacion hibrida (opcional)
except Exception:
    auto_updater = None

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ============================================================================
# CONSTANTES / TEMA
# ============================================================================
VERSION   = "2.6.7"
AUTOR     = "Adrián Castro"
ROJO_ES   = "#E11D2D"
AZUL_ES   = "#1F3864"
GRIS_BG   = "#F4F5F7"
VERDE_OK  = "#1E7E34"
AMARILLO  = "#B8860B"
BLANCO    = "#FFFFFF"

CARPETAS_MADRE = ("ARQUITECTONICOS", "MECANICOS", "ESTRUCTURALES", "ELECTRICOS")
MADRE_A_PREFIJO = {"ARQUITECTONICOS": "ARQ", "ESTRUCTURALES": "ESTR",
                   "MECANICOS": "MEC", "ELECTRICOS": "ELEC"}
ORDEN_CAT = {"ARQ": 0, "ESTR": 1, "MEC": 2, "ELEC": 3}
DISCIPLINA_SINGULAR = {"ARQUITECTONICOS": "ARQUITECTONICO", "MECANICOS": "MECANICO",
                       "ESTRUCTURALES": "ESTRUCTURAL", "ELECTRICOS": "ELECTRICO"}

NOMBRE_MOTOR   = "generate_caratulas.py"
NOMBRE_JSON    = "datos_materiales.json"
NOMBRE_REPORTE = "generate_caratulas_report.txt"
NOMBRE_CARPETA_FINAL = "COMPILADO FINAL SUBMITTAL"
DEFAULT_BASE   = r"C:\Users\castr\Downloads\Submitals ES"

IMG_EXT = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
ADMISIBLES = (".pdf",) + IMG_EXT + (".docx", ".doc")
MAX_FICHA = 2000
MAX_COMBINADO = 6000
MAX_NORMATIVA = 500
MODELO_GPT = "gpt-4o-mini"
TESSERACT_DEFECTO = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
SIN_ESP = "SIN ESPECIFICAR"
VISION_MODEL = "gpt-4o"   # v2.5.1: OpenAI Vision (una sola API)
TESSERACT_OK = None
LOCK_PORT = 50573        # instancia unica (socket)
_LOCK_SOCK = None

CARATULAS = {
    "clasica": ("template_caratula.html",
                "Tabla visual refresh/assets/logo_es_crop.png", "logo_path"),
    "ministerio_salud": ("template_ministerio_salud.html",
                         "Tabla visual refresh/assets/ministerio_salud_banner.png", "logo_ministerio"),
}
CARATULA_NOMBRES = {"clasica": "Clásica (ES Constructora)", "ministerio_salud": "Ministerio de Salud"}

CAMPOS_PROYECTO = [
    ("proyecto", "Proyecto"),
    ("cliente", "Cliente / Institución"),
    ("contrato", "Contrato / Licitación"),
    ("monto", "Monto"),
    ("plazo", "Plazo"),
    ("nombre_cargo", "Responsable (nombre y cargo)"),
    ("fecha", "Fecha (dd/mm/aaaa)"),
    ("fecha_emision", "Fecha de emisión"),
]


def app_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resource_path(rel):
    base = getattr(sys, "_MEIPASS", None)
    if base:
        p = Path(base) / rel
        if p.exists():
            return p
    return app_dir() / rel


ICONO_APP_ICO = "assets/icono_app.ico"     # icono de ventana/barra de tareas (Windows)
ICONO_APP_PNG = "assets/icono_app.png"     # respaldo multiplataforma para iconphoto()
ICONO_HEADER_PNG = "assets/icono_header.png"  # logo pequeno para el header de la interfaz


def _aplicar_icono_ventana(win):
    """Aplica el icono del Generador de Submittals a una ventana Tk.
    iconbitmap (.ico) es lo que Windows usa para la barra de tareas; iconphoto
    (.png) queda como respaldo si .ico no esta disponible o falla."""
    try:
        ico = resource_path(ICONO_APP_ICO)
        if ico.exists():
            win.iconbitmap(str(ico))
            return
    except Exception:
        pass
    try:
        png = resource_path(ICONO_APP_PNG)
        if png.exists():
            win._icono_img = tk.PhotoImage(file=str(png))
            win.iconphoto(True, win._icono_img)
    except Exception:
        pass


CONFIG_PATH = app_dir() / "submitals_config.json"

# Carpeta propia de la app para los datos de idioma de Tesseract (v2.6.6):
# el instalador de Tesseract-OCR solo trae "eng"/"osd" por defecto y su
# carpeta ("Program Files\Tesseract-OCR\tessdata") requiere permisos de
# administrador para agregar "spa" -> se guarda una copia aparte aqui, sin
# necesitar admin, y se apunta pytesseract a ella via TESSDATA_PREFIX.
TESSDATA_DIR = app_dir() / "tessdata_es"
TESSDATA_URLS = {
    "eng": "https://github.com/tesseract-ocr/tessdata/raw/main/eng.traineddata",
    "osd": "https://github.com/tesseract-ocr/tessdata/raw/main/osd.traineddata",
    "spa": "https://github.com/tesseract-ocr/tessdata/raw/main/spa.traineddata",
}

_pw_local = app_dir() / "ms-playwright"
# v2.6.2: antes bastaba con que la carpeta EXISTIERA para redirigir Playwright
# ahi, aunque estuviera vacia. Una carpeta "ms-playwright" vacia (ej. creada
# por error o como placeholder) terminaba OCULTANDO la instalacion valida de
# Chromium en la cache global de Playwright (%LOCALAPPDATA%\ms-playwright),
# forzando una caida silenciosa a pdfkit/wkhtmltopdf (peor soporte de CSS
# moderno => caratulas mal renderizadas). Ahora solo se usa la carpeta local
# si ademas tiene contenido (navegador realmente instalado ahi).
if _pw_local.exists() and any(_pw_local.iterdir()):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(_pw_local)


# ============================================================================
# INSTANCIA UNICA (evita multiples ventanas)  -- v2.4
# ============================================================================
def instancia_unica(port=LOCK_PORT):
    """True si es la unica instancia; False si ya hay otra corriendo."""
    global _LOCK_SOCK
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 0)
        s.bind(("127.0.0.1", port))
        s.listen(1)
        _LOCK_SOCK = s   # se mantiene vivo mientras el programa corre
        return True
    except OSError:
        try:
            s.close()
        except Exception:
            pass
        return False


# ============================================================================
# BOOTSTRAP DE DEPENDENCIAS  (v2.4: nunca relanza el .exe)
# ============================================================================
PIP_DEPS = [("jinja2", "jinja2"), ("openai", "openai"), ("pypdf", "pypdf"),
            ("fitz", "pymupdf"), ("pytesseract", "pytesseract"), ("PIL", "Pillow"),
            ("docx", "python-docx"), ("playwright", "playwright"),
            ("requests", "requests"), ("bs4", "beautifulsoup4"), ("openpyxl", "openpyxl"),
            ("pdfplumber", "pdfplumber")]


def _tesseract_binario():
    """Ubica el ejecutable de tesseract instalado en el sistema, si existe."""
    en_path = shutil.which("tesseract")
    if en_path:
        return Path(en_path)
    for cand in (TESSERACT_DEFECTO, str(app_dir() / "tesseract" / "tesseract.exe")):
        if os.path.exists(cand):
            return Path(cand)
    return None


def _instalar_tesseract_si_falta(logf):
    """Instala Tesseract-OCR con winget si no se encuentra en el sistema.
    A diferencia de los paquetes de PIP_DEPS, esto es un binario del sistema
    (no una libreria de Python de este interprete), por lo que se instala
    igual este empaquetado (.exe) o no: subprocess aqui llama a "winget", NO
    a sys.executable, asi que no hay riesgo de relanzar el programa."""
    if _tesseract_binario():
        return
    winget = shutil.which("winget")
    if not winget:
        logf("Tesseract-OCR no encontrado y 'winget' no esta disponible; instalelo "
             "manualmente desde https://github.com/UB-Mannheim/tesseract/releases")
        return
    logf("Tesseract-OCR no encontrado; instalando con winget (puede tardar un momento)...")
    try:
        subprocess.check_call([
            winget, "install", "--id", "UB-Mannheim.TesseractOCR", "-e", "--silent",
            "--accept-package-agreements", "--accept-source-agreements",
        ])
    except Exception as e:
        logf(f"No se pudo instalar Tesseract-OCR automaticamente: {str(e)[:200]}")
        return
    logf("Tesseract-OCR instalado." if _tesseract_binario()
         else "winget termino pero no se encontro tesseract.exe; revise manualmente.")


def _asegurar_idioma_espanol(logf):
    """Garantiza que 'spa.traineddata' (idioma espanol) este disponible para
    pytesseract, copiandolo/descargandolo a TESSDATA_DIR (propia de la app,
    sin permisos de administrador) en vez de la carpeta de instalacion de
    Tesseract. Ver _config_tesseract(), que apunta TESSDATA_PREFIX ahi."""
    binario = _tesseract_binario()
    if not binario:
        return False

    TESSDATA_DIR.mkdir(parents=True, exist_ok=True)
    origen = binario.parent / "tessdata"
    faltan = [lang for lang in ("eng", "osd", "spa")
             if not (TESSDATA_DIR / f"{lang}.traineddata").exists()]
    if faltan:
        logf("Descargando datos de idioma para OCR (" + ", ".join(faltan) + ")...")
        for lang in faltan:
            destino = TESSDATA_DIR / f"{lang}.traineddata"
            try:
                origen_lang = origen / f"{lang}.traineddata"
                if origen_lang.exists():
                    shutil.copy2(origen_lang, destino)
                    continue
            except Exception:
                pass
            try:
                import requests
                r = requests.get(TESSDATA_URLS[lang], timeout=60)
                r.raise_for_status()
                destino.write_bytes(r.content)
            except Exception as e:
                logf(f"No se pudo obtener el idioma '{lang}' para OCR: {str(e)[:150]}")
    return (TESSDATA_DIR / "spa.traineddata").exists()


def bootstrap(logf):
    if sys.version_info < (3, 9):
        raise RuntimeError(f"Se requiere Python 3.9 o superior "
                           f"(detectado {sys.version.split()[0]}).")
    congelado = getattr(sys, "frozen", False)

    # IMPORTANTE: si esta empaquetado (.exe), sys.executable es el PROPIO .exe.
    # Ejecutar subprocess([sys.executable, ...]) RELANZARIA el programa en bucle
    # (bug de "multiples ventanas"). Por eso, congelado => NO se instala nada.
    if congelado:
        logf("Ejecutable empaquetado: librerias incluidas (no se instala nada).")
    else:
        faltan = [pkg for mod, pkg in PIP_DEPS if importlib.util.find_spec(mod) is None]
        if faltan:
            logf("Instalando dependencias: " + ", ".join(faltan))
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", *faltan])
            except Exception as e:
                raise RuntimeError(
                    "No se pudieron instalar las dependencias.\n"
                    "Verifique su conexion a internet (o ejecute como administrador).\n"
                    "Detalle: " + str(e)[:200])
        else:
            logf("Dependencias de Python: OK")

    global TESSERACT_OK
    _instalar_tesseract_si_falta(logf)
    binario = _tesseract_binario()
    idioma_ok = _asegurar_idioma_espanol(logf) if binario else False
    TESSERACT_OK = bool(binario) and idioma_ok
    if TESSERACT_OK:
        logf("Tesseract-OCR: OK (con idioma español)")
    elif binario:
        logf("Tesseract-OCR: instalado pero sin idioma español (el OCR quedara limitado)")
    else:
        logf("Tesseract-OCR: NO instalado (el OCR de imagenes quedara limitado)")


# ============================================================================
# CONFIGURACION PERSISTENTE
# ============================================================================
CONFIG_DEFECTO = {
    "version": VERSION,
    "caratula_seleccionada": "clasica",
    "carpetas_recientes": [],
    "ultimo_json": "",
    "opciones": {
        "solo_faltantes": True,
        "forzar_regeneracion": False,
        "mostrar_log": True,
        "generar_json_automatico": True,
        "usar_json_existente": False,
    },
    "datos_proyecto": {k: "" for k, _ in CAMPOS_PROYECTO},
    "api": {"openai_key_encrypted": "", "anthropic_key_encrypted": "", "ultima_validacion": ""},
    "mantenimiento": {"ultima_limpieza_cache": "", "veces_reseted": 0},
}


def _config_base():
    return json.loads(json.dumps(CONFIG_DEFECTO))


def cargar_config():
    try:
        cfg = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        base = _config_base()
        base["version"] = VERSION
        base["caratula_seleccionada"] = cfg.get("caratula_seleccionada", "clasica")
        base["carpetas_recientes"] = cfg.get("carpetas_recientes", [])
        base["ultimo_json"] = cfg.get("ultimo_json", "")
        base["opciones"].update(cfg.get("opciones", {}))
        base["datos_proyecto"].update(cfg.get("datos_proyecto", {}))
        base["api"].update(cfg.get("api", {}))
        base["mantenimiento"].update(cfg.get("mantenimiento", {}))
        if cfg.get("ultima_carpeta") and not base["carpetas_recientes"]:
            base["carpetas_recientes"] = [cfg["ultima_carpeta"]]
        return base
    except Exception:
        return _config_base()


def guardar_config(cfg):
    try:
        cfg["version"] = VERSION
        CONFIG_PATH.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass


def cifrar_api_key(key):
    try:
        return base64.b64encode(key.encode("utf-8")).decode("ascii")
    except Exception:
        return ""


def descifrar_api_key(enc):
    try:
        return base64.b64decode(enc.encode("ascii")).decode("utf-8")
    except Exception:
        return ""


def cargar_api_key(cfg):
    return descifrar_api_key(cfg.get("api", {}).get("openai_key_encrypted", ""))


def cargar_claude_key(cfg):
    return descifrar_api_key(cfg.get("api", {}).get("anthropic_key_encrypted", ""))


def ahora_iso():
    return datetime.now().isoformat(timespec="seconds")


def sanitizar(nombre):
    s = str(nombre).replace('"', 'in').replace('/', '-').replace('\\', '-')
    s = re.sub(r'[:*?<>|]', '-', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# ============================================================================
# EXCEPCIONES
# ============================================================================
class _AuthError(Exception):
    pass


class _Cancelado(Exception):
    pass


# ============================================================================
# ERRORES LEGIBLES
# ============================================================================
def error_legible(e):
    s = str(e)
    if isinstance(e, FileNotFoundError) or "No such file" in s:
        return "Archivo o carpeta no encontrada. Verifique las rutas seleccionadas."
    if isinstance(e, PermissionError) or "Permission denied" in s or "Acceso denegado" in s:
        return ("Sin permisos / archivo en uso. Cierre el PDF si esta abierto en un "
                "visor y verifique que la carpeta no sea de solo lectura.")
    if isinstance(e, json.JSONDecodeError):
        return (f"El archivo JSON tiene un error de formato en la linea "
                f"{e.lineno}, columna {e.colno}.")
    if "playwright" in s.lower() and ("install" in s.lower() or "Executable doesn't exist" in s):
        return ("Falta el navegador de Playwright. Ejecute:\n"
                "python -m playwright install chromium")
    if isinstance(e, (ImportError, ModuleNotFoundError)):
        return ("Falta una libreria de Python. Ejecute:\n"
                "pip install jinja2 pypdf pymupdf playwright openai pytesseract Pillow "
                "python-docx requests beautifulsoup4")
    return f"Ocurrio un error inesperado:\n{s[:300]}"


# ============================================================================
# TOOLTIP
# ============================================================================
class Tooltip:
    def __init__(self, widget, texto):
        self.widget, self.texto, self.tip = widget, texto, None
        widget.bind("<Enter>", self._mostrar)
        widget.bind("<Leave>", self._ocultar)

    def _mostrar(self, _=None):
        if self.tip or not self.texto:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self.tip = tk.Toplevel(self.widget)
        self.tip.wm_overrideredirect(True)
        self.tip.wm_geometry(f"+{x}+{y}")
        tk.Label(self.tip, text=self.texto, justify="left", bg="#FFFFE0",
                 relief="solid", borderwidth=1, font=("Segoe UI", 9),
                 wraplength=380, padx=6, pady=4).pack()

    def _ocultar(self, _=None):
        if self.tip:
            self.tip.destroy()
            self.tip = None


# ============================================================================
# LECTURA ROBUSTA DE DOCUMENTOS
# ============================================================================
def buscar_archivos_carpeta(carpeta):
    try:
        files = []
        for p in carpeta.iterdir():
            if not p.is_file():
                continue
            up = p.name.upper()
            if up.startswith("CARATULA") or up.endswith("-CMP.PDF"):
                continue
            if p.suffix.lower() in ADMISIBLES:
                files.append(p)
        return sorted(files, key=lambda x: x.name.lower())
    except Exception:
        return []


def _leer_pdf_texto(path):
    from pypdf import PdfReader
    r = PdfReader(str(path))
    txt = ""
    for pg in r.pages[:2]:
        txt += (pg.extract_text() or "")
    return " ".join(txt.split())


def _config_tesseract():
    import pytesseract
    binario = _tesseract_binario()
    if binario:
        pytesseract.pytesseract.tesseract_cmd = str(binario)
    if (TESSDATA_DIR / "spa.traineddata").exists():
        os.environ["TESSDATA_PREFIX"] = str(TESSDATA_DIR)
    return pytesseract


def ocr_mejorado(img):
    from PIL import Image, ImageEnhance, ImageOps
    pytesseract = _config_tesseract()
    abrir_local = not isinstance(img, Image.Image)
    if abrir_local:
        img = Image.open(str(img))
    try:
        g = img.convert("L")
        g = ImageOps.autocontrast(g)
        g = ImageEnhance.Contrast(g).enhance(1.8)
        g = ImageEnhance.Sharpness(g).enhance(2.0)
        return " ".join(pytesseract.image_to_string(g, lang="spa+eng").split())
    finally:
        if abrir_local:
            try:
                img.close()
            except Exception:
                pass


def _ocr_pdf(path):
    import fitz
    from PIL import Image
    doc = fitz.open(str(path))
    out = ""
    try:
        for pg in doc[:2]:
            pix = pg.get_pixmap(matrix=fitz.Matrix(2, 2))
            bio = io.BytesIO(pix.tobytes("png"))
            img = Image.open(bio)
            try:
                out += " " + ocr_mejorado(img)
            finally:
                img.close()
                bio.close()
        return out.strip()
    finally:
        doc.close()


def extraer_texto_robusto(path):
    ext = path.suffix.lower()
    if ext == ".pdf":
        t = _leer_pdf_texto(path)
        if len(t.strip()) < 20:
            t = _ocr_pdf(path)
        return t[:MAX_FICHA]
    if ext in IMG_EXT:
        return ocr_mejorado(path)[:MAX_FICHA]
    if ext == ".docx":
        import docx
        d = docx.Document(str(path))
        return " ".join(p.text for p in d.paragraphs if p.text.strip())[:MAX_FICHA]
    if ext == ".doc":
        raise ValueError("formato .doc no soportado (convierta a .docx o PDF)")
    raise ValueError(f"formato no soportado: {ext}")


# ============================================================================
# DETECCION Y RE-EXTRACCION DE TABLAS TECNICAS DENSAS  (v2.6.5)
# ============================================================================
# Las fichas de accesorios/tuberia/perfileria con tablas de dimensiones (varios
# diametros x varias columnas, con tolerancias) se leen sin error con
# pypdf/OCR, pero _leer_pdf_texto() colapsa todos los saltos de linea a
# espacios (" ".join(txt.split())) -> la tabla queda "aplanada" en una sola
# tira de numeros sin fila/columna (ej. "13 19 25 32 38 ... 21.6 26.9 33.7
# ... 0.2 0.3 0.3 ..."). ChatGPT no puede reconstruir esa estructura y la
# marca/descripcion salen vacias, genericas o sin sentido. Estas funciones
# detectan ese patron y, si el documento es un PDF, lo releen con pdfplumber
# (que si respeta filas/columnas) para armar un resumen que ChatGPT SI puede
# interpretar. Si pdfplumber no esta instalado o no encuentra tabla, no pasa
# nada: se sigue usando el texto plano de siempre (comportamiento identico al
# de antes de este cambio).
RE_TOKEN_NUMERICO = re.compile(r'(?<!\w)\d{1,4}(?:[.,]\d+)?(?!\w)')
PALABRAS_TABLA_TECNICA = ("tolerancia", "diámetro nominal", "diametro nominal",
                          "dimensiones", "ovalidad", "diámetro exterior",
                          "diametro exterior")


def es_tabla_tecnica_densa(texto):
    """True si 'texto' parece el volcado plano (sin estructura) de una tabla
    de especificaciones/dimensiones -> candidato a releerse con pdfplumber en
    vez de enviarse tal cual a ChatGPT."""
    if not texto or len(texto) < 60:
        return False
    palabras = re.findall(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]{3,}", texto)
    numeros = RE_TOKEN_NUMERICO.findall(texto)
    if len(numeros) < 15:
        return False
    ratio_numeros = len(numeros) / max(1, len(palabras) + len(numeros))
    if ratio_numeros < 0.22:
        return False
    texto_l = texto.lower()
    menciona_tabla = any(p in texto_l for p in PALABRAS_TABLA_TECNICA)
    muchas_tolerancias = texto.count("±") >= 3
    return menciona_tabla or muchas_tolerancias


def extraer_resumen_tabla_pdfplumber(path, max_filas=8, max_paginas=3):
    """Relee un PDF con pdfplumber (respeta filas/columnas de las tablas) y
    arma un resumen de texto legible para ChatGPT. Devuelve "" si no se pudo
    extraer ninguna tabla o si pdfplumber no esta disponible; el llamador debe
    seguir funcionando con el texto plano original en ese caso."""
    try:
        import pdfplumber
    except ImportError:
        return ""
    try:
        resumen = []
        with pdfplumber.open(str(path)) as pdf:
            for pagina in pdf.pages[:max_paginas]:
                try:
                    tabla = pagina.extract_table()
                except Exception:
                    tabla = None
                texto_pagina = (pagina.extract_text() or "").strip()
                if texto_pagina:
                    resumen.append(texto_pagina[:400])
                if not tabla or len(tabla) < 2:
                    continue
                encabezado = [str(c or "").strip() for c in tabla[0]]
                filas = [f for f in tabla[1:] if any(str(c or "").strip() for c in f)]
                if not filas:
                    continue
                if any(encabezado):
                    resumen.append("Columnas de la tabla: " +
                                   " | ".join(c for c in encabezado if c))
                resumen.append(f"Tabla con {len(filas)} fila(s)/variante(s), "
                               f"muestra de las primeras {min(max_filas, len(filas))}:")
                for f in filas[:max_filas]:
                    valores = [str(c or "").strip() for c in f]
                    resumen.append(" | ".join(v for v in valores if v))
        return "\n".join(resumen).strip()
    except Exception:
        return ""


# ============================================================================
# ChatGPT  (v2.4: deteccion de idioma + traduccion)
# ============================================================================
PROMPT_TXT = (
    "Eres un experto en construccion, materiales de obra, normativas tecnicas y "
    "traduccion tecnica.\n"
    "Te doy 1 O MAS documentos de UN MISMO material, que pueden estar en otro idioma.\n"
    "1. Detecta el idioma principal.\n"
    "2. Si NO esta en espanol, TRADUCE al espanol con precision tecnica y trabaja "
    "sobre la version traducida.\n"
    "Luego extrae de forma INTEGRADA (todo en espanol):\n"
    " - marca/fabricante (la principal; si no aparece: \"SIN ESPECIFICAR\")\n"
    " - descripcion tecnica breve (maximo 200 caracteres)\n"
    " - normativas/estandares (todas, sin duplicados, orden alfabetico; max 500; "
    "si no hay: \"SIN ESPECIFICAR\")\n"
    "Si un documento viene marcado como \"[TABLA DE ESPECIFICACIONES TECNICAS...]\" "
    "es una tabla de dimensiones con varias filas/variantes (distintos diametros, "
    "medidas, etc.): la descripcion debe referirse al PRODUCTO EN GENERAL (no a "
    "una fila especifica) e ignorar los valores de tolerancia numerica.\n\n"
    "DOCUMENTOS:\n<<CONTENIDO>>\n\n"
    "Responde UNICAMENTE en JSON valido (sin markdown):\n"
    '{"idioma_original": "espanol|ingles|frances|portugues|...", '
    '"fue_traducido": true/false, "marca": "string", '
    '"descripcion": "string (max 200)", "normativa": "string (max 500 o SIN ESPECIFICAR)"}'
)

# v2.6: clasificacion de la relacion entre 2+ fichas tecnicas de una misma carpeta
PROMPT_RELACION_FICHAS = (
    "Eres un experto en construccion, materiales de obra y control de calidad de "
    "submittals.\n"
    "Te doy el contenido de varias fichas tecnicas que se encontraron JUNTAS en la "
    "misma carpeta de un submittal. Si se incluye tambien una IMAGEN de la "
    "portada de alguna ficha, usala para identificar el logo o nombre de la "
    "empresa fabricante: muchas fichas tienen el nombre de la marca UNICAMENTE "
    "como logo grafico en la portada, y NO aparece como texto en el resto del "
    "documento; en ese caso la imagen es la unica forma de identificar la marca.\n"
    "Para cada ficha identifica su empresa fabricante/proveedor (buscalo tanto en "
    "el texto como en el logo de la imagen de portada, si la hay) y el tipo de "
    "producto que describe (categoria general, ej: 'tuberia de PVC', 'clavo de "
    "acero', 'pintura acrilica').\n"
    "Luego clasifica la relacion entre TODAS las fichas en una sola categoria:\n"
    " - MISMO_PROVEEDOR: TODAS mencionan o muestran (en texto o en el logo de "
    "portada) el MISMO nombre de empresa/marca, o son documentos complementarios "
    "(ficha + certificado + ensayo, etc.) de un mismo producto de una misma "
    "empresa.\n"
    " - MISMA_FAMILIA_DISTINTA_MARCA: se identifican 2 o mas nombres de empresa/"
    "marca DIFERENTES entre las fichas (en texto o en el logo), aunque describan "
    "el mismo tipo de producto (alternativas comercialmente intercambiables). "
    "IMPORTANTE: si identificas 2 o mas marcas distintas, SIEMPRE clasifica aqui, "
    "NUNCA en MISMO_PROVEEDOR, sin importar cuan parecido sea el resto del "
    "contenido. Si NO logras identificar ninguna marca en una o mas fichas (ni "
    "en texto ni en imagen), NO asumas que es la misma empresa que las demas: "
    "usa esta categoria tambien en ese caso de duda, indicandolo en la "
    "justificacion.\n"
    " - DISCREPANCIA: describen tipos de producto distintos e incompatibles entre "
    "si (ej. clavos y tuberia de PVC), lo que sugiere que la carpeta mezcla por "
    "error fichas de materiales diferentes.\n\n"
    "FICHAS:\n<<CONTENIDO>>\n\n"
    "Responde UNICAMENTE en JSON valido (sin markdown):\n"
    '{"relacion": "MISMO_PROVEEDOR|MISMA_FAMILIA_DISTINTA_MARCA|DISCREPANCIA", '
    '"marcas": ["string", ...], "tipos_producto": ["string", ...], '
    '"justificacion": "string breve (max 200 caracteres), en espanol"}'
)


def _parse_json_gpt(texto):
    a = texto.find("{")
    b = texto.rfind("}")
    if a >= 0 and b > a:
        return json.loads(texto[a:b + 1])
    raise ValueError("la respuesta no contiene JSON")


def test_openai(api_key):
    if not api_key.strip():
        return False, "Ingrese una API Key."
    try:
        import openai
        from openai import OpenAI
    except Exception:
        return False, "Falta la libreria 'openai'. Ejecute: pip install openai"
    try:
        OpenAI(api_key=api_key, timeout=15).models.list()
        return True, "Conexion exitosa."
    except Exception as e:
        try:
            import openai
            if isinstance(e, openai.AuthenticationError):
                return False, "API key invalida o sin creditos."
        except Exception:
            pass
        return False, f"No se pudo conectar: {str(e)[:150]}"


def _trunc_desc(texto, n=200):
    texto = (texto or "").strip()
    if len(texto) <= n:
        return texto
    cut = texto[:n]
    m = max(cut.rfind("."), cut.rfind(" "))
    if m > 0:
        cut = cut[:m]
    return cut.rstrip() + "..."


def _trunc_norm(texto):
    texto = (texto or "").strip()
    if not texto:
        return SIN_ESP
    if len(texto) <= MAX_NORMATIVA:
        return texto
    cut = texto[:MAX_NORMATIVA]
    c = cut.rfind(",")
    return (cut[:c] if c > 100 else cut).strip()


def _desc_parece_valida(desc):
    """Heuristica (v2.6.5) para descartar una 'descripcion' que ChatGPT
    devolvio pero que en realidad es ruido numerico (una tabla de
    especificaciones mal interpretada) en vez de una frase real en espanol."""
    desc = (desc or "").strip()
    if len(desc) < 8:
        return False
    letras = sum(c.isalpha() for c in desc)
    return (letras / max(1, len(desc))) >= 0.4


def _chatgpt_json(prompt, api_key, q, cons, modelo, max_tokens=320, temperature=0.3):
    """Llama a OpenAI y devuelve el JSON parseado de la respuesta, con reintentos
    ante fallos transitorios. Usada tanto para la extraccion de marca/descripcion
    como para la clasificacion de relacion entre fichas."""
    import openai
    from openai import OpenAI
    client = OpenAI(api_key=api_key, timeout=25)
    intentos, ultimo = 0, None
    while intentos < 3:
        intentos += 1
        try:
            r = client.chat.completions.create(
                model=modelo, temperature=temperature, max_tokens=max_tokens,
                messages=[{"role": "user", "content": prompt}])
            return _parse_json_gpt(r.choices[0].message.content or "")
        except openai.AuthenticationError as e:
            raise _AuthError(str(e))
        except openai.RateLimitError as e:
            ultimo = e
            if q:
                q.put(("WARN", f"{cons}: limite de tasa, esperando 30s..."))
            time.sleep(30)
        except (openai.APITimeoutError, openai.APIConnectionError) as e:
            ultimo = e
            time.sleep(2)
        except Exception as e:
            ultimo = e
            time.sleep(1)
    raise RuntimeError(f"sin respuesta valida tras {intentos} intentos ({ultimo})")


def extraer_con_chatgpt(contenido, api_key, q=None, cons="-"):
    """Devuelve (marca, desc, norm, idioma_original, fue_traducido)."""
    prompt = PROMPT_TXT.replace("<<CONTENIDO>>", contenido)
    data = _chatgpt_json(prompt, api_key, q, cons, MODELO_GPT, max_tokens=320)
    marca = str(data.get("marca", "")).strip()
    desc = str(data.get("descripcion", "")).strip()
    norm = str(data.get("normativa", SIN_ESP)).strip()
    idioma = str(data.get("idioma_original", "español")).strip() or "español"
    trad = bool(data.get("fue_traducido", False))
    return marca, desc, norm, idioma, trad


RELACIONES_FICHAS_VALIDAS = ("MISMO_PROVEEDOR", "MISMA_FAMILIA_DISTINTA_MARCA", "DISCREPANCIA")


def _render_portada_b64(path, max_px=900):
    """Renderiza la primera pagina de un PDF (o abre la imagen, si ya es una)
    como JPEG en base64 para enviarla a OpenAI Vision. Muchas fichas tienen el
    nombre del fabricante UNICAMENTE como logo grafico en la portada (no como
    texto seleccionable), por lo que la extraccion de texto normal nunca lo ve;
    la imagen es la unica forma de identificar la marca en esos casos.
    Devuelve None si no se pudo generar (archivo no soportado, PDF corrupto,
    faltan PyMuPDF/Pillow, etc.) -- el llamador debe seguir funcionando sin ella.
    """
    try:
        path = Path(path)
        ext = path.suffix.lower()
        from PIL import Image
        if ext == ".pdf":
            import fitz
            doc = fitz.open(str(path))
            try:
                if len(doc) == 0:
                    return None
                pix = doc[0].get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
                img = Image.open(io.BytesIO(pix.tobytes("png")))
            finally:
                doc.close()
        elif ext in IMG_EXT:
            img = Image.open(str(path))
        else:
            return None
        img = img.convert("RGB")
        if max(img.size) > max_px:
            ratio = max_px / max(img.size)
            img = img.resize((max(1, int(img.width * ratio)), max(1, int(img.height * ratio))))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=70)
        return base64.b64encode(buf.getvalue()).decode("ascii")
    except Exception:
        return None


def _marcas_claramente_distintas(marcas):
    """True si la lista de marcas identificadas contiene 2 o mas nombres
    claramente distintos (no variantes del mismo nombre/mismas siglas)."""
    normalizadas = []
    for m in marcas:
        norm = re.sub(r'[^A-Z0-9]', '', m.upper())
        if not norm:
            continue
        if not any(difflib.SequenceMatcher(None, norm, existente).ratio() >= 0.7
                  for existente in normalizadas):
            normalizadas.append(norm)
    return len(normalizadas) >= 2


def analizar_relacion_fichas(textos_por_doc, api_key, q=None, cons="-", imagenes_portada=None):
    """Clasifica la relacion entre 2+ fichas tecnicas halladas en una misma carpeta.

    textos_por_doc: lista de tuplas (nombre_archivo, texto_extraido).
    imagenes_portada: lista opcional, alineada con textos_por_doc, de imagenes
        de portada en base64 (o None por documento) para que la IA (Vision)
        pueda identificar marcas que solo aparecen como logo grafico.
    Devuelve un dict {"relacion", "marcas", "tipos_producto", "justificacion"} o
    None si la clasificacion no fue concluyente/posible (no bloquea el proceso;
    simplemente no se agregan aspectos adicionales para esa carpeta).
    """
    contenido = "\n\n".join(
        f"FICHA {i} ({nombre}):\n{texto[:800]}"
        for i, (nombre, texto) in enumerate(textos_por_doc, 1))
    prompt_texto = PROMPT_RELACION_FICHAS.replace("<<CONTENIDO>>", contenido)

    imagenes_portada = imagenes_portada or []
    if any(imagenes_portada):
        modelo = VISION_MODEL
        prompt = [{"type": "text", "text": prompt_texto}]
        for i, ((nombre, _), b64) in enumerate(zip(textos_por_doc, imagenes_portada), 1):
            if b64:
                prompt.append({"type": "text",
                              "text": f"Imagen de portada de FICHA {i} ({nombre}):"})
                prompt.append({"type": "image_url",
                              "image_url": {"url": f"data:image/jpeg;base64,{b64}",
                                           "detail": "low"}})
    else:
        modelo = MODELO_GPT
        prompt = prompt_texto

    try:
        data = _chatgpt_json(prompt, api_key, q, cons, modelo, max_tokens=300)
    except _AuthError:
        raise
    except Exception as e:
        if q:
            q.put(("WARN", f"{cons}: no se pudo analizar la relacion entre fichas ({e})"))
        return None
    relacion = str(data.get("relacion", "")).strip().upper()
    if relacion not in RELACIONES_FICHAS_VALIDAS:
        return None
    marcas = [str(m).strip() for m in data.get("marcas", []) if str(m).strip()]
    tipos = [str(t).strip() for t in data.get("tipos_producto", []) if str(t).strip()]
    justificacion = str(data.get("justificacion", "")).strip()

    # v2.6.4: salvaguarda determinista -- si la IA misma identifico 2+ marcas
    # claramente distintas pero igual clasifico como MISMO_PROVEEDOR (inconsistencia),
    # se corrige aqui sin depender de un nuevo intento a la IA.
    if relacion == "MISMO_PROVEEDOR" and _marcas_claramente_distintas(marcas):
        relacion = "MISMA_FAMILIA_DISTINTA_MARCA"
        if not justificacion:
            justificacion = "Se detectaron nombres de fabricante distintos entre las fichas."

    return {"relacion": relacion, "marcas": marcas, "tipos_producto": tipos,
            "justificacion": justificacion}


def construir_texto_aspectos_adicionales(relacion_info, n_docs):
    """Redacta el texto profesional de 'Aspectos adicionales' segun la relacion
    detectada entre las fichas tecnicas de la carpeta. Para DISCREPANCIA no se
    redacta texto: esa situacion solo genera una advertencia en el log."""
    relacion = relacion_info.get("relacion")
    if relacion == "MISMO_PROVEEDOR":
        return (
            f"Se adjuntan {n_docs} fichas técnicas del mismo proveedor, las cuales "
            "se complementan entre sí para sustentar de forma integral la "
            "especificación técnica del material presentado."
        )
    if relacion == "MISMA_FAMILIA_DISTINTA_MARCA":
        return (
            f"Se adjuntan {n_docs} fichas técnicas de distintas marcas comerciales "
            "que corresponden al mismo tipo de producto. Debido a posibles "
            f"limitaciones de existencias en el mercado, se solicita la aprobación "
            f"de las {n_docs} marcas incluidas en este submittal, de manera que, "
            "ante una eventual falta de stock de alguna de ellas al momento de la "
            "instalación, se cuente con la aprobación previa para emplear "
            "cualquiera de las marcas alternativas aquí presentadas, garantizando "
            "en todos los casos el cumplimiento de los estándares de calidad y "
            "seguridad requeridos para la obra."
        )
    return ""


# ============================================================================
# COMPILADO
# ============================================================================
def imagen_a_pdf_reader(path):
    from PIL import Image
    from pypdf import PdfReader
    im = Image.open(str(path))
    try:
        if im.mode not in ("RGB", "L"):
            im = im.convert("RGB")
        buf = io.BytesIO()
        im.save(buf, format="PDF", resolution=150.0)
        buf.seek(0)
        return PdfReader(buf)
    finally:
        try:
            im.close()
        except Exception:
            pass


def generar_compilado(caratula_path, doc_paths, out_path, q=None, cons="-"):
    from pypdf import PdfWriter, PdfReader
    w = PdfWriter()
    # Solo la 1a pagina de la caratula (evita la 2a hoja en blanco que agrega Chromium)
    w.add_page(PdfReader(str(caratula_path)).pages[0])
    anexados = 0
    for d in sorted(doc_paths, key=lambda x: x.name.lower()):
        ext = d.suffix.lower()
        try:
            if ext == ".pdf":
                w.append(PdfReader(str(d)))
                anexados += 1
            elif ext in IMG_EXT:
                w.append(imagen_a_pdf_reader(d))
                anexados += 1
            else:
                if q:
                    q.put(("WARN", f"{cons}: '{d.name}' no se anexa (formato {ext})"))
        except Exception as e:
            if q:
                q.put(("WARN", f"{cons}: no se pudo anexar '{d.name}' ({e})"))
    tmp = out_path.with_suffix(".cmp.tmp")
    try:
        with open(tmp, "wb") as f:
            w.write(f)
        os.replace(tmp, out_path)
    finally:
        try:
            w.close()
        except Exception:
            pass
    return anexados


# ============================================================================
# BORRAR TODAS LAS CARATULAS Y COMPILADOS (v2.6.2) — "empezar de 0"
# ============================================================================
def borrar_caratulas_y_compilados(base, q=None):
    """
    Elimina TODAS las caratulas (CARATULA*.pdf), los compilados individuales
    (*-CMP.pdf), los compilados por disciplina (CMP SUBMITTAL *.pdf) y el
    datos_materiales.json bajo la carpeta base, para regenerar todo desde cero
    en la proxima corrida (incluyendo una nueva lectura de fichas con ChatGPT,
    ya que los datos extraidos tambien se eliminan).

    NO toca las fichas tecnicas originales de cada carpeta (los documentos
    fuente en cada subcarpeta se conservan intactos; solo se borran los PDF
    generados y el JSON de datos). Esto tambien resuelve el caso en que se
    sobreescribe una caratula pero el compilado de esa carpeta no se vuelve a
    generar porque "solo faltantes" lo encuentra ya existente: al borrar todo,
    la proxima corrida regenera todo junto y consistente.

    Devuelve un dict:
        {"caratulas": int, "compilados_individuales": int,
         "compilados_disciplina": int, "json_eliminado": bool,
         "errores": [(ruta, detalle), ...]}
    """
    base = Path(base)
    conteo = {"caratulas": 0, "compilados_individuales": 0, "compilados_disciplina": 0,
             "json_eliminado": False}
    errores = []

    json_path = base / NOMBRE_JSON
    if json_path.exists():
        try:
            json_path.unlink()
            conteo["json_eliminado"] = True
            if q:
                q.put(("LOG", f"Eliminado: {json_path.relative_to(base)}"))
        except Exception as e:
            errores.append((str(json_path), str(e)))
            if q:
                q.put(("WARN", f"No se pudo eliminar {json_path.name}: {e}"))

    for madre in CARPETAS_MADRE:
        carpeta_madre = base / madre
        if not carpeta_madre.is_dir():
            continue

        for p in carpeta_madre.glob("CMP SUBMITTAL *.pdf"):
            try:
                p.unlink()
                conteo["compilados_disciplina"] += 1
                if q:
                    q.put(("LOG", f"Eliminado: {p.relative_to(base)}"))
            except Exception as e:
                errores.append((str(p), str(e)))
                if q:
                    q.put(("WARN", f"No se pudo eliminar {p.name}: {e}"))

        for sub in carpeta_madre.iterdir():
            if not sub.is_dir():
                continue
            for p in sub.glob("*.pdf"):
                up = p.name.upper()
                if up.startswith("CARATULA"):
                    clave = "caratulas"
                elif up.endswith("-CMP.PDF"):
                    clave = "compilados_individuales"
                else:
                    continue
                try:
                    p.unlink()
                    conteo[clave] += 1
                    if q:
                        q.put(("LOG", f"Eliminado: {p.relative_to(base)}"))
                except Exception as e:
                    errores.append((str(p), str(e)))
                    if q:
                        q.put(("WARN", f"No se pudo eliminar {p.name}: {e}"))

    conteo["errores"] = errores
    return conteo


# ============================================================================
# EXPORTAR VERSION FINAL (v2.6.7)
# ============================================================================
def exportar_version_final(base, q=None):
    """
    Crea (recreandola desde cero) la carpeta 'COMPILADO FINAL SUBMITTAL' junto
    a las carpetas madre, con la version "limpia" del submittal lista para
    enviar a aprobacion: por cada carpeta madre, solo los CMP individuales de
    cada material (caratula + ficha tecnica ya combinadas) y su compilado de
    disciplina (CMP SUBMITTAL <DISC>.pdf); en la raiz, el excel guia.

        COMPILADO FINAL SUBMITTAL/
            <CARPETA MADRE>/
                <material>-CMP.pdf
                CMP SUBMITTAL <DISCIPLINA>.pdf
            Guía Materiales.xlsx

    Solo copia archivos ya generados; no modifica ni borra nada dentro de las
    carpetas madre originales.

    Devuelve un dict:
        {"carpeta_salida": str, "cmp_copiados": int,
         "compilados_disc_copiados": int, "excel_copiado": bool,
         "disciplinas_sin_archivos": [str, ...], "errores": [(ruta, detalle), ...]}
    """
    base = Path(base)
    salida = base / NOMBRE_CARPETA_FINAL

    if salida.exists():
        shutil.rmtree(salida)
    salida.mkdir(parents=True, exist_ok=True)

    resultado = {"carpeta_salida": str(salida), "cmp_copiados": 0,
                "compilados_disc_copiados": 0, "excel_copiado": False,
                "disciplinas_sin_archivos": [], "errores": []}

    for madre in CARPETAS_MADRE:
        carpeta_madre = base / madre
        if not carpeta_madre.is_dir():
            continue

        cmp_disciplina = list(carpeta_madre.glob("CMP SUBMITTAL *.pdf"))
        cmp_individuales = [p for sub in carpeta_madre.iterdir() if sub.is_dir()
                            for p in sub.glob("*-CMP.pdf")]

        if not cmp_individuales and not cmp_disciplina:
            resultado["disciplinas_sin_archivos"].append(madre)
            continue

        carpeta_destino = salida / madre
        carpeta_destino.mkdir(parents=True, exist_ok=True)

        for p in sorted(cmp_individuales, key=lambda x: x.name.lower()) + cmp_disciplina:
            clave = "compilados_disc_copiados" if p in cmp_disciplina else "cmp_copiados"
            try:
                shutil.copy2(p, carpeta_destino / p.name)
                resultado[clave] += 1
                if q:
                    q.put(("LOG", f"Copiado: {madre}/{p.name}"))
            except Exception as e:
                resultado["errores"].append((str(p), str(e)))
                if q:
                    q.put(("WARN", f"No se pudo copiar '{p.name}': {e}"))

    excel_path = base / "Guía Materiales.xlsx"
    if excel_path.exists():
        try:
            shutil.copy2(excel_path, salida / excel_path.name)
            resultado["excel_copiado"] = True
            if q:
                q.put(("LOG", f"Copiado: {excel_path.name}"))
        except Exception as e:
            resultado["errores"].append((str(excel_path), str(e)))
            if q:
                q.put(("WARN", f"No se pudo copiar el excel: {e}"))

    return resultado


# ============================================================================
# COMPILADO POR DISCIPLINA (v2.5.2)
# ============================================================================
def compilar_por_disciplina(carpeta_base, disciplina, q=None):
    """
    Genera un unico PDF con las caratulas y fichas de TODOS los materiales de
    una disciplina (ej: ARQ01 + ARQ02 + ARQ03 -> CMP SUBMITTAL ARQUITECTONICO.pdf).
    Solo reutiliza PDFs ya generados (pypdf); no usa ChatGPT ni consume API.

    Devuelve un dict:
        {"exitoso": bool, "archivo_generado": str, "ruta_completa": str,
         "total_materiales": int, "total_paginas": int, "error": str}
    """
    from pypdf import PdfWriter, PdfReader

    if disciplina not in CARPETAS_MADRE:
        return {"exitoso": False, "error": f"Disciplina invalida: {disciplina}"}

    carpeta_disciplina = Path(carpeta_base) / disciplina
    if not carpeta_disciplina.is_dir():
        return {"exitoso": False, "error": f"Carpeta no encontrada: {carpeta_disciplina}"}

    pfx = MADRE_A_PREFIJO[disciplina]
    entradas = []
    for sub in sorted(carpeta_disciplina.iterdir()):
        if not sub.is_dir():
            continue
        m = re.match(rf"^({pfx})(\d+)-(.*)$", sub.name)
        if m:
            entradas.append((int(m.group(2)), sub))
    entradas.sort(key=lambda e: e[0])

    if not entradas:
        return {"exitoso": False, "error": f"No hay materiales ({pfx}##-...) en {disciplina}"}

    w = PdfWriter()
    total_paginas = 0
    materiales_procesados = 0
    try:
        for _, sub in entradas:
            caratulas = sorted(sub.glob("CARATULA*.pdf"))
            if not caratulas:
                if q:
                    q.put(("WARN", f"{disciplina}: '{sub.name}' sin caratula, se omite del compilado"))
                continue
            try:
                r = PdfReader(str(caratulas[0]))
                w.append(r)
                total_paginas += len(r.pages)
            except Exception as e:
                if q:
                    q.put(("WARN", f"{disciplina}: no se pudo leer caratula de '{sub.name}' ({e})"))
                continue

            for d in buscar_archivos_carpeta(sub):
                ext = d.suffix.lower()
                try:
                    if ext == ".pdf":
                        r2 = PdfReader(str(d))
                    elif ext in IMG_EXT:
                        r2 = imagen_a_pdf_reader(d)
                    else:
                        if q:
                            q.put(("WARN", f"{disciplina}: '{sub.name}/{d.name}' no se anexa (formato {ext})"))
                        continue
                    w.append(r2)
                    total_paginas += len(r2.pages)
                except Exception as e:
                    if q:
                        q.put(("WARN", f"{disciplina}: no se pudo anexar '{sub.name}/{d.name}' ({e})"))

            materiales_procesados += 1

        if materiales_procesados == 0:
            return {"exitoso": False, "error": f"Ningun material de {disciplina} tiene caratula generada."}

        disciplina_sing = DISCIPLINA_SINGULAR.get(disciplina, disciplina)
        nombre_compilado = f"CMP SUBMITTAL {disciplina_sing}.pdf"
        ruta_compilado = carpeta_disciplina / nombre_compilado
        tmp = ruta_compilado.with_suffix(".tmp")
        with open(tmp, "wb") as f:
            w.write(f)
        os.replace(tmp, ruta_compilado)
    finally:
        try:
            w.close()
        except Exception:
            pass

    return {
        "exitoso": True,
        "archivo_generado": nombre_compilado,
        "ruta_completa": str(ruta_compilado),
        "total_materiales": materiales_procesados,
        "total_paginas": total_paginas,
    }


def hilo_compilados_disciplina(carpeta_base, disciplinas, q):
    t0 = time.time()
    resultados = []
    total = len(disciplinas)
    q.put(("DISC_FASE", "Generando compilados por disciplina..."))
    for i, disciplina in enumerate(disciplinas, start=1):
        q.put(("DISC_PROG", i, total, disciplina))
        q.put(("LOG", f"Procesando disciplina: {disciplina}"))
        try:
            resultado = compilar_por_disciplina(carpeta_base, disciplina, q)
        except Exception as e:
            resultado = {"exitoso": False, "error": error_legible(e)}
        resultados.append({"disciplina": disciplina, "resultado": resultado})
        if resultado.get("exitoso"):
            q.put(("LOG", f"{disciplina}: {resultado['archivo_generado']} "
                          f"({resultado['total_materiales']} material(es), "
                          f"{resultado['total_paginas']} pagina(s))"))
        else:
            q.put(("WARN", f"{disciplina}: {resultado.get('error')}"))
    q.put(("DISC_COMPLETE", resultados, time.time() - t0))


# ============================================================================
# GENERAR EXCEL CON DATOS DE MATERIALES
# ============================================================================
def generar_excel_materiales(json_path, carpeta_base):
    """
    Genera un Excel 'Guía Materiales.xlsx' en carpeta_base con los datos del JSON.
    Una hoja por disciplina (ARQ, ESTR, ELEC, MEC).
    Actualiza si ya existe.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    json_path = Path(json_path)
    if not json_path.exists():
        return {"exitoso": False, "error": f"JSON no encontrado: {json_path}"}

    try:
        data = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as e:
        return {"exitoso": False, "error": f"Error leyendo JSON: {e}"}

    materiales = data.get("materiales", []) if isinstance(data, dict) else data
    if not isinstance(materiales, list):
        return {"exitoso": False, "error": "JSON debe contener array 'materiales'"}

    # Agrupar por disciplina
    por_disciplina = {
        "ARQUITECTONICOS": [],
        "ESTRUCTURALES": [],
        "MECANICOS": [],
        "ELECTRICOS": []
    }

    for item in materiales:
        consec = str(item.get("consecutivo", "")).strip()
        if not consec:
            continue
        # Determinar disciplina por prefijo del consecutivo
        # Extraer letras al inicio hasta el primer dígito
        prefijo = re.sub(r'\d.*', '', consec).upper()
        if prefijo == "ARQ":
            por_disciplina["ARQUITECTONICOS"].append(item)
        elif prefijo == "ESTR":
            por_disciplina["ESTRUCTURALES"].append(item)
        elif prefijo == "MEC":
            por_disciplina["MECANICOS"].append(item)
        elif prefijo == "ELEC":
            por_disciplina["ELECTRICOS"].append(item)

    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja vacía por defecto

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Crear hoja por disciplina
    nombres_hojas = {
        "ARQUITECTONICOS": "Arquitectónicos",
        "ESTRUCTURALES": "Estructurales",
        "MECANICOS": "Mecánicos",
        "ELECTRICOS": "Eléctricos"
    }

    total_materiales = 0
    for disciplina, nombre_hoja in nombres_hojas.items():
        items = por_disciplina[disciplina]
        if not items:
            continue

        ws = wb.create_sheet(title=nombre_hoja)

        # Encabezados
        encabezados = ["Consecutivo", "Familia", "Descripción", "Normativa", "Estado"]
        for col, header in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Datos
        for row_idx, item in enumerate(items, 2):
            ws.cell(row=row_idx, column=1, value=item.get("consecutivo", ""))
            ws.cell(row=row_idx, column=2, value=item.get("marca", ""))  # Familia = marca
            # v2.6.6: Descripcion = nombre del material segun el nombre de la
            # carpeta (item["nombre"]), NO el texto largo que ChatGPT extrae
            # del JSON (item["descripcion"]).
            ws.cell(row=row_idx, column=3, value=item.get("nombre", ""))
            ws.cell(row=row_idx, column=4, value=item.get("normativa", "SIN ESPECIFICAR"))

            # Estado: si la carpeta está vacía o estado != FICHA_DISPONIBLE -> FALTANTE
            estado = "FALTANTE"
            if item.get("estado") == "FICHA_DISPONIBLE" and not item.get("carpeta_vacia"):
                estado = "DISPONIBLE"
            ws.cell(row=row_idx, column=5, value=estado)

        # Ajustar anchos de columna
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 15

        total_materiales += len(items)

    # Guardar
    ruta_excel = Path(carpeta_base) / "Guía Materiales.xlsx"
    try:
        wb.save(str(ruta_excel))
    except Exception as e:
        return {"exitoso": False, "error": f"Error guardando Excel: {e}"}

    return {
        "exitoso": True,
        "archivo": str(ruta_excel),
        "total_materiales": total_materiales,
    }


# ============================================================================
# CONSTRUIR datos_materiales.json
# ============================================================================
def construir_materiales(base, api_key, q, cancelar, gc=None):
    entradas = []
    for madre, pfx in MADRE_A_PREFIJO.items():
        d = base / madre
        if not d.is_dir():
            continue
        for sub in sorted(d.iterdir()):
            if not sub.is_dir():
                continue
            m = re.match(rf"^({pfx})(\d+)-(.*)$", sub.name)
            if not m:
                q.put(("WARN", f"Subcarpeta ignorada (nombre no valido): {sub.name}"))
                continue
            entradas.append((m.group(1), int(m.group(2)), m.group(3), sub))
    entradas.sort(key=lambda e: (ORDEN_CAT.get(e[0], 9), e[1]))

    total = len(entradas)
    docs_totales = 0
    q.put(("LOG", f"Subcarpetas detectadas: {total}"))
    materiales = []
    vistos = {}
    for i, (cat, num, nombre, sub) in enumerate(entradas, 1):
        if cancelar.is_set():
            raise _Cancelado()
        cons = f"{cat}{num:02d}"
        clave = nombre.upper()
        if clave in vistos:
            vistos[clave] += 1
            nombre = f"{nombre} ({vistos[clave]})"
        else:
            vistos[clave] = 1
        q.put(("JSON_PROG", i, total, cons))

        docs = buscar_archivos_carpeta(sub)
        docs_totales += len(docs)
        nombres_docs = [d.name for d in docs]

        if not docs:
            materiales.append({
                "consecutivo": cons, "nombre": nombre, "categoria": cat,
                "marca": "", "descripcion": "", "normativa": SIN_ESP,
                "idioma_original": "", "fue_traducido": False,
                "documentos_encontrados": [], "compilado_generado": None,
                "estado": "CARPETA_VACÍA", "carpeta_vacia": True,
                "aspectos_adicionales": "", "ruta_carpeta": str(sub)})
            continue

        bloques, textos_por_doc, tot = [], [], 0
        for d in docs:
            if cancelar.is_set():
                raise _Cancelado()
            try:
                t = extraer_texto_robusto(d)
                if t.strip():
                    seg = t[:1500]
                    # v2.6.5: si el texto plano parece una tabla tecnica densa
                    # (dimensiones/tolerancias "aplanadas" sin fila/columna),
                    # se relee el PDF con pdfplumber para conservar la
                    # estructura; sin esto, ChatGPT recibe numeros sueltos sin
                    # sentido y la marca/descripcion salen vacias o inventadas.
                    if d.suffix.lower() == ".pdf" and es_tabla_tecnica_densa(seg):
                        resumen_tabla = extraer_resumen_tabla_pdfplumber(d)
                        if resumen_tabla:
                            seg = ("[TABLA DE ESPECIFICACIONES TECNICAS - "
                                   "releida por columnas]\n" + resumen_tabla)[:1500]
                            q.put(("LOG", f"{cons}: '{d.name}' parece tabla "
                                          "tecnica densa; releida con pdfplumber"))
                    bloques.append(f"DOCUMENTO ({d.name}):\n{seg}")
                    textos_por_doc.append((d.name, seg))
                    tot += len(seg)
                    if tot >= MAX_COMBINADO:
                        break
                else:
                    q.put(("WARN", f"{cons}: '{d.name}' sin texto legible"))
            except Exception as e:
                q.put(("WARN", f"{cons}: error leyendo '{d.name}' ({e})"))

        compilado = sanitizar(f"{cons}-{nombre}-CMP") + ".pdf"

        if not bloques:
            materiales.append({
                "consecutivo": cons, "nombre": nombre, "categoria": cat,
                "marca": "POR DEFINIR",
                "descripcion": "Ficha tecnica no legible o incompleta",
                "normativa": SIN_ESP, "idioma_original": "", "fue_traducido": False,
                "documentos_encontrados": nombres_docs, "compilado_generado": compilado,
                "estado": "FICHA_INCOMPLETA", "carpeta_vacia": False,
                "aspectos_adicionales": "", "ruta_carpeta": str(sub)})
            continue

        # v2.6: si hay 2+ fichas legibles en la carpeta, se le consulta a la
        # misma IA (OpenAI) como se relacionan entre si, para redactar los
        # "aspectos adicionales" de la caratula o advertir de una posible
        # mezcla incorrecta de materiales distintos en una misma carpeta.
        aspectos_adicionales = ""
        if len(textos_por_doc) >= 2:
            # v2.6.4: tambien se renderiza la portada de cada ficha, porque el
            # nombre del fabricante muchas veces solo aparece como logo grafico
            # (no como texto extraible) y sin la imagen la IA no puede
            # distinguir marcas distintas (ver analizar_relacion_fichas).
            imagenes_portada = [_render_portada_b64(sub / nombre_doc)
                               for nombre_doc, _ in textos_por_doc]
            relacion_info = analizar_relacion_fichas(
                textos_por_doc, api_key, q, cons, imagenes_portada)
            if relacion_info:
                aspectos_adicionales = construir_texto_aspectos_adicionales(
                    relacion_info, len(textos_por_doc))
                if relacion_info["relacion"] == "DISCREPANCIA":
                    detalle = ("; ".join(relacion_info["tipos_producto"])
                               or relacion_info["justificacion"]
                               or "tipos de producto no compatibles entre si")
                    nombres_fichas = ", ".join(n for n, _ in textos_por_doc)
                    msg = (f"Posible error de organizacion: la carpeta '{nombre}' "
                           f"contiene fichas tecnicas de productos distintos e "
                           f"incompatibles entre si ({detalle}). "
                           f"Documentos: {nombres_fichas}. Verifique que no se "
                           f"hayan mezclado fichas de materiales diferentes en "
                           f"esta carpeta.")
                    if gc:
                        gc.log(logging.WARNING, msg, cons)
                    else:
                        q.put(("WARN", f"{cons}: {msg}"))

        combinado = "\n\n".join(bloques)
        try:
            marca, desc, norm, idioma, trad = extraer_con_chatgpt(combinado, api_key, q, cons)
            if not marca or marca.upper() in (SIN_ESP, "NO ESPECIFICADA"):
                marca = "POR DEFINIR"
            desc = _trunc_desc(desc) or "Sin especificacion disponible"
            if not _desc_parece_valida(desc):
                # v2.6.5: ChatGPT devolvio una "descripcion" que en realidad es
                # ruido numerico (tabla mal interpretada) -> se usa un texto
                # generico en vez de guardar basura en el submittal.
                q.put(("WARN", f"{cons}: descripcion de ChatGPT no parece valida "
                               f"('{desc[:60]}'); se usa una generica"))
                desc = ("Especificación técnica de material de obra. Ver ficha "
                       "técnica adjunta para el detalle dimensional completo.")
            norm = _trunc_norm(norm)
            if not norm or norm.upper() in ("NO ESPECIFICADA", "NINGUNA", "N/A"):
                norm = SIN_ESP
            estado = "FICHA_DISPONIBLE"
            if trad:
                q.put(("WARN", f"[TRADUCCION] {cons}: ficha en {idioma} traducida al espanol"))
        except _AuthError:
            raise
        except Exception as e:
            q.put(("WARN", f"{cons}: ChatGPT fallo ({e}); marcado incompleto"))
            marca, desc, norm, estado = ("POR DEFINIR",
                                         "Ficha tecnica no legible o incompleta",
                                         SIN_ESP, "FICHA_INCOMPLETA")
            idioma, trad = "", False

        materiales.append({
            "consecutivo": cons, "nombre": nombre, "categoria": cat,
            "marca": marca, "descripcion": desc, "normativa": norm,
            "idioma_original": idioma, "fue_traducido": trad,
            "documentos_encontrados": nombres_docs, "compilado_generado": compilado,
            "estado": estado, "carpeta_vacia": False,
            "aspectos_adicionales": aspectos_adicionales, "ruta_carpeta": str(sub)})
    return materiales, docs_totales


def resumen_materiales(m, docs_totales=0):
    return {
        "total": len(m),
        "fichas_disponibles": sum(x["estado"] == "FICHA_DISPONIBLE" for x in m),
        "carpetas_vacias": sum(bool(x["carpeta_vacia"]) for x in m),
        "fichas_incompletas": sum(x["estado"] == "FICHA_INCOMPLETA" for x in m),
        "documentos_totales": docs_totales,
        "compilados_generados": sum(1 for x in m if x.get("documentos_encontrados")),
        "traducidos": sum(1 for x in m if x.get("fue_traducido")),
    }


# ============================================================================
# DETECCION Y RESOLUCION DE MATERIALES DUPLICADOS (v2.6)
# ============================================================================
CARPETA_CUARENTENA = "_DUPLICADOS_ELIMINADOS"
# v2.6.1: umbrales mas estrictos (eran demasiado laxos: materiales de la misma
# familia/marca y normativa, pero de tamano o tipo distinto, se marcaban como
# duplicados solo por compartir texto de plantilla en la descripcion).
UMBRAL_DUPLICADO = 0.95
UMBRAL_DUPLICADO_MISMO_NOMBRE = 0.75

_RE_DIM_UNIDAD = re.compile(r'\d+(?:[.,]\d+)?\s*(?:MM|CM|IN|PULG|")', re.IGNORECASE)
_RE_DIM_PAR = re.compile(r'\d+(?:[.,]\d+)?\s*[xX]\s*\d+(?:[.,]\d+)?')
_RE_DIM_FRACCION = re.compile(r'\b\d+\s*/\s*\d+\b')


def _nombre_base(nombre):
    """Quita el sufijo ' (N)' que construir_materiales() agrega en memoria
    cuando dos carpetas comparten el mismo nombre, para comparar el nombre
    'real' del material sin ese contador."""
    return re.sub(r'\s*\(\d+\)\s*$', '', str(nombre)).strip()


def _clave_similitud(item):
    return re.sub(r'\s+', ' ',
                  f"{item.get('marca', '')} {item.get('descripcion', '')} "
                  f"{item.get('normativa', '')}").strip().upper()


def _extraer_dimensiones(*textos):
    """Extrae medidas/dimensiones (ej. '50MM', '2IN', '70x1.5', '3/4') de uno o
    mas textos, normalizadas para comparar. Se usa para evitar marcar como
    duplicados materiales de la misma familia/marca pero de tamano distinto
    (ej. costanera 3"x2" vs 4"x2", union PVC de 50mm vs 100mm)."""
    texto = " ".join(str(t) for t in textos if t)
    tokens = set()
    for patron in (_RE_DIM_UNIDAD, _RE_DIM_PAR, _RE_DIM_FRACCION):
        for m in patron.finditer(texto):
            tokens.add(re.sub(r'\s+', '', m.group(0)).upper().replace(',', '.'))
    return tokens


def _dimensiones_compatibles(a, b):
    """False si ambos materiales tienen medidas detectadas y estas difieren:
    en ese caso NO deben considerarse duplicados aunque el resto del texto
    (marca, normativa, texto de plantilla de la ficha) sea casi identico."""
    dims_a = _extraer_dimensiones(a.get("nombre", ""), a.get("descripcion", ""))
    dims_b = _extraer_dimensiones(b.get("nombre", ""), b.get("descripcion", ""))
    if dims_a and dims_b and dims_a != dims_b:
        return False
    return True


def detectar_duplicados(materiales, umbral=UMBRAL_DUPLICADO):
    """
    Detecta materiales cuya marca+descripcion+normativa (o nombre) son
    practicamente identicos a los de otro material del submittal, lo que
    sugiere que la MISMA ficha tecnica quedo registrada dos veces bajo
    consecutivos (y carpetas) distintos.

    Es deliberadamente estricto: materiales de la misma familia/marca/norma
    pero de TAMANO distinto (costanera 3"x2" vs 4"x2", union PVC de 50mm vs
    100mm) NUNCA se marcan como duplicados, sin importar que tan parecido sea
    el resto del texto (comparten mucho texto de plantilla/normativa). Un
    material solo entra a un grupo si coincide con TODOS los que ya estan en
    el grupo (no solo con el primero), para evitar cadenas de similitud que
    terminen agrupando materiales que en realidad son distintos entre si.

    Solo compara materiales con ficha disponible y marca identificada
    (carpetas vacias o con ficha ilegible no aportan datos para comparar).

    Devuelve una lista de grupos:
        [{"materiales": [item, item, ...], "similitud_min": float}, ...]
    cada uno con 2 o mas materiales relacionados entre si.
    """
    activos = [m for m in materiales
              if not m.get("carpeta_vacia") and m.get("estado") == "FICHA_DISPONIBLE"
              and str(m.get("marca", "")).strip().upper() not in ("", "POR DEFINIR")]

    grupos = []
    usados = set()
    for i in range(len(activos)):
        a = activos[i]
        if a["consecutivo"] in usados:
            continue
        grupo = [a]
        similitudes = []
        for j in range(i + 1, len(activos)):
            b = activos[j]
            if b["consecutivo"] in usados:
                continue
            if not all(_dimensiones_compatibles(existente, b) for existente in grupo):
                continue
            ratios_con_grupo = []
            coincide_con_todos = True
            for existente in grupo:
                ratio = difflib.SequenceMatcher(
                    None, _clave_similitud(existente), _clave_similitud(b)).ratio()
                mismo_nombre = (_nombre_base(existente["nombre"]).upper()
                               == _nombre_base(b["nombre"]).upper())
                if ratio >= umbral or (mismo_nombre and ratio >= UMBRAL_DUPLICADO_MISMO_NOMBRE):
                    ratios_con_grupo.append(ratio)
                else:
                    coincide_con_todos = False
                    break
            if coincide_con_todos:
                grupo.append(b)
                similitudes.extend(ratios_con_grupo)
        if len(grupo) > 1:
            for g in grupo:
                usados.add(g["consecutivo"])
            grupos.append({"materiales": grupo,
                          "similitud_min": round(min(similitudes), 2) if similitudes else 1.0})
    return grupos


def resolver_duplicados(base, materiales, consecutivos_a_eliminar, q=None):
    """
    Elimina (mueve a cuarentena reversible) los materiales indicados y
    reordena los consecutivos de cada disciplina afectada para cerrar los
    huecos que deja la eliminacion, dejando el submittal listo para volver
    a generar las caratulas.

    Afectaciones que se resuelven aqui:
      1. La carpeta del material eliminado se MUEVE (no se borra) a
         base/_DUPLICADOS_ELIMINADOS/<consecutivo>-<nombre>_<timestamp>/,
         para poder recuperarla si la eliminacion fue un error.
      2. Los consecutivos posteriores de la MISMA disciplina se renumeran
         para cerrar el hueco (ej. ARQ01,ARQ03,ARQ04 -> ARQ01,ARQ02,ARQ03),
         renombrando las carpetas en disco en orden ascendente (evita
         colisiones porque la numeracion solo se comprime, nunca se invierte).
      3. En las carpetas renombradas se borran la CARATULA*.pdf y el
         *-CMP.pdf existentes (referencian el consecutivo/nombre de archivo
         viejo); se regeneraran limpias en la proxima corrida.
      4. El campo 'compilado_generado' de cada material renumerado se
         recalcula con el nuevo consecutivo.
      5. El compilado por disciplina (CMP SUBMITTAL <disciplina>.pdf) de
         cada disciplina afectada queda obsoleto (su contenido y paginacion
         cambiaron) y se elimina; debe regenerarse con "Generar Compilados".

    Devuelve un dict con el detalle de lo realizado y la lista actualizada
    de materiales (ya sin los eliminados y con los consecutivos al dia).
    """
    base = Path(base)
    a_eliminar = set(consecutivos_a_eliminar)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    eliminados = []
    disciplinas_afectadas = set()
    restantes = []

    for m in materiales:
        cons = m["consecutivo"]
        if cons not in a_eliminar:
            restantes.append(m)
            continue
        origen = Path(m["ruta_carpeta"])
        if not origen.is_absolute():
            origen = base / origen
        cuarentena_raiz = base / CARPETA_CUARENTENA
        destino = cuarentena_raiz / f"{cons}-{sanitizar(m.get('nombre', cons))}_{ts}"
        try:
            if origen.exists():
                cuarentena_raiz.mkdir(parents=True, exist_ok=True)
                shutil.move(str(origen), str(destino))
            disciplinas_afectadas.add(m["categoria"])
            eliminados.append({"consecutivo": cons, "nombre": m.get("nombre", ""),
                              "ruta_cuarentena": str(destino)})
            if q:
                q.put(("LOG", f"{cons}: material duplicado movido a cuarentena -> {destino}"))
        except Exception as e:
            if q:
                q.put(("WARN", f"{cons}: no se pudo mover la carpeta a cuarentena ({e}); "
                                f"se conserva sin cambios"))
            restantes.append(m)

    renumerados = []
    for cat in disciplinas_afectadas:
        items_cat = sorted(
            (m for m in restantes if m["categoria"] == cat),
            key=lambda m: int(m["consecutivo"][len(cat):]))
        for nuevo_num, m in enumerate(items_cat, 1):
            nuevo_cons = f"{cat}{nuevo_num:02d}"
            if nuevo_cons == m["consecutivo"]:
                continue
            viejo_cons = m["consecutivo"]
            carpeta_vieja = Path(m["ruta_carpeta"])
            mm = re.match(rf'^{cat}\d+-(.*)$', carpeta_vieja.name)
            nombre_disco = mm.group(1) if mm else m.get("nombre", nuevo_cons)
            carpeta_nueva = carpeta_vieja.parent / f"{nuevo_cons}-{nombre_disco}"
            try:
                if carpeta_vieja.exists():
                    os.rename(str(carpeta_vieja), str(carpeta_nueva))
                for p in (carpeta_nueva.glob("*.pdf") if carpeta_nueva.exists() else []):
                    up = p.name.upper()
                    if up.startswith("CARATULA") or up.endswith("-CMP.PDF"):
                        try:
                            p.unlink()
                        except Exception:
                            pass
                m["consecutivo"] = nuevo_cons
                m["ruta_carpeta"] = str(carpeta_nueva)
                m["compilado_generado"] = sanitizar(f"{nuevo_cons}-{nombre_disco}-CMP") + ".pdf"
                renumerados.append({"categoria": cat, "consecutivo_viejo": viejo_cons,
                                   "consecutivo_nuevo": nuevo_cons, "nombre": nombre_disco})
                if q:
                    q.put(("LOG", f"Reordenado: {viejo_cons} -> {nuevo_cons} ({nombre_disco})"))
            except Exception as e:
                if q:
                    q.put(("WARN", f"No se pudo renombrar {viejo_cons} -> {nuevo_cons}: {e}"))

    restantes.sort(key=lambda m: (ORDEN_CAT.get(m["categoria"], 9),
                                  int(m["consecutivo"][len(m["categoria"]):])))

    compilados_obsoletos = []
    for cat in disciplinas_afectadas:
        madre = next((k for k, v in MADRE_A_PREFIJO.items() if v == cat), None)
        if not madre:
            continue
        disc_sing = DISCIPLINA_SINGULAR.get(madre, madre)
        cmp_path = base / madre / f"CMP SUBMITTAL {disc_sing}.pdf"
        if cmp_path.exists():
            try:
                cmp_path.unlink()
                compilados_obsoletos.append(str(cmp_path))
                if q:
                    q.put(("LOG", f"Compilado de disciplina obsoleto eliminado: {cmp_path.name}"))
            except Exception as e:
                if q:
                    q.put(("WARN", f"No se pudo eliminar compilado obsoleto {cmp_path.name}: {e}"))

    return {"eliminados": eliminados, "renumerados": renumerados,
            "materiales": restantes, "disciplinas_afectadas": sorted(disciplinas_afectadas),
            "compilados_obsoletos": compilados_obsoletos}


# ============================================================================
# CARGA DEL MOTOR
# ============================================================================
def _cargar_motor(base, q):
    ruta_motor = base / NOMBRE_MOTOR
    if not ruta_motor.exists():
        q.put(("ERROR", f"No se encontro {NOMBRE_MOTOR} en la carpeta base:\n{base}"))
        return None
    lg = logging.getLogger("caratulas")
    for h in list(lg.handlers):
        try:
            h.close()
        except Exception:
            pass
        lg.removeHandler(h)
    spec = importlib.util.spec_from_file_location("generate_caratulas", str(ruta_motor))
    gc = importlib.util.module_from_spec(spec)
    sys.modules["generate_caratulas"] = gc
    spec.loader.exec_module(gc)
    for h in list(lg.handlers):
        if isinstance(h, logging.StreamHandler) and not isinstance(h, logging.FileHandler):
            if getattr(h, "stream", None) is None or sys.stdout is None:
                lg.removeHandler(h)

    class _QH(logging.Handler):
        def emit(self, record):
            try:
                q.put(("LOG", self.format(record)))
            except Exception:
                pass
    qh = _QH()
    qh.setFormatter(logging.Formatter("%(levelname)s %(consec)s %(message)s"))
    qh.addFilter(gc._CtxFilter())
    lg.addHandler(qh)
    return gc


# ============================================================================
# HILO DE TRABAJO PRINCIPAL
# ============================================================================
def hilo_trabajo(modo, carpeta_base, ruta_json, api_key, opciones,
                 caratula, datos_proyecto, q, cancelar):
    t0 = time.time()
    base = Path(carpeta_base)
    try:
        gc = _cargar_motor(base, q)
        if gc is None:
            return

        docs_totales = 0
        if modo == "auto":
            q.put(("FASE", "Leyendo fichas y extrayendo datos con ChatGPT..."))
            for lib in ("fitz", "pytesseract"):
                try:
                    __import__(lib)
                except Exception:
                    q.put(("WARN", f"Falta '{lib}': los PDF escaneados/imagenes podrian no leerse."))
            try:
                materiales, docs_totales = construir_materiales(base, api_key, q, cancelar, gc)
            except _AuthError:
                q.put(("ERROR", "La API Key de OpenAI es invalida o no tiene creditos.\n"
                                "Puede cambiarla desde el boton 'Mantenimiento'."))
                return
            except _Cancelado:
                q.put(("LOG", "Proceso cancelado por la usuaria."))
                q.put(("CANCELLED",))
                return
            if cancelar.is_set():
                q.put(("CANCELLED",))
                return
            res = resumen_materiales(materiales, docs_totales)
            q.put(("LOG", f"JSON: {res['total']} materiales, {res['documentos_totales']} documentos, "
                          f"{res['fichas_disponibles']} con ficha, {res['carpetas_vacias']} vacias, "
                          f"{res['fichas_incompletas']} incompletas, {res['traducidos']} traducidas"))
            out = {"resumen": res, "materiales": materiales}
            try:
                (base / NOMBRE_JSON).write_text(
                    json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
                q.put(("LOG", f"Guardado: {NOMBRE_JSON}"))
            except Exception as e:
                q.put(("WARN", f"No se pudo guardar {NOMBRE_JSON}: {e}"))
        else:
            jp = Path(ruta_json)
            if not jp.exists():
                q.put(("ERROR", f"No se encontro el archivo JSON:\n{jp}"))
                return
            try:
                data = json.loads(jp.read_text(encoding="utf-8"))
            except json.JSONDecodeError as e:
                q.put(("ERROR", error_legible(e)))
                return
            materiales = data.get("materiales") if isinstance(data, dict) else data
            if not isinstance(materiales, list) or not materiales:
                q.put(("ERROR", "El JSON no contiene la lista 'materiales' o esta vacia."))
                return
            for it in materiales:
                if isinstance(it, dict) and "normativa" not in it:
                    it["normativa"] = SIN_ESP

        # ---------- FASE 4: seleccionar plantilla y generar caratulas -------
        q.put(("FASE", f"Generando caratulas ({CARATULA_NOMBRES.get(caratula, caratula)})..."))
        try:
            import jinja2  # noqa
        except ImportError:
            q.put(("ERROR", "Falta jinja2.\nInstale: pip install jinja2 pypdf playwright"))
            return
        engines = gc.available_engines()
        if not engines:
            q.put(("ERROR", "No hay motor de PDF.\nInstale:\npip install playwright\n"
                            "python -m playwright install chromium"))
            return

        tpl_name, logo_rel, _logo_var = CARATULAS.get(caratula, CARATULAS["clasica"])
        tpl_path = resource_path(tpl_name)
        if not tpl_path.exists():
            q.put(("ERROR", f"No se encontro la plantilla '{tpl_name}'.\n"
                            "Si necesita esta caratula, solicite una actualizacion del software."))
            return
        tpl_text = tpl_path.read_text(encoding="utf-8")
        if not tpl_text.strip() or "<html" not in tpl_text.lower():
            q.put(("ERROR", f"La plantilla '{tpl_name}' esta vacia o corrupta."))
            return
        from jinja2 import Template
        template = Template(tpl_text)

        logo_path = resource_path(logo_rel)
        logo_uri = ""
        if logo_path.exists():
            logo_uri = gc.file_uri(logo_path)
        else:
            q.put(("WARN", f"No se encontro el logo '{logo_rel}'. La caratula se genera sin logo."))
        gc.LOGO_URI = logo_uri
        gc.stats["total"] = len(materiales)
        q.put(("LOG", f"Motores de PDF: {', '.join(n for n, _ in engines)}"))

        if opciones.get("forzar_regeneracion"):
            borrados, bloqueados = 0, 0
            for item in materiales:
                ruta = item.get("ruta_carpeta")
                if not ruta:
                    continue
                carpeta = gc.to_absolute(ruta)
                if carpeta.exists():
                    for p in carpeta.glob("*.pdf"):
                        up = p.name.upper()
                        if up.startswith("CARATULA") or up.endswith("-CMP.PDF"):
                            try:
                                p.unlink()
                                borrados += 1
                            except Exception:
                                bloqueados += 1
                                q.put(("WARN", f"No se pudo reemplazar '{p.name}': "
                                               f"cierrelo si esta abierto en un visor."))
            q.put(("LOG", f"Forzar (sobreescribir): {borrados} PDF(s) borrado(s)"
                          + (f", {bloqueados} bloqueado(s)" if bloqueados else "")))

        def _extra_ctx(item):
            if caratula != "ministerio_salud":
                return None
            dp = datos_proyecto or {}
            return {
                "logo_ministerio": logo_uri,
                "version": "v1",
                "registro": str(item.get("consecutivo", "")).strip(),
                "fecha_emision": dp.get("fecha_emision", ""),
                "proyecto": dp.get("proyecto", ""),
                "cliente": dp.get("cliente", ""),
                "plazo": dp.get("plazo", ""),
                "contrato": dp.get("contrato", ""),
                "monto": dp.get("monto", ""),
                "nombre_cargo": dp.get("nombre_cargo", ""),
                "fecha": dp.get("fecha", ""),
                "documentacion_tecnica": ("Ficha técnica" if item.get("documentos_encontrados") else ""),
                "observaciones_material": item.get("aspectos_adicionales", ""),
                "estado": "", "fecha_revision": "", "observaciones_respuesta": "", "revisa": "",
            }

        total = len(materiales)
        for i, item in enumerate(materiales, start=1):
            if cancelar.is_set():
                q.put(("CANCELLED",))
                return
            try:
                gc.process_material(item, template, engines, extra_ctx=_extra_ctx(item))
            except Exception as e:
                c = str(item.get("consecutivo", "?"))
                gc.stats["fallidos"] += 1
                gc.lst_errores.append((c, f"inesperado: {e}"))
                q.put(("WARN", f"{c}: {error_legible(e)}"))
            q.put(("PDF_GENERATED", i, total, dict(gc.stats)))

        # ---------- FASE 5: compilados --------------------------------------
        q.put(("FASE", "Compilando documentos (caratula + fichas)..."))
        compilados = 0
        solo_faltantes = opciones.get("solo_faltantes", True)
        for i, item in enumerate(materiales, start=1):
            if cancelar.is_set():
                q.put(("CANCELLED",))
                return
            q.put(("COMP_PROG", i, total))
            if item.get("carpeta_vacia"):
                continue
            carpeta = gc.to_absolute(item.get("ruta_carpeta", ""))
            if not carpeta.exists():
                continue
            docs = buscar_archivos_carpeta(carpeta)
            if not docs:
                continue
            caratulas = [p for p in carpeta.glob("*.pdf") if p.name.upper().startswith("CARATULA")]
            if not caratulas:
                continue
            cons = item.get("consecutivo", "?")
            nombre = item.get("nombre", cons)
            cmp_name = item.get("compilado_generado") or (sanitizar(f"{cons}-{nombre}-CMP") + ".pdf")
            cmp_path = carpeta / cmp_name
            if solo_faltantes and cmp_path.exists():
                continue
            try:
                anexos = generar_compilado(caratulas[0], docs, cmp_path, q, cons)
                compilados += 1
                q.put(("LOG", f"{cons}: compilado -> {cmp_name} ({anexos} doc(s))"))
            except Exception as e:
                q.put(("WARN", f"{cons}: no se pudo compilar ({error_legible(e)})"))

        gc.write_report()

        # ---------- GENERAR EXCEL CON DATOS DE MATERIALES -----
        q.put(("LOG", "Generando Excel con datos de materiales..."))
        try:
            resultado_excel = generar_excel_materiales(base / NOMBRE_JSON, base)
            if resultado_excel.get("exitoso"):
                q.put(("LOG", f"✅ Excel generado: {resultado_excel['archivo']} "
                              f"({resultado_excel['total_materiales']} materiales)"))
            else:
                q.put(("WARN", f"No se pudo generar Excel: {resultado_excel.get('error')}"))
        except Exception as e:
            q.put(("WARN", f"Error generando Excel: {error_legible(e)}"))

        stats_final = dict(gc.stats)
        stats_final["compilados"] = compilados
        stats_final["documentos_totales"] = docs_totales
        q.put(("COMPLETE", stats_final, time.time() - t0))

    except Exception as e:
        q.put(("ERROR", error_legible(e)))


def hilo_test_api(api_key, q, evento="APITEST"):
    ok, msg = test_openai(api_key)
    q.put((evento, ok, msg, api_key))


# ============================================================================
# APLICACION PRINCIPAL
# ============================================================================
class SubmitalsGUI(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(f"Generador de Submittals - ES Constructora | v{VERSION} | Elaborado por {AUTOR}")
        _aplicar_icono_ventana(self)
        self.resizable(True, True)
        self.minsize(700, 480)
        self.configure(bg=GRIS_BG)

        self.cfg = cargar_config()

        # Crear carpetas necesarias si no existen
        self._crear_carpetas_iniciales()
        self.cola = queue.Queue()
        self.cancelar = threading.Event()
        self.procesando = False
        self.compilando_disc = False   # v2.5.2: compilado por disciplina en curso
        self.thread_activo = None      # v2.4: control de hilo
        self.log_lineas = []
        self.maint = None

        self.var_api = tk.StringVar(value=cargar_api_key(self.cfg))
        self.var_auto = tk.BooleanVar(value=self.cfg["opciones"].get("generar_json_automatico", True))
        self.var_existente = tk.BooleanVar(value=self.cfg["opciones"].get("usar_json_existente", False))
        if not self.var_auto.get() and not self.var_existente.get():
            self.var_auto.set(True)

        self.var_caratula = tk.StringVar(value=self.cfg.get("caratula_seleccionada", "clasica"))
        self.datos_proyecto = dict(self.cfg.get("datos_proyecto", {}))

        self.var_carpeta = tk.StringVar()
        self.var_json = tk.StringVar()
        self.var_solo = tk.BooleanVar(value=self.cfg["opciones"]["solo_faltantes"])
        self.var_forzar = tk.BooleanVar(value=self.cfg["opciones"]["forzar_regeneracion"])
        self.var_log = tk.BooleanVar(value=self.cfg["opciones"]["mostrar_log"])

        self._construir_ui()
        self._cargar_valores_iniciales()
        self._on_caratula(inicial=True)
        self._validar()
        self.after(100, self._revisar_cola)
        self.protocol("WM_DELETE_WINDOW", self._cerrar)
        self._ajustar_tamano_pantalla()
        if auto_updater and auto_updater.configurado():
            threading.Thread(target=self._chk_update, daemon=True).start()

    def _ajustar_tamano_pantalla(self):
        """Ajusta el tamano inicial de la ventana para que siempre quepa en pantalla."""
        self.update_idletasks()
        pantalla_h = self.winfo_screenheight()
        pantalla_w = self.winfo_screenwidth()
        ancho = 780
        # Deja margen para la barra de tareas y el borde de la ventana
        alto = min(760, pantalla_h - 80)
        x = max(0, (pantalla_w - ancho) // 2)
        y = max(0, (pantalla_h - alto) // 2 - 20)
        self.geometry(f"{ancho}x{alto}+{x}+{y}")

    def _crear_carpetas_iniciales(self):
        """Crea las carpetas necesarias si no existen."""
        try:
            base = Path(DEFAULT_BASE)
            base.mkdir(parents=True, exist_ok=True)

            # Crear carpetas de disciplinas
            for disciplina in CARPETAS_MADRE:
                (base / disciplina).mkdir(parents=True, exist_ok=True)

            # v2.6.2: YA NO se crea aqui una carpeta "ms-playwright" vacia.
            # Una carpeta vacia con ese nombre hacia que el arranque (arriba)
            # redirigiera Playwright hacia ella en vez de usar la cache global
            # donde Chromium SI esta instalado, rompiendo el render (caida
            # silenciosa a pdfkit, con peor soporte de CSS moderno).

        except Exception as e:
            pass  # Fallar silenciosamente, no es crítico

    # ------------------------------------------------------------------ UI --
    def _construir_ui(self):
        head = tk.Frame(self, bg=AZUL_ES)
        head.pack(fill="x")
        fila_titulo = tk.Frame(head, bg=AZUL_ES)
        fila_titulo.pack(pady=(8, 6))
        try:
            logo_path = resource_path(ICONO_HEADER_PNG)
            if logo_path.exists():
                self._logo_header_img = tk.PhotoImage(file=str(logo_path))
                tk.Label(fila_titulo, image=self._logo_header_img,
                         bg=AZUL_ES).pack(side="left", padx=(0, 10))
        except Exception:
            pass
        texto_titulo = tk.Frame(fila_titulo, bg=AZUL_ES)
        texto_titulo.pack(side="left")
        tk.Label(texto_titulo, text="GENERADOR DE SUBMITTALS", bg=AZUL_ES, fg=BLANCO,
                 font=("Segoe UI", 14, "bold")).pack()
        tk.Label(texto_titulo, text=f"ES CONSTRUCTORA  ·  v{VERSION}  ·  Elaborado por {AUTOR}",
                 bg=AZUL_ES, fg="#BFC8DC", font=("Segoe UI", 8)).pack()
        tk.Frame(self, bg=ROJO_ES, height=3).pack(fill="x")

        # Contenedor con scroll: garantiza que TODO el contenido sea accesible
        # aunque la pantalla del usuario sea pequena y la ventana no entre completa.
        contenedor = tk.Frame(self, bg=GRIS_BG)
        contenedor.pack(fill="both", expand=True)

        self._canvas_ui = tk.Canvas(contenedor, bg=GRIS_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(contenedor, orient="vertical", command=self._canvas_ui.yview)
        self._canvas_ui.configure(yscrollcommand=scrollbar.set)
        self._canvas_ui.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        cuerpo = tk.Frame(self._canvas_ui, bg=GRIS_BG, padx=24, pady=6)
        cuerpo_id = self._canvas_ui.create_window((0, 0), window=cuerpo, anchor="nw")

        def _on_cuerpo_configure(_e=None):
            self._canvas_ui.configure(scrollregion=self._canvas_ui.bbox("all"))
        cuerpo.bind("<Configure>", _on_cuerpo_configure)

        def _on_canvas_configure(event):
            self._canvas_ui.itemconfig(cuerpo_id, width=event.width)
        self._canvas_ui.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(event):
            self._canvas_ui.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self._canvas_ui.bind("<Enter>", lambda _e: self._canvas_ui.bind_all("<MouseWheel>", _on_mousewheel))
        self._canvas_ui.bind("<Leave>", lambda _e: self._canvas_ui.unbind_all("<MouseWheel>"))

        # SECCION MODO DE GENERACION + CONFIGURACION
        fh = tk.Frame(cuerpo, bg=GRIS_BG)
        fh.pack(fill="x", pady=(4, 1))
        tk.Label(fh, text="⚙️ MODO DE GENERACION", bg=GRIS_BG, fg=ROJO_ES,
                 font=("Segoe UI", 11, "bold")).pack(side="left")
        self.btn_maint = tk.Button(fh, text="🔧 Configuracion", command=self._abrir_mantenimiento,
                                   bg=AZUL_ES, fg=BLANCO, font=("Segoe UI", 8),
                                   relief="flat", padx=8, pady=2, cursor="hand2")
        self.btn_maint.pack(side="right")
        tk.Frame(cuerpo, bg="#C9CED6", height=1).pack(fill="x", pady=(0, 4))

        chk_auto = tk.Checkbutton(cuerpo, text="Generar JSON automaticamente desde fichas (requiere API Key)",
                                  variable=self.var_auto, bg=GRIS_BG, anchor="w",
                                  font=("Segoe UI", 10), command=self._chk_auto)
        chk_auto.pack(fill="x")
        chk_exist = tk.Checkbutton(cuerpo, text="Usar JSON pre-existente",
                                   variable=self.var_existente, bg=GRIS_BG, anchor="w",
                                   font=("Segoe UI", 10), command=self._chk_existente)
        chk_exist.pack(fill="x")
        self.lbl_api = tk.Label(cuerpo, text="", bg=GRIS_BG, font=("Segoe UI", 8), anchor="w")
        self.lbl_api.pack(fill="x", pady=(1, 0))

        # SECCION TIPO DE CARATULA
        self._titulo_seccion(cuerpo, "🧾 TIPO DE CARÁTULA")
        fsel = tk.Frame(cuerpo, bg=GRIS_BG)
        fsel.pack(fill="x")
        tk.Radiobutton(fsel, text="Clásica (ES Constructora)", variable=self.var_caratula,
                       value="clasica", bg=GRIS_BG, font=("Segoe UI", 10),
                       command=self._on_caratula).pack(side="left")
        tk.Radiobutton(fsel, text="Ministerio de Salud", variable=self.var_caratula,
                       value="ministerio_salud", bg=GRIS_BG, font=("Segoe UI", 10),
                       command=self._on_caratula).pack(side="left", padx=(14, 0))
        self.btn_datos = tk.Button(fsel, text="📋 Datos del proyecto", command=self._abrir_datos_proyecto,
                                   bg=AZUL_ES, fg=BLANCO, relief="flat", font=("Segoe UI", 8),
                                   padx=8, cursor="hand2")
        self.btn_datos.pack(side="right")
        tk.Label(cuerpo, text="⚠ Si la carátula que necesita no está en esta lista, solicite al "
                              "administrador una actualización del software.",
                 bg=GRIS_BG, fg=AMARILLO, font=("Segoe UI", 8), wraplength=660,
                 justify="left", anchor="w").pack(fill="x", pady=(2, 0))

        # SECCION PROYECTO
        self._titulo_seccion(cuerpo, "📁 CONFIGURACION DEL PROYECTO")
        tk.Label(cuerpo, text="Carpeta Base del Proyecto:", bg=GRIS_BG,
                 fg=AZUL_ES, font=("Segoe UI", 10, "bold"), anchor="w").pack(fill="x")
        f1 = tk.Frame(cuerpo, bg=GRIS_BG)
        f1.pack(fill="x", pady=(2, 2))
        self.cbo_carpeta = ttk.Combobox(f1, textvariable=self.var_carpeta, state="readonly",
                                        font=("Segoe UI", 9), values=self.cfg["carpetas_recientes"])
        self.cbo_carpeta.pack(side="left", fill="x", expand=True, ipady=3)
        self.cbo_carpeta.bind("<<ComboboxSelected>>", lambda e: self._carpeta_cambiada())
        tk.Button(f1, text="Examinar…", command=self._examinar_carpeta, bg=AZUL_ES, fg=BLANCO,
                  font=("Segoe UI", 9), relief="flat", padx=10, cursor="hand2").pack(side="left", padx=(6, 0))
        self.lbl_val_carpeta = tk.Label(cuerpo, text="", bg=GRIS_BG, font=("Segoe UI", 8), anchor="w")
        self.lbl_val_carpeta.pack(fill="x")

        self.lbl_json_titulo = tk.Label(cuerpo, text="Archivo JSON de Datos:", bg=GRIS_BG,
                                        fg=AZUL_ES, font=("Segoe UI", 10, "bold"), anchor="w")
        self.lbl_json_titulo.pack(fill="x", pady=(6, 0))
        f2 = tk.Frame(cuerpo, bg=GRIS_BG)
        f2.pack(fill="x", pady=(2, 2))
        self.ent_json = tk.Entry(f2, textvariable=self.var_json, state="readonly", font=("Segoe UI", 9))
        self.ent_json.pack(side="left", fill="x", expand=True, ipady=3)
        self.btn_json = tk.Button(f2, text="Examinar…", command=self._examinar_json, bg=AZUL_ES, fg=BLANCO,
                                  font=("Segoe UI", 9), relief="flat", padx=10, cursor="hand2")
        self.btn_json.pack(side="left", padx=(6, 0))
        self.lbl_val_json = tk.Label(cuerpo, text="", bg=GRIS_BG, font=("Segoe UI", 8), anchor="w")
        self.lbl_val_json.pack(fill="x")

        # SECCION OPCIONES
        self._titulo_seccion(cuerpo, "⚙ OPCIONES")
        c1 = tk.Checkbutton(cuerpo, text="Solo generar faltantes (caratulas y compilados)",
                            variable=self.var_solo, bg=GRIS_BG, anchor="w",
                            font=("Segoe UI", 10), command=self._chk_solo)
        c1.pack(fill="x")
        Tooltip(c1, "Salta las carpetas que ya tienen su caratula y su compilado.")
        c2 = tk.Checkbutton(cuerpo, text="Forzar regeneracion (SOBREESCRIBE las existentes)",
                            variable=self.var_forzar, bg=GRIS_BG, anchor="w",
                            font=("Segoe UI", 10), command=self._chk_forzar)
        c2.pack(fill="x")
        Tooltip(c2, "Reemplaza las caratulas y compilados existentes por los nuevos.")
        c3 = tk.Checkbutton(cuerpo, text="Mostrar log al finalizar", variable=self.var_log,
                            bg=GRIS_BG, anchor="w", font=("Segoe UI", 10))
        c3.pack(fill="x")

        # SECCION COMPILADO POR DISCIPLINA (v2.5.2)
        self._titulo_seccion(cuerpo, "📦 COMPILADO POR DISCIPLINA")
        tk.Label(cuerpo, text="Genera un PDF unico por disciplina con TODAS sus caratulas y fichas "
                              "(ej: CMP SUBMITTAL ARQUITECTONICO.pdf). No usa ChatGPT ni consume API.",
                 bg=GRIS_BG, fg="#6C757D", font=("Segoe UI", 8), wraplength=660,
                 justify="left", anchor="w").pack(fill="x")
        fdisc = tk.Frame(cuerpo, bg=GRIS_BG)
        fdisc.pack(fill="x", pady=(2, 4))
        self.var_disc_arq = tk.BooleanVar(value=True)
        self.var_disc_estr = tk.BooleanVar(value=True)
        self.var_disc_mec = tk.BooleanVar(value=True)
        self.var_disc_elec = tk.BooleanVar(value=True)
        tk.Checkbutton(fdisc, text="ARQ", variable=self.var_disc_arq, bg=GRIS_BG,
                       font=("Segoe UI", 9)).pack(side="left")
        tk.Checkbutton(fdisc, text="ESTR", variable=self.var_disc_estr, bg=GRIS_BG,
                       font=("Segoe UI", 9)).pack(side="left", padx=(8, 0))
        tk.Checkbutton(fdisc, text="MEC", variable=self.var_disc_mec, bg=GRIS_BG,
                       font=("Segoe UI", 9)).pack(side="left", padx=(8, 0))
        tk.Checkbutton(fdisc, text="ELEC", variable=self.var_disc_elec, bg=GRIS_BG,
                       font=("Segoe UI", 9)).pack(side="left", padx=(8, 0))
        self.btn_compilado_disc = tk.Button(fdisc, text="📦 Generar Compilados",
                                            command=self._generar_compilados_disciplina,
                                            bg="#FF9800", fg=BLANCO, font=("Segoe UI", 9, "bold"),
                                            relief="flat", padx=10, cursor="hand2")
        self.btn_compilado_disc.pack(side="right")

        # BOTONES
        fbtn = tk.Frame(cuerpo, bg=GRIS_BG)
        fbtn.pack(pady=10)
        self.btn_generar = tk.Button(fbtn, text="🚀 GENERAR", command=self._generar,
                                     bg=ROJO_ES, fg=BLANCO, font=("Segoe UI", 13, "bold"),
                                     relief="flat", padx=24, pady=12, cursor="hand2",
                                     disabledforeground="#DDDDDD",
                                     activebackground="#B01623", activeforeground=BLANCO)
        self.btn_generar.pack(side="left")
        self.tip_generar = Tooltip(self.btn_generar, "")
        self.btn_cancelar = tk.Button(fbtn, text="✋ CANCELAR", command=self._cancelar_proceso,
                                      bg="#6C757D", fg=BLANCO, font=("Segoe UI", 11, "bold"),
                                      relief="flat", padx=16, pady=12, cursor="hand2", state="disabled")
        self.btn_cancelar.pack(side="left", padx=(10, 0))

        # PROGRESO
        self._titulo_seccion(cuerpo, "📊 PROGRESO")
        st = ttk.Style(self)
        st.theme_use("default")
        st.configure("Verde.Horizontal.TProgressbar", troughcolor="#E1E4EA",
                     background="#28A745", thickness=20)
        self.barra = ttk.Progressbar(cuerpo, style="Verde.Horizontal.TProgressbar", maximum=100, value=0)
        self.barra.pack(fill="x", pady=(2, 2))
        self.lbl_progreso = tk.Label(cuerpo, text="Listo para comenzar.", bg=GRIS_BG,
                                     fg=AZUL_ES, font=("Segoe UI", 9))
        self.lbl_progreso.pack()

        # RESULTADOS
        self._titulo_seccion(cuerpo, "📋 RESULTADOS")
        fres = tk.Frame(cuerpo, bg=BLANCO, relief="solid", borderwidth=1)
        fres.pack(fill="x", pady=(2, 4))
        self.lbl_ok = self._fila_resultado(fres, "✅ Caratulas:", VERDE_OK)
        self.lbl_cmp = self._fila_resultado(fres, "📚 Compilados:", AZUL_ES)
        self.lbl_skip = self._fila_resultado(fres, "⏭ Saltados:", AMARILLO)
        self.lbl_err = self._fila_resultado(fres, "❌ Errores:", ROJO_ES)
        self.lbl_tiempo = self._fila_resultado(fres, "⏳ Tiempo:", AZUL_ES)

        self.txt_log = tk.Text(cuerpo, height=4, font=("Consolas", 8), bg="#1E1E1E",
                               fg="#D4D4D4", state="disabled", relief="flat", wrap="none")
        self.txt_log.pack(fill="x", pady=(4, 4))

        facc = tk.Frame(cuerpo, bg=GRIS_BG)
        facc.pack(pady=6)
        est = dict(font=("Segoe UI", 9), relief="flat", padx=10, pady=6, cursor="hand2",
                   bg=AZUL_ES, fg=BLANCO)
        tk.Button(facc, text="📋 Ver Log Completo", command=self._ver_log, **est).grid(row=0, column=0, padx=4, pady=2)
        tk.Button(facc, text="📁 Abrir Resultados", command=self._abrir_resultados, **est).grid(row=0, column=1, padx=4, pady=2)
        tk.Button(facc, text="📊 Actualizar Excel", command=self._generar_excel_ahora, **est).grid(row=0, column=2, padx=4, pady=2)
        tk.Button(facc, text="🔄 Nuevo Proceso", command=self._nuevo_proceso, **est).grid(row=0, column=3, padx=4, pady=2)
        tk.Button(facc, text="❌ Cerrar", command=self._cerrar, font=("Segoe UI", 9), relief="flat",
                  padx=10, pady=6, cursor="hand2", bg="#6C757D", fg=BLANCO).grid(row=0, column=4, padx=4, pady=2)
        tk.Button(facc, text="🧬 Detectar Duplicados", command=self._detectar_duplicados_click,
                  font=("Segoe UI", 9), relief="flat", padx=10, pady=6, cursor="hand2",
                  bg="#FF9800", fg=BLANCO).grid(row=1, column=0, columnspan=5, padx=4, pady=(4, 2), sticky="ew")
        tk.Button(facc, text="🗑️ Borrar Carátulas y Compilados (empezar de 0)",
                  command=self._borrar_caratulas_compilados_click,
                  font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=6, cursor="hand2",
                  bg="#B00020", fg=BLANCO).grid(row=2, column=0, columnspan=5, padx=4, pady=(2, 2), sticky="ew")
        tk.Button(facc, text="📤 Exportar Versión Final",
                  command=self._exportar_version_final_click,
                  font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=6, cursor="hand2",
                  bg=VERDE_OK, fg=BLANCO).grid(row=3, column=0, columnspan=5, padx=4, pady=(2, 2), sticky="ew")

    def _titulo_seccion(self, parent, texto):
        tk.Label(parent, text=texto, bg=GRIS_BG, fg=ROJO_ES, font=("Segoe UI", 10, "bold"),
                 anchor="w").pack(fill="x", pady=(5, 1))
        tk.Frame(parent, bg="#C9CED6", height=1).pack(fill="x", pady=(0, 3))

    def _fila_resultado(self, parent, etiqueta, color):
        f = tk.Frame(parent, bg=BLANCO)
        f.pack(fill="x", padx=10, pady=1)
        tk.Label(f, text=etiqueta, bg=BLANCO, fg=color, font=("Segoe UI", 10, "bold"),
                 width=14, anchor="w").pack(side="left")
        lbl = tk.Label(f, text="—", bg=BLANCO, fg=color, font=("Segoe UI", 10), anchor="w")
        lbl.pack(side="left")
        return lbl

    # --------------------------------------------- selector de caratula
    def _on_caratula(self, inicial=False):
        cara = self.var_caratula.get()
        self.cfg["caratula_seleccionada"] = cara
        guardar_config(self.cfg)
        es_min = (cara == "ministerio_salud")
        try:
            self.btn_datos.config(state=("normal" if es_min else "disabled"))
        except Exception:
            pass
        if es_min and not inicial and not any((self.datos_proyecto or {}).values()):
            self._abrir_datos_proyecto()
        self._validar()

    def _abrir_datos_proyecto(self):
        try:
            dlg = tk.Toplevel(self)
            dlg.title("Datos del proyecto (carátula Ministerio de Salud)")
            dlg.configure(bg=GRIS_BG)
            dlg.transient(self)
            dlg.resizable(False, False)
            dlg.grab_set()
            tk.Label(dlg, text="DATOS DEL PROYECTO", bg=AZUL_ES, fg=BLANCO,
                     font=("Segoe UI", 12, "bold"), padx=14, pady=8).pack(fill="x")
            cont = tk.Frame(dlg, bg=GRIS_BG, padx=18, pady=12)
            cont.pack(fill="both", expand=True)
            tk.Label(cont, text="Estos datos se aplican a TODAS las carátulas del lote.\n"
                                "Los que deje vacíos quedarán en blanco para llenarlos a mano.",
                     bg=GRIS_BG, fg=AZUL_ES, font=("Segoe UI", 8), justify="left").pack(fill="x", pady=(0, 8))
            ents = {}
            for clave, etiqueta in CAMPOS_PROYECTO:
                tk.Label(cont, text=etiqueta, bg=GRIS_BG, fg=AZUL_ES,
                         font=("Segoe UI", 9, "bold"), anchor="w").pack(fill="x")
                e = tk.Entry(cont, font=("Segoe UI", 9))
                e.insert(0, self.datos_proyecto.get(clave, ""))
                e.pack(fill="x", ipady=2, pady=(0, 4))
                ents[clave] = e
            tk.Label(cont, text="Nota: 'Versión' será siempre v1 y 'Registro' será el "
                                "consecutivo de cada material (automático).",
                     bg=GRIS_BG, fg="#6C757D", font=("Segoe UI", 8), justify="left",
                     wraplength=360).pack(fill="x", pady=(4, 8))

            def _guardar():
                for clave, _ in CAMPOS_PROYECTO:
                    self.datos_proyecto[clave] = ents[clave].get().strip()
                self.cfg["datos_proyecto"] = dict(self.datos_proyecto)
                guardar_config(self.cfg)
                dlg.grab_release()
                dlg.destroy()
                self._validar()

            fb = tk.Frame(cont, bg=GRIS_BG)
            fb.pack(fill="x")
            tk.Button(fb, text="💾 Guardar", command=_guardar, bg=ROJO_ES, fg=BLANCO,
                      relief="flat", font=("Segoe UI", 9, "bold"), padx=14, pady=5,
                      cursor="hand2").pack(side="left")
            tk.Button(fb, text="Cancelar", command=lambda: (dlg.grab_release(), dlg.destroy()),
                      bg="#6C757D", fg=BLANCO, relief="flat", font=("Segoe UI", 9),
                      padx=14, pady=5, cursor="hand2").pack(side="left", padx=(8, 0))
            dlg.update_idletasks()
            dlg.geometry(f"+{self.winfo_rootx() + 60}+{self.winfo_rooty() + 40}")
        except Exception as e:
            messagebox.showerror("Datos del proyecto", f"No se pudo abrir: {e}")

    # --------------------------------------------- PANEL MANTENIMIENTO
    def _estado_key(self, k):
        return "✅ Configurada" if (k or "").strip() else "Sin configurar"

    def _abrir_mantenimiento(self):
        if self.procesando:
            messagebox.showinfo("Mantenimiento", "No se puede abrir durante un proceso en curso.")
            return
        try:
            dlg = tk.Toplevel(self)
            dlg.title("Configuracion y mantenimiento")
            dlg.configure(bg=GRIS_BG)
            dlg.transient(self)
            dlg.resizable(False, False)
            dlg.grab_set()
            self.maint = {"dlg": dlg}
            tk.Label(dlg, text="CONFIGURACION Y MANTENIMIENTO", bg=AZUL_ES, fg=BLANCO,
                     font=("Segoe UI", 12, "bold"), padx=14, pady=8).pack(fill="x")
            cont = tk.Frame(dlg, bg=GRIS_BG, padx=18, pady=12)
            cont.pack(fill="both", expand=True)

            tk.Label(cont, text="🔑 API KEY OpenAI", bg=GRIS_BG, fg=ROJO_ES,
                     font=("Segoe UI", 10, "bold"), anchor="w").pack(fill="x")
            # v2.4: SEGURIDAD -> solo estado, sin mostrar ni copiar la clave
            self.maint["lbl_actual"] = tk.Label(
                cont, text="Estado:  " + self._estado_key(self.var_api.get()),
                bg=GRIS_BG, fg=AZUL_ES, font=("Segoe UI", 9, "bold"), anchor="w")
            self.maint["lbl_actual"].pack(fill="x", pady=(2, 4))
            fb = tk.Frame(cont, bg=GRIS_BG)
            fb.pack(fill="x")
            tk.Button(fb, text="🗑 Limpiar", command=self._maint_limpiar_key, bg="#6C757D", fg=BLANCO,
                      relief="flat", font=("Segoe UI", 9), padx=8, cursor="hand2").pack(side="left")

            tk.Label(cont, text="Configurar / cambiar clave:", bg=GRIS_BG, fg=AZUL_ES,
                     font=("Segoe UI", 9, "bold"), anchor="w").pack(fill="x", pady=(8, 0))
            fc = tk.Frame(cont, bg=GRIS_BG)
            fc.pack(fill="x")
            self.maint["ent_nueva"] = tk.Entry(fc, show="*", font=("Segoe UI", 9))
            self.maint["ent_nueva"].pack(side="left", fill="x", expand=True, ipady=3)
            self.maint["btn_ok"] = tk.Button(fc, text="✅ OK", command=self._maint_cambiar_key,
                                             bg=ROJO_ES, fg=BLANCO, relief="flat",
                                             font=("Segoe UI", 9, "bold"), padx=10, cursor="hand2")
            self.maint["btn_ok"].pack(side="left", padx=(6, 0))
            self.maint["status"] = tk.Label(cont, text="", bg=GRIS_BG, font=("Segoe UI", 8), anchor="w")
            self.maint["status"].pack(fill="x")
            tk.Label(cont, text="⚠️ Por seguridad, la clave no se puede ver ni copiar. "
                                "Si necesita otra, use 'Cambiar clave'.",
                     bg=GRIS_BG, fg="#6C757D", font=("Segoe UI", 8), justify="left",
                     wraplength=360, anchor="w").pack(fill="x", pady=(2, 0))

            tk.Frame(cont, bg="#C9CED6", height=1).pack(fill="x", pady=8)
            tk.Label(cont, text="🔄 OPCIONES DE RESET", bg=GRIS_BG, fg=ROJO_ES,
                     font=("Segoe UI", 10, "bold"), anchor="w").pack(fill="x")
            fr = tk.Frame(cont, bg=GRIS_BG)
            fr.pack(fill="x", pady=(2, 0))
            tk.Button(fr, text="🔄 Limpiar cache", command=self._maint_limpiar_cache, bg=AZUL_ES,
                      fg=BLANCO, relief="flat", font=("Segoe UI", 9), padx=8, cursor="hand2").grid(row=0, column=0, padx=2, pady=2)
            tk.Button(fr, text="🗑 Resetear config", command=self._maint_reset_config, bg=AZUL_ES,
                      fg=BLANCO, relief="flat", font=("Segoe UI", 9), padx=8, cursor="hand2").grid(row=0, column=1, padx=2, pady=2)
            tk.Button(fr, text="📁 Abrir carpeta config", command=self._maint_abrir_config, bg=AZUL_ES,
                      fg=BLANCO, relief="flat", font=("Segoe UI", 9), padx=8, cursor="hand2").grid(row=0, column=2, padx=2, pady=2)

            tk.Frame(cont, bg="#C9CED6", height=1).pack(fill="x", pady=8)
            tk.Label(cont, text="ℹ️ Informacion", bg=GRIS_BG, fg=ROJO_ES,
                     font=("Segoe UI", 10, "bold"), anchor="w").pack(fill="x")
            uv = self.cfg.get("api", {}).get("ultima_validacion", "") or "nunca"
            proyecto = Path(self.var_carpeta.get()).name if self.var_carpeta.get() else "—"
            tess = "OK" if TESSERACT_OK else ("no instalado" if TESSERACT_OK is not None else "sin verificar")
            info = (f"Version: {VERSION} - Elaborado por {AUTOR}\n"
                    f"Ultima validacion de API: {uv}\n"
                    f"Tesseract-OCR: {tess}\n"
                    f"Config: {CONFIG_PATH}\n"
                    f"Proyecto actual: {proyecto}")
            tk.Label(cont, text=info, bg=GRIS_BG, fg=AZUL_ES, font=("Segoe UI", 8),
                     justify="left", anchor="w").pack(fill="x", pady=(2, 8))
            tk.Button(cont, text="Cerrar", command=self._maint_cerrar, bg="#6C757D", fg=BLANCO,
                      relief="flat", font=("Segoe UI", 9), padx=14, pady=5, cursor="hand2").pack()

            dlg.protocol("WM_DELETE_WINDOW", self._maint_cerrar)
            dlg.update_idletasks()
            dlg.geometry(f"+{self.winfo_rootx() + 60}+{self.winfo_rooty() + 60}")
        except Exception as e:
            self._log_agregar(f"Error al abrir mantenimiento: {e}")
            messagebox.showerror("Mantenimiento", f"⚠️ No se pudo abrir el panel:\n{e}")

    def _maint_cerrar(self):
        if self.maint and self.maint.get("dlg"):
            try:
                self.maint["dlg"].grab_release()
                self.maint["dlg"].destroy()
            except Exception:
                pass
        self.maint = None

    def _maint_limpiar_key(self):
        if not messagebox.askyesno("Limpiar clave", "¿Borrar la API key guardada?"):
            return
        self.var_api.set("")
        self.cfg.setdefault("api", {})["openai_key_encrypted"] = ""
        guardar_config(self.cfg)
        if self.maint:
            self.maint["lbl_actual"].config(text="Estado:  " + self._estado_key(""))
            self.maint["status"].config(text="Clave borrada.", fg=AMARILLO)
        self._validar()

    def _maint_cambiar_key(self):
        nueva = self.maint["ent_nueva"].get().strip()
        if not nueva.startswith("sk-"):
            self.maint["status"].config(text="❌ Formato invalido: debe empezar con 'sk-'.", fg=ROJO_ES)
            return
        self.maint["btn_ok"].config(state="disabled")
        self.maint["status"].config(text="Probando la nueva clave con OpenAI…", fg=AZUL_ES)
        threading.Thread(target=hilo_test_api, args=(nueva, self.cola, "MAINT_TEST"), daemon=True).start()

    def _maint_limpiar_cache(self):
        borrados = 0
        try:
            base = Path(self.var_carpeta.get() or "")
            objetivos = []
            if base.is_dir():
                for p in base.glob("*.html"):
                    if p.name != "template_caratula.html" and re.match(r"tmp.*\.html$", p.name, re.I):
                        objetivos.append(p)
                for p in base.rglob("*.cmp.tmp"):
                    objetivos.append(p)
            for d in (base, app_dir()):
                pc = d / "__pycache__"
                if pc.is_dir():
                    objetivos.extend(pc.glob("*"))
            for p in objetivos:
                try:
                    p.unlink()
                    borrados += 1
                except Exception:
                    pass
            self.cfg.setdefault("mantenimiento", {})["ultima_limpieza_cache"] = ahora_iso()
            guardar_config(self.cfg)
            if self.maint:
                self.maint["status"].config(text=f"Cache limpiada: {borrados} archivo(s).", fg=VERDE_OK)
        except Exception as e:
            messagebox.showerror("Limpiar cache", error_legible(e))

    def _maint_reset_config(self):
        if not messagebox.askyesno("Resetear configuracion",
                                   "Restaura submitals_config.json a valores por defecto "
                                   "(borra la API key y preferencias).\n\n¿Continuar?"):
            return
        veces = self.cfg.get("mantenimiento", {}).get("veces_reseted", 0) + 1
        self.cfg = _config_base()
        self.cfg["mantenimiento"]["veces_reseted"] = veces
        guardar_config(self.cfg)
        self.var_api.set("")
        self.var_auto.set(True)
        self.var_existente.set(False)
        self.var_solo.set(True)
        self.var_forzar.set(False)
        self.var_log.set(True)
        self.var_caratula.set("clasica")
        self.datos_proyecto = dict(self.cfg["datos_proyecto"])
        if self.maint:
            self.maint["lbl_actual"].config(text="Estado:  " + self._estado_key(""))
            self.maint["status"].config(text=f"Configuracion reseteada (#{veces}).", fg=VERDE_OK)
        self._on_caratula(inicial=True)
        self._validar()

    def _maint_abrir_config(self):
        try:
            os.startfile(str(app_dir()))
        except Exception as e:
            messagebox.showerror("Abrir carpeta", error_legible(e))

    # ------------------------------------------------- valores iniciales
    def _cargar_valores_iniciales(self):
        for c in self.cfg["carpetas_recientes"]:
            if Path(c).is_dir():
                self.var_carpeta.set(c)
                break
        else:
            if Path(DEFAULT_BASE).is_dir():
                self.var_carpeta.set(DEFAULT_BASE)
        if self.var_carpeta.get():
            self._carpeta_cambiada(buscar_json=not self.cfg.get("ultimo_json"))
        uj = self.cfg.get("ultimo_json", "")
        if uj and Path(uj).is_file():
            self.var_json.set(uj)

    # ----------------------------------------------------------- validaciones
    def _carpeta_valida(self, ruta):
        p = Path(ruta) if ruta else None
        return bool(p and p.is_dir() and any((p / m).is_dir() for m in CARPETAS_MADRE))

    def _json_valido(self, ruta):
        try:
            data = json.loads(Path(ruta).read_text(encoding="utf-8"))
            mats = data.get("materiales") if isinstance(data, dict) else data
            return isinstance(mats, list) and len(mats) > 0
        except Exception:
            return False

    def _modo_auto(self):
        return self.var_auto.get()

    def _validar(self):
        ok_c = self._carpeta_valida(self.var_carpeta.get())
        ok_j = bool(self.var_json.get()) and self._json_valido(self.var_json.get())
        auto = self._modo_auto()
        api_ok = bool(self.var_api.get().strip())

        if not self.var_carpeta.get():
            self.lbl_val_carpeta.config(text="Seleccione la carpeta del proyecto.", fg=AMARILLO)
        elif ok_c:
            self.lbl_val_carpeta.config(text="✓ Carpeta valida.", fg=VERDE_OK)
        else:
            self.lbl_val_carpeta.config(text="✗ No contiene ARQUITECTONICOS/MECANICOS/ESTRUCTURALES.", fg=ROJO_ES)

        try:
            self.ent_json.config(state=("disabled" if auto else "readonly"))
            self.btn_json.config(state=("disabled" if auto else "normal"))
        except Exception:
            pass
        if auto:
            self.lbl_json_titulo.config(fg="#9AA0A8")
            self.lbl_val_json.config(text="El JSON se generara automaticamente desde las fichas.", fg=AZUL_ES)
        else:
            self.lbl_json_titulo.config(fg=AZUL_ES)
            if not self.var_json.get():
                self.lbl_val_json.config(text="Seleccione datos_materiales.json.", fg=AMARILLO)
            elif ok_j:
                self.lbl_val_json.config(text="✓ JSON valido.", fg=VERDE_OK)
            else:
                self.lbl_val_json.config(text="✗ JSON invalido o sin 'materiales'.", fg=ROJO_ES)

        if auto and not api_ok:
            self.lbl_api.config(text="⚠ Falta la API Key. Configúrela en '🔧 Configuracion' (arriba).", fg=AMARILLO)
        elif auto and api_ok:
            self.lbl_api.config(text="✓ API Key configurada.", fg=VERDE_OK)
        else:
            self.lbl_api.config(text="Modo JSON existente: no se necesita API Key.", fg="#9AA0A8")

        if auto:
            habilitar = ok_c and api_ok
            hint = "" if habilitar else "Modo automatico: falta carpeta valida o API Key."
        else:
            habilitar = ok_c and ok_j
            hint = "" if habilitar else "Modo existente: falta carpeta valida o JSON valido."
        habilitar = habilitar and not self.procesando
        self.btn_generar.config(state=("normal" if habilitar else "disabled"),
                                bg=(ROJO_ES if habilitar else "#9AA0A8"))
        self.tip_generar.texto = "" if habilitar else hint
        return habilitar

    def _chk_auto(self):
        self.var_existente.set(not self.var_auto.get())
        self._validar()

    def _chk_existente(self):
        self.var_auto.set(not self.var_existente.get())
        self._validar()

    # -------------------------------------------------------------- examinar
    def _examinar_carpeta(self):
        inicial = self.var_carpeta.get() or DEFAULT_BASE
        ruta = filedialog.askdirectory(title="Seleccione la carpeta base del proyecto",
                                       initialdir=inicial if Path(inicial).is_dir() else None)
        if ruta:
            self.var_carpeta.set(os.path.normpath(ruta))
            self._carpeta_cambiada()

    def _carpeta_cambiada(self, buscar_json=True):
        base = self.var_carpeta.get()
        if self._carpeta_valida(base):
            rec = [c for c in self.cfg["carpetas_recientes"] if c != base]
            self.cfg["carpetas_recientes"] = ([base] + rec)[:3]
            self.cbo_carpeta.config(values=self.cfg["carpetas_recientes"])
            guardar_config(self.cfg)
            if buscar_json:
                candidato = Path(base) / NOMBRE_JSON
                if candidato.is_file():
                    self.var_json.set(str(candidato))
        self._validar()

    def _examinar_json(self):
        inicial = self.var_carpeta.get() or None
        ruta = filedialog.askopenfilename(title="Seleccione el archivo JSON de datos",
                                          initialdir=inicial, filetypes=[("Archivos JSON", "*.json")])
        if not ruta:
            return
        ruta = os.path.normpath(ruta)
        if not self._json_valido(ruta):
            messagebox.showerror("JSON invalido", "No es un JSON valido o no contiene 'materiales'.")
        self.var_json.set(ruta)
        self.cfg["ultimo_json"] = ruta
        guardar_config(self.cfg)
        self._validar()

    def _chk_solo(self):
        if self.var_solo.get():
            self.var_forzar.set(False)

    def _chk_forzar(self):
        if self.var_forzar.get():
            self.var_solo.set(False)
            if not messagebox.askyesno("Forzar regeneracion",
                                       "Se SOBREESCRIBIRÁN las carátulas y compilados existentes "
                                       "con los nuevos.\n\n¿Desea continuar?"):
                self.var_forzar.set(False)
                self.var_solo.set(True)

    # ---------------------------------------------------------------- generar
    def _generar(self):
        # v2.4: control de hilo -> evita procesos duplicados
        if self.procesando or (self.thread_activo and self.thread_activo.is_alive()) \
                or self.compilando_disc:
            messagebox.showwarning("En proceso", "El programa ya esta generando. "
                                                  "Espere a que termine.")
            return
        if not self._validar():
            return
        base = Path(self.var_carpeta.get())
        auto = self._modo_auto()
        cara = self.var_caratula.get()

        avisos = []
        tpl_name, logo_rel, _ = CARATULAS.get(cara, CARATULAS["clasica"])
        if not resource_path(tpl_name).exists():
            avisos.append(f"• No se encontro la plantilla '{tpl_name}'.")
        if not resource_path(logo_rel).exists():
            avisos.append(f"• No se encontro el logo '{logo_rel}'.")
        if avisos and not messagebox.askyesno("Advertencia",
                "Se detectaron posibles problemas:\n\n" + "\n".join(avisos) +
                "\n\n¿Desea continuar de todos modos?"):
            return

        ruta_json = self.var_json.get()
        if auto:
            existente = base / NOMBRE_JSON
            if existente.is_file():
                r = messagebox.askyesnocancel("JSON existente",
                    f"Ya existe {NOMBRE_JSON}.\n\nSI = Sobrescribir (releer fichas con ChatGPT)\n"
                    "NO = Usar el existente (no consume la API)\nCANCELAR = No hacer nada")
                if r is None:
                    return
                if r is False:
                    auto = False
                    ruta_json = str(existente)

        self.cfg["opciones"] = {
            "solo_faltantes": self.var_solo.get(),
            "forzar_regeneracion": self.var_forzar.get(),
            "mostrar_log": self.var_log.get(),
            "generar_json_automatico": self.var_auto.get(),
            "usar_json_existente": self.var_existente.get(),
        }
        self.cfg["caratula_seleccionada"] = cara
        self.cfg["datos_proyecto"] = dict(self.datos_proyecto)
        self.cfg["api"]["openai_key_encrypted"] = cifrar_api_key(self.var_api.get().strip())
        guardar_config(self.cfg)

        self.procesando = True
        self.cancelar.clear()
        self.log_lineas.clear()
        self._log_limpiar()
        self.barra["value"] = 0
        self.lbl_progreso.config(text="Verificando configuracion…")
        for lbl in (self.lbl_ok, self.lbl_cmp, self.lbl_skip, self.lbl_err, self.lbl_tiempo):
            lbl.config(text="—")
        self.btn_generar.config(text="⏳ PROCESANDO…", state="disabled", bg=AZUL_ES)
        self.btn_cancelar.config(state="normal")
        self.btn_maint.config(state="disabled")
        self.btn_compilado_disc.config(state="disabled")

        opciones = dict(self.cfg["opciones"])
        modo = "auto" if auto else "existente"
        self.thread_activo = threading.Thread(
            target=hilo_trabajo,
            args=(modo, str(base), ruta_json, self.var_api.get().strip(),
                  opciones, cara, dict(self.datos_proyecto),
                  self.cola, self.cancelar), daemon=True)
        self.thread_activo.start()

    def _cancelar_proceso(self):
        if self.procesando:
            self.cancelar.set()
            self.lbl_progreso.config(text="Cancelando…")
            self.btn_cancelar.config(state="disabled")

    # ------------------------------------------- compilado por disciplina (v2.5.2)
    def _generar_compilados_disciplina(self):
        if self.procesando or self.compilando_disc:
            messagebox.showwarning("En proceso", "Espere a que termine el proceso actual.")
            return
        base = self.var_carpeta.get()
        if not self._carpeta_valida(base):
            messagebox.showwarning("Carpeta", "Seleccione una carpeta de proyecto valida "
                                              "(con ARQUITECTONICOS/MECANICOS/ESTRUCTURALES/ELECTRICOS).")
            return

        disciplinas = []
        if self.var_disc_arq.get():
            disciplinas.append("ARQUITECTONICOS")
        if self.var_disc_estr.get():
            disciplinas.append("ESTRUCTURALES")
        if self.var_disc_mec.get():
            disciplinas.append("MECANICOS")
        if self.var_disc_elec.get():
            disciplinas.append("ELECTRICOS")
        if not disciplinas:
            messagebox.showwarning("Compilado por disciplina", "Seleccione al menos una disciplina.")
            return

        self.compilando_disc = True
        self.btn_compilado_disc.config(state="disabled", text="Generando…")
        threading.Thread(target=hilo_compilados_disciplina,
                         args=(base, disciplinas, self.cola), daemon=True).start()

    def _mostrar_resultados_compilados(self, resultados):
        exitosos = sum(1 for r in resultados if r["resultado"]["exitoso"])
        total = len(resultados)
        mensaje = f"Compilados generados: {exitosos}/{total}\n\n"
        for r in resultados:
            disciplina = r["disciplina"]
            resultado = r["resultado"]
            if resultado["exitoso"]:
                mensaje += (f"✅ {disciplina}:\n"
                           f"   Archivo: {resultado['archivo_generado']}\n"
                           f"   Materiales: {resultado['total_materiales']}\n"
                           f"   Paginas: {resultado['total_paginas']}\n\n")
            else:
                mensaje += f"❌ {disciplina}:\n   Error: {resultado['error']}\n\n"
        messagebox.showinfo("Compilados por disciplina", mensaje.strip())

    # ------------------------------------------------------ cola de eventos
    def _revisar_cola(self):
        try:
            while True:
                ev = self.cola.get_nowait()
                tipo = ev[0]
                if tipo == "LOG":
                    self._log_agregar(ev[1])
                elif tipo == "WARN":
                    self._log_agregar("AVISO: " + ev[1])
                elif tipo == "FASE":
                    self.barra["value"] = 0
                    self.lbl_progreso.config(text=ev[1])
                    self._log_agregar("== " + ev[1] + " ==")
                elif tipo == "JSON_PROG":
                    _, n, total, cons = ev
                    pct = int(n * 100 / max(total, 1))
                    self.barra["value"] = pct
                    self.lbl_progreso.config(
                        text=f"Leyendo fichas y extrayendo (con traduccion si aplica): {cons}  ({n}/{total}, {pct}%)")
                elif tipo == "PDF_GENERATED":
                    _, n, total, stats = ev
                    pct = int(n * 100 / max(total, 1))
                    self.barra["value"] = pct
                    self.lbl_progreso.config(text=f"Generando caratulas: {pct}%  ({n}/{total})")
                    self._pintar_stats(stats)
                elif tipo == "COMP_PROG":
                    _, n, total = ev
                    pct = int(n * 100 / max(total, 1))
                    self.barra["value"] = pct
                    self.lbl_progreso.config(text=f"Compilando documentos: {pct}%  ({n}/{total})")
                elif tipo == "DISC_FASE":
                    self.lbl_progreso.config(text=ev[1])
                    self._log_agregar("== " + ev[1] + " ==")
                elif tipo == "DISC_PROG":
                    _, n, total, disciplina = ev
                    self.lbl_progreso.config(
                        text=f"Compilando por disciplina: {disciplina}  ({n}/{total})")
                elif tipo == "DISC_COMPLETE":
                    _, resultados, _seg = ev
                    self.compilando_disc = False
                    self.btn_compilado_disc.config(state="normal", text="📦 Generar Compilados")
                    exitosos = sum(1 for r in resultados if r["resultado"]["exitoso"])
                    self.lbl_progreso.config(
                        text=f"✅ Compilados por disciplina: {exitosos}/{len(resultados)}.")
                    self._mostrar_resultados_compilados(resultados)
                elif tipo == "MAINT_TEST":
                    _, ok, msg, key = ev
                    if self.maint:
                        self.maint["btn_ok"].config(state="normal")
                    if ok:
                        self.var_api.set(key)
                        self.cfg.setdefault("api", {})["openai_key_encrypted"] = cifrar_api_key(key)
                        self.cfg["api"]["ultima_validacion"] = ahora_iso()
                        guardar_config(self.cfg)
                        self._maint_cerrar()
                        messagebox.showinfo("API Key", "✅ API key actualizada correctamente.")
                        self._validar()
                    elif self.maint:
                        self.maint["status"].config(
                            text="❌ " + msg + "  (platform.openai.com/api-keys)", fg=ROJO_ES)
                elif tipo == "CANCELLED":
                    self.procesando = False
                    self._boton_normal()
                    self.lbl_progreso.config(text="Proceso cancelado.")
                    self._validar()
                elif tipo == "ERROR":
                    self.procesando = False
                    self._boton_normal()
                    self.lbl_progreso.config(text="Proceso detenido por un error.")
                    messagebox.showerror("Error", ev[1])
                    self._validar()
                elif tipo == "COMPLETE":
                    _, stats, seg = ev
                    self.procesando = False
                    self._boton_normal()
                    self.barra["value"] = 100
                    self._pintar_stats(stats)
                    self.lbl_cmp.config(text=f"{stats.get('compilados', 0)} PDFs")
                    h, r = divmod(int(seg), 3600)
                    m, s = divmod(r, 60)
                    self.lbl_tiempo.config(text=f"{h:02d}:{m:02d}:{s:02d}")
                    self.lbl_progreso.config(
                        text=f"✅ ¡Listo! {stats.get('ok', 0)} caratulas y {stats.get('compilados', 0)} compilados.")
                    self._validar()
                    if stats.get("fallidos"):
                        messagebox.showwarning("Completado con errores",
                            f"Se generaron {stats['ok']} caratulas, pero hubo "
                            f"{stats['fallidos']} error(es).\nRevise el log completo.")
                    if stats.get("motor_fallback"):
                        messagebox.showwarning(
                            "Playwright/Chromium no disponible",
                            f"{stats['motor_fallback']} caratula(s) se generaron con un motor "
                            "de respaldo (pdfkit/weasyprint) porque Playwright/Chromium fallo. "
                            "Ese motor renderiza peor el diseno de la caratula (puede verse mal "
                            "visualmente).\n\n"
                            "Para corregirlo, cierre el programa y ejecute:\n"
                            "  python -m playwright install chromium\n\n"
                            "y vuelva a generar las caratulas afectadas.")
                    if self.var_log.get():
                        self._ver_log(silencioso=True)
                elif tipo == "UPDATE":
                    self._on_update(ev[1])
        except queue.Empty:
            pass
        self.after(100, self._revisar_cola)

    # ---- v2.6.7: auto-actualizacion --------------------------------------
    def _chk_update(self):
        try:
            info = auto_updater.verificar_actualizacion(
                logf=lambda m: self.cola.put(("LOG", m)))
            if info.get("disponible"):
                self.cola.put(("UPDATE", info))
        except Exception:
            pass

    def _on_update(self, info):
        if not messagebox.askyesno(
                "✨ Actualizacion disponible",
                f"Version {info.get('version_remota', '?')}\n\n"
                f"{info.get('changelog', '')}\n\n¿Actualizar ahora?"):
            return
        if hasattr(self, "lbl_progreso"):
            self.lbl_progreso.config(text="Actualizando…")
        threading.Thread(target=self._aplicar_update, args=(info,), daemon=True).start()

    def _aplicar_update(self, info):
        try:
            ok, msg, reiniciar = auto_updater.aplicar_actualizacion(
                info, progreso=lambda p, t: self.cola.put(("LOG", f"[{p}%] {t}")))
            self.cola.put(("LOG", msg))
            if info.get("requiere_exe"):
                _, mx = auto_updater.descargar_exe_y_preparar_swap(
                    info, logf=lambda m: self.cola.put(("LOG", m)))
                self.cola.put(("LOG", mx))
            if ok and reiniciar:
                self.cola.put(("LOG", "Reiniciando la aplicacion…"))
                self.after(1500, auto_updater.reiniciar_app)
        except Exception as e:
            self.cola.put(("LOG", f"Error en actualizacion: {e}"))

    def _pintar_stats(self, st):
        saltados = st.get("existentes", 0) + st.get("vacias", 0) + st.get("incompletas", 0)
        self.lbl_ok.config(text=f"{st.get('ok', 0)} PDFs")
        self.lbl_cmp.config(text=f"{st.get('compilados', 0)} PDFs")
        self.lbl_skip.config(text=f"{saltados}  (ya existen: {st.get('existentes', 0)}, "
                                  f"vacias: {st.get('vacias', 0)}, incompletas: {st.get('incompletas', 0)})")
        self.lbl_err.config(text=str(st.get("fallidos", 0)))

    def _boton_normal(self):
        self.btn_generar.config(text="🚀 GENERAR")
        self.btn_cancelar.config(state="disabled")
        self.btn_maint.config(state="normal")
        self.btn_compilado_disc.config(state="normal")

    def _log_agregar(self, linea):
        self.log_lineas.append(linea)
        self.txt_log.config(state="normal")
        self.txt_log.insert("end", linea + "\n")
        self.txt_log.see("end")
        self.txt_log.config(state="disabled")

    def _log_limpiar(self):
        self.txt_log.config(state="normal")
        self.txt_log.delete("1.0", "end")
        self.txt_log.config(state="disabled")

    def _ver_log(self, silencioso=False):
        reporte = Path(self.var_carpeta.get()) / NOMBRE_REPORTE
        if reporte.is_file():
            try:
                os.startfile(str(reporte))
                return
            except Exception as e:
                if not silencioso:
                    messagebox.showerror("Error", error_legible(e))
                return
        if not silencioso:
            messagebox.showinfo("Sin reporte", "Todavia no existe un reporte.")

    def _abrir_resultados(self):
        base = Path(self.var_carpeta.get() or "")
        destino = None
        for m in CARPETAS_MADRE:
            if (base / m).is_dir():
                destino = base / m
                break
        destino = destino or (base if base.is_dir() else None)
        if destino:
            try:
                os.startfile(str(destino))
            except Exception as e:
                messagebox.showerror("Error", error_legible(e))
        else:
            messagebox.showinfo("Sin carpeta", "Seleccione primero una carpeta valida.")

    def _generar_excel_ahora(self):
        carpeta = Path(self.var_carpeta.get() or "")
        if not carpeta.is_dir():
            messagebox.showwarning("Carpeta invalida",
                                  "Seleccione una carpeta valida primero.")
            return

        json_path = carpeta / NOMBRE_JSON
        if not json_path.exists():
            messagebox.showwarning("JSON no encontrado",
                                  f"No existe {NOMBRE_JSON} en la carpeta.\n"
                                  f"Primero debe generar o cargar los datos.")
            return

        try:
            resultado = generar_excel_materiales(json_path, carpeta)
            if resultado.get("exitoso"):
                total = resultado.get("total_materiales", 0)
                messagebox.showinfo("Excel generado",
                                   f"✅ Excel actualizado exitosamente\n\n"
                                   f"Archivo: Guía Materiales.xlsx\n"
                                   f"Materiales: {total}\n\n"
                                   f"Ubicación: {carpeta}")
            else:
                error = resultado.get("error", "Error desconocido")
                messagebox.showerror("Error al generar Excel", error)
        except Exception as e:
            messagebox.showerror("Error", error_legible(e))

    # ---------------------------------------- borrar todo y empezar de 0 (v2.6.2)
    def _borrar_caratulas_compilados_click(self):
        if self.procesando or self.compilando_disc:
            messagebox.showwarning("En proceso", "Espere a que termine el proceso actual.")
            return

        carpeta = Path(self.var_carpeta.get() or "")
        if not carpeta.is_dir():
            messagebox.showwarning("Carpeta invalida", "Seleccione una carpeta valida primero.")
            return

        json_path = carpeta / NOMBRE_JSON
        hay_json = json_path.exists()

        total_preview = 1 if hay_json else 0
        for madre in CARPETAS_MADRE:
            cm = carpeta / madre
            if not cm.is_dir():
                continue
            total_preview += len(list(cm.glob("CMP SUBMITTAL *.pdf")))
            for sub in cm.iterdir():
                if not sub.is_dir():
                    continue
                for p in sub.glob("*.pdf"):
                    up = p.name.upper()
                    if up.startswith("CARATULA") or up.endswith("-CMP.PDF"):
                        total_preview += 1

        if total_preview == 0:
            messagebox.showinfo("Nada que borrar",
                               "No se encontraron caratulas, compilados ni "
                               f"{NOMBRE_JSON} en esta carpeta.")
            return

        if not messagebox.askyesno(
                "Confirmar borrado total",
                f"Se eliminaran {total_preview} archivo(s) bajo '{carpeta.name}':\n\n"
                "  • Todas las CARATULA*.pdf\n"
                "  • Todos los compilados individuales (*-CMP.pdf)\n"
                "  • Todos los compilados por disciplina (CMP SUBMITTAL *.pdf)\n"
                f"  • {NOMBRE_JSON} (los datos de marca/descripcion/normativa "
                "ya extraidos con ChatGPT se pierden)\n\n"
                "NO se tocan las fichas técnicas originales (los documentos "
                "fuente de cada carpeta se conservan).\n\n"
                "⚠ Esta acción NO se puede deshacer: los archivos se borran "
                "permanentemente (no van a una carpeta de respaldo). La próxima "
                "generación tendrá que volver a leer todas las fichas con "
                "ChatGPT (consume la API de OpenAI de nuevo).\n\n"
                "¿Desea continuar y empezar de cero?"):
            return

        try:
            resultado = borrar_caratulas_y_compilados(carpeta, q=None)
        except Exception as e:
            messagebox.showerror("Error", error_legible(e))
            return

        total_borrados = (resultado["caratulas"] + resultado["compilados_individuales"]
                          + resultado["compilados_disciplina"]
                          + (1 if resultado["json_eliminado"] else 0))
        msg = [f"✅ Se eliminaron {total_borrados} archivo(s):",
              f"  • Carátulas: {resultado['caratulas']}",
              f"  • Compilados individuales (CMP): {resultado['compilados_individuales']}",
              f"  • Compilados por disciplina: {resultado['compilados_disciplina']}",
              f"  • {NOMBRE_JSON}: {'eliminado' if resultado['json_eliminado'] else 'no existia'}"]
        if resultado["errores"]:
            msg.append(f"\n⚠ {len(resultado['errores'])} archivo(s) no se pudieron eliminar "
                      "(pueden estar abiertos en un visor de PDF/Excel). Ciérrelos e intente de nuevo.")
        msg.append("\n➡ Ya puede presionar 🚀 GENERAR (modo automático) para leer las "
                   "fichas y regenerar todo desde cero.")
        messagebox.showinfo("Borrado completado", "\n".join(msg))

        # El JSON ya no existe: se vuelve al modo automatico para la proxima
        # corrida (el modo "usar JSON existente" quedaria apuntando a un
        # archivo borrado).
        if resultado["json_eliminado"]:
            self.var_auto.set(True)
            self.var_existente.set(False)
            self.var_json.set("")
            self._validar()

    # ------------------------------------------ exportar version final (v2.6.7)
    def _exportar_version_final_click(self):
        if self.procesando or self.compilando_disc:
            messagebox.showwarning("En proceso", "Espere a que termine el proceso actual.")
            return

        carpeta = Path(self.var_carpeta.get() or "")
        if not carpeta.is_dir():
            messagebox.showwarning("Carpeta invalida", "Seleccione una carpeta valida primero.")
            return

        salida = carpeta / NOMBRE_CARPETA_FINAL
        if salida.exists():
            if not messagebox.askyesno(
                    "Sobrescribir versión final",
                    f"Ya existe la carpeta '{NOMBRE_CARPETA_FINAL}'.\n\n"
                    "Se eliminará por completo y se volverá a generar desde cero "
                    "con los archivos actuales.\n\n¿Desea continuar?"):
                return

        try:
            resultado = exportar_version_final(carpeta, q=None)
        except Exception as e:
            messagebox.showerror("Error", error_legible(e))
            return

        msg = [f"✅ Versión final exportada en:\n{resultado['carpeta_salida']}\n",
              f"  • CMP individuales copiados: {resultado['cmp_copiados']}",
              f"  • Compilados por disciplina copiados: {resultado['compilados_disc_copiados']}",
              f"  • Excel guía: {'copiado' if resultado['excel_copiado'] else 'no encontrado'}"]
        if resultado["disciplinas_sin_archivos"]:
            msg.append(f"\n⚠ Sin archivos para: {', '.join(resultado['disciplinas_sin_archivos'])}")
        if resultado["errores"]:
            msg.append(f"\n⚠ {len(resultado['errores'])} archivo(s) no se pudieron copiar "
                       "(pueden estar abiertos en otro programa).")
        msg.append("\n¿Desea abrir la carpeta ahora?")

        if messagebox.askyesno("Exportación completada", "\n".join(msg)):
            try:
                os.startfile(resultado["carpeta_salida"])
            except Exception as e:
                messagebox.showerror("Error", error_legible(e))

    # -------------------------------------------------- duplicados (v2.6)
    def _detectar_duplicados_click(self):
        if self.procesando or self.compilando_disc:
            messagebox.showwarning("En proceso", "Espere a que termine el proceso actual.")
            return

        carpeta = Path(self.var_carpeta.get() or "")
        if not carpeta.is_dir():
            messagebox.showwarning("Carpeta invalida", "Seleccione una carpeta valida primero.")
            return

        json_path = carpeta / NOMBRE_JSON
        if not json_path.exists():
            messagebox.showwarning("JSON no encontrado",
                                  f"No existe {NOMBRE_JSON} en la carpeta.\n"
                                  f"Primero debe generar o cargar los datos.")
            return

        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception as e:
            messagebox.showerror("Error leyendo JSON", error_legible(e))
            return
        materiales = data.get("materiales") if isinstance(data, dict) else data
        if not isinstance(materiales, list) or not materiales:
            messagebox.showwarning("JSON vacio", "El JSON no contiene materiales.")
            return

        grupos = detectar_duplicados(materiales)
        if not grupos:
            messagebox.showinfo("Sin duplicados",
                               "No se detectaron materiales duplicados en este submittal.")
            return

        self._mostrar_dialogo_duplicados(carpeta, json_path, materiales, grupos)

    def _mostrar_dialogo_duplicados(self, base, json_path, materiales, grupos):
        win = tk.Toplevel(self)
        win.title(f"Materiales duplicados detectados ({len(grupos)} grupo(s))")
        win.configure(bg=GRIS_BG)
        win.geometry("760x560")
        win.transient(self)
        win.grab_set()

        tk.Label(win, bg=GRIS_BG, fg=AZUL_ES, font=("Segoe UI", 11, "bold"),
                anchor="w", justify="left",
                text=("Se detectaron materiales que parecen ser la MISMA ficha tecnica "
                      "registrada en mas de una carpeta.\n"
                      "Elija cual consecutivo CONSERVAR en cada grupo; los demas se "
                      f"moveran a '{CARPETA_CUARENTENA}' (no se borran) y los consecutivos "
                      "posteriores de esa disciplina se reordenaran.")
               ).pack(fill="x", padx=14, pady=(14, 6))

        cont = tk.Frame(win, bg=GRIS_BG)
        cont.pack(fill="both", expand=True, padx=14, pady=6)
        canvas = tk.Canvas(cont, bg=BLANCO, highlightthickness=0)
        scrollbar = ttk.Scrollbar(cont, orient="vertical", command=canvas.yview)
        frame = tk.Frame(canvas, bg=BLANCO)
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        variables = []
        for gi, grupo in enumerate(grupos, 1):
            items = grupo["materiales"]
            fg_box = tk.LabelFrame(
                frame, bg=BLANCO, fg=AZUL_ES, font=("Segoe UI", 9, "bold"),
                text=f"Grupo {gi} — similitud {int(grupo['similitud_min'] * 100)}%")
            fg_box.pack(fill="x", padx=8, pady=6)
            var = tk.StringVar(value=items[0]["consecutivo"])
            variables.append(var)
            for it in items:
                desc = (it.get("descripcion", "") or "")[:90]
                texto = (f"CONSERVAR  {it['consecutivo']}  ({it['categoria']})  —  "
                        f"{it.get('nombre', '')}  |  Marca: {it.get('marca', '')}  |  {desc}")
                tk.Radiobutton(fg_box, text=texto, variable=var, value=it["consecutivo"],
                              bg=BLANCO, anchor="w", justify="left", wraplength=680,
                              font=("Segoe UI", 8)).pack(fill="x", anchor="w", padx=6, pady=2)

        def _confirmar():
            a_eliminar = []
            detalle = []
            for grupo, var in zip(grupos, variables):
                conservar = var.get()
                for it in grupo["materiales"]:
                    if it["consecutivo"] != conservar:
                        a_eliminar.append(it["consecutivo"])
                        detalle.append(f"  - {it['consecutivo']} ({it.get('nombre', '')}) "
                                      f"→ se elimina; se conserva {conservar}")
            if not a_eliminar:
                messagebox.showinfo("Nada que hacer", "No hay materiales para eliminar.")
                return
            if not messagebox.askyesno(
                    "Confirmar correccion",
                    "Se realizaran los siguientes cambios:\n\n" + "\n".join(detalle) +
                    f"\n\nLas carpetas eliminadas se MOVERAN a "
                    f"'{CARPETA_CUARENTENA}' (no se borran del disco).\n"
                    "Los consecutivos posteriores de cada disciplina afectada se "
                    "reordenaran y sus caratulas/compilados quedaran pendientes de "
                    "regenerar.\n\n¿Desea continuar?"):
                return
            self._ejecutar_resolucion_duplicados(win, base, json_path, materiales, a_eliminar)

        fbtn = tk.Frame(win, bg=GRIS_BG)
        fbtn.pack(fill="x", padx=14, pady=(6, 14))
        tk.Button(fbtn, text="✅ Confirmar y corregir", command=_confirmar,
                 bg=ROJO_ES, fg=BLANCO, font=("Segoe UI", 10, "bold"),
                 relief="flat", padx=14, pady=8, cursor="hand2").pack(side="left")
        tk.Button(fbtn, text="Cancelar", command=win.destroy,
                 bg="#6C757D", fg=BLANCO, font=("Segoe UI", 10),
                 relief="flat", padx=14, pady=8, cursor="hand2").pack(side="left", padx=(8, 0))

    def _ejecutar_resolucion_duplicados(self, win, base, json_path, materiales, a_eliminar):
        try:
            resultado = resolver_duplicados(base, materiales, a_eliminar, q=None)
        except Exception as e:
            messagebox.showerror("Error", error_legible(e))
            return

        nuevos_materiales = resultado["materiales"]
        docs_totales = sum(len(m.get("documentos_encontrados", []) or []) for m in nuevos_materiales)
        res = resumen_materiales(nuevos_materiales, docs_totales)
        try:
            json_path.write_text(
                json.dumps({"resumen": res, "materiales": nuevos_materiales},
                          ensure_ascii=False, indent=2),
                encoding="utf-8")
        except Exception as e:
            messagebox.showerror("Error guardando JSON", error_legible(e))
            return

        try:
            generar_excel_materiales(json_path, base)
        except Exception:
            pass

        win.destroy()

        eliminados = resultado["eliminados"]
        renumerados = resultado["renumerados"]
        compilados = resultado["compilados_obsoletos"]

        msg = [f"✅ Se eliminaron {len(eliminados)} material(es) duplicado(s):"]
        for e in eliminados:
            msg.append(f"  • {e['consecutivo']} ({e['nombre']}) → movido a "
                      f"'{Path(e['ruta_cuarentena']).parent.name}/{Path(e['ruta_cuarentena']).name}'")
        if renumerados:
            msg.append("\n🔄 Consecutivos reordenados:")
            for r in renumerados:
                msg.append(f"  • {r['consecutivo_viejo']} → {r['consecutivo_nuevo']} ({r['nombre']})")
        if compilados:
            msg.append(f"\n⚠ Se eliminaron {len(compilados)} compilado(s) por disciplina "
                      "obsoleto(s); genere uno nuevo con '📦 Generar Compilados'.")
        msg.append(f"\n📊 {NOMBRE_JSON} y Guía Materiales.xlsx actualizados.")
        msg.append("\n➡ Presione 🚀 GENERAR (modo 'usar JSON existente') para regenerar las "
                   "caratulas de los materiales renumerados (no vuelve a consultar la API).")
        messagebox.showinfo("Duplicados corregidos", "\n".join(msg))

        # Prepara el modo "usar JSON existente" con el JSON ya actualizado, listo
        # para que un solo clic en GENERAR regenere justo lo que quedo pendiente.
        self.var_auto.set(False)
        self.var_existente.set(True)
        self.var_json.set(str(json_path))
        self.cfg["ultimo_json"] = str(json_path)
        guardar_config(self.cfg)
        self._validar()

    def _nuevo_proceso(self):
        if self.procesando:
            if not messagebox.askyesno("Proceso en curso", "Hay un proceso en curso. ¿Cancelarlo?"):
                return
            self.cancelar.set()
            self.procesando = False
            self._boton_normal()
        self.barra["value"] = 0
        self.lbl_progreso.config(text="Listo para comenzar.")
        for lbl in (self.lbl_ok, self.lbl_cmp, self.lbl_skip, self.lbl_err, self.lbl_tiempo):
            lbl.config(text="—")
        self._log_limpiar()
        self._validar()

    def _cerrar(self):
        if self.procesando:
            if not messagebox.askyesno("Proceso en curso", "Hay un proceso en curso. ¿Cerrar igual?"):
                return
            self.cancelar.set()
        guardar_config(self.cfg)
        self.destroy()


# ============================================================================
# ARRANQUE (instancia unica + bootstrap)
# ============================================================================
def _arrancar():
    if not instancia_unica():
        r = tk.Tk()
        r.withdraw()
        messagebox.showwarning("Ya esta abierto",
                               "El Generador de Submittals ya se esta ejecutando.\n"
                               "Use la ventana que ya esta abierta.")
        r.destroy()
        return

    splash = tk.Tk()
    splash.title(f"Generador de Submittals v{VERSION}")
    _aplicar_icono_ventana(splash)
    splash.configure(bg=GRIS_BG)
    splash.geometry("480x300")
    tk.Label(splash, text="PREPARANDO EL PROGRAMA", bg=AZUL_ES, fg=BLANCO,
             font=("Segoe UI", 13, "bold"), pady=10).pack(fill="x")
    tk.Label(splash, text="Verificando e instalando dependencias…", bg=GRIS_BG,
             fg=AZUL_ES, font=("Segoe UI", 9)).pack(pady=(8, 2))
    txt = tk.Text(splash, height=10, font=("Consolas", 8), bg="#1E1E1E", fg="#D4D4D4",
                  state="disabled", relief="flat")
    txt.pack(fill="both", expand=True, padx=12, pady=6)
    pb = ttk.Progressbar(splash, mode="indeterminate")
    pb.pack(fill="x", padx=12, pady=(0, 10))
    pb.start(12)

    q2 = queue.Queue()
    estado = {"err": None}

    def logf(m):
        q2.put(m)

    def worker():
        try:
            bootstrap(logf)
        except Exception as e:
            estado["err"] = e
        q2.put(None)

    threading.Thread(target=worker, daemon=True).start()

    def poll():
        try:
            while True:
                m = q2.get_nowait()
                if m is None:
                    splash.quit()
                    return
                txt.config(state="normal")
                txt.insert("end", m + "\n")
                txt.see("end")
                txt.config(state="disabled")
        except queue.Empty:
            pass
        splash.after(120, poll)

    splash.after(120, poll)
    splash.mainloop()
    try:
        splash.destroy()
    except Exception:
        pass

    if estado["err"]:
        r = tk.Tk()
        r.withdraw()
        messagebox.showerror("No se pudo iniciar", str(estado["err"]))
        r.destroy()
        return
    SubmitalsGUI().mainloop()


if __name__ == "__main__":
    multiprocessing.freeze_support()   # v2.4: evita relanzamientos en .exe
    _arrancar()
